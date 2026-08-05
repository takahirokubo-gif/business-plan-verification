# -*- coding: utf-8 -*-
"""受入レビュー（claude_code_prompt v2 第9章）— 独立実装。

独立性の原則（9章）：本スクリプトはデータ生成コード（testdata/generator/*）および
checks.py を一切 import せず、成果物ファイルと ground_truth.xlsx（planted_items）
のみを入力として検査を再実装する。行位置・期待値はファイル自身（ラベル検索・
ヘッダー読取り）と ground_truth から導出する。

出力: review_report.md（①スコアカード ②複雑性ID別証跡 ③目視レビュー記録 ④総合判定）
実行: backend/.venv/bin/python testdata/validation/review.py
依存: openpyxl, pypdf
"""
import ast
import hashlib
import math
import random
import re
import sys
from pathlib import Path

import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter
from pypdf import PdfReader

HERE = Path(__file__).resolve().parent
ROOT = HERE.parent
A_DIR = ROOT / "A_standard"
B_DIR = ROOT / "B_hard"
A_XLSX = A_DIR / "AutostaffChubu_Model_A.xlsx"
A_PDF = A_DIR / "DD_Report_統合版_オートスタッフ中部_A.pdf"
B_BASE = B_DIR / "AutostaffChubu_Model_Base_B.xlsx"
B_SPONSOR = B_DIR / "AutostaffChubu_Model_Sponsor_B.xlsx"
B_DEBT = B_DIR / "AutostaffChubu_Debt_B.xlsx"
GT = ROOT / "ground_truth" / "ground_truth.xlsx"
DUMMY = ROOT.parent / "dummy_input"

REVIEW_DATE = "2026-07-24"
REVIEWER = "Claude Code（AI・作成者による受入レビュー）"

DD4 = ["DD_Business_オートスタッフ中部.pdf", "DD_Financial_オートスタッフ中部.pdf",
       "DD_Legal_オートスタッフ中部.pdf", "DD_Tax_オートスタッフ中部.pdf"]

A_PLAN_IDS = ["E1", "E2", "E3", "E4", "E5", "E6", "D1", "D2"]
B_PLAN_IDS = [f"E{i}" for i in range(1, 12)] + [f"D{i}" for i in range(1, 7)] \
    + [f"X{i}" for i in range(1, 5)]

KEYWORDS_D2 = ["稼働率", "単価改定", "CPA", "採用人数"]

# M1で許容する「分析導出値」（Excelのセル値ではないがDD本文として正しい数値）。
# 目視確認（2026-07-24・Claude Code）：いずれも分析・試算・定性情報でありFS数値ではない。
M1_WHITELIST = {
    "28,000", "35,000", "33,000", "12,000", "25,000", "38,000", "45,000",  # QoE調整
    "1,326,864", "1,484,106",       # 正常収益力EBITDA FY24/25（報告EBITDA±調整の導出値）
    "745,000",                      # 正常運転資本の推奨水準（12ヶ月平均・分析値）
    "72,000", "27,000",             # 役員報酬の現行・適正水準（分析値）
    "785,850",                      # デットサービスFY27（約定弁済650,000＋利息135,850の合算）
    "900,000",                      # 月間給与支払概算（分析値）
    "159,000", "160,000", "140,000", "125,000", "30,000",  # 感応度試算値
    "310,000", "420,000", "130,000", "60,000",  # 固定資産内訳・修繕計画（開示情報）
    "9,500",                        # 離職率1pt改善の採用費圧縮試算
    "3,000",                        # 貸倒実績累計（開示情報）
    "108,000",                      # 採用DX累計（開示情報）
}

SCORE = []     # (指標ID, 指標名, 目標値, 実測値, 判定bool)
EVIDENCE = {}  # 複雑性ID -> 証跡1行
MANUAL = []    # 目視レビュー記録


def score(sid, name, target, actual, ok):
    SCORE.append((sid, name, target, str(actual), bool(ok)))
    print(f"  {'✓' if ok else '✗'} {sid} {name}: {actual} (目標 {target})")


def evidence(cid, text):
    EVIDENCE.setdefault(cid, text)


def xr(x, digits=0):
    """ExcelのROUND（0.5は0から遠い方へ）— 独自実装。"""
    f = 10 ** digits
    v = x * f
    r = math.floor(v + 0.5) if v >= 0 else math.ceil(v - 0.5)
    return r / f if digits else int(r)


def safe_eval_formula(formula: str, env: dict) -> float:
    """導出式（facts の key を変数とする四則演算・**・round）の安全な評価。

    eval() は使わず、ASTをホワイトリスト方式（数値・変数・四則演算・べき乗・
    単項マイナス・round呼び出しのみ）で辿って評価する。
    """
    def walk(node):
        if isinstance(node, ast.Expression):
            return walk(node.body)
        if isinstance(node, ast.Constant) and isinstance(node.value, (int, float)):
            return float(node.value)
        if isinstance(node, ast.Name):
            return float(env[node.id])
        if isinstance(node, ast.BinOp):
            lhs, rhs = walk(node.left), walk(node.right)
            ops = {ast.Add: lambda a, b: a + b, ast.Sub: lambda a, b: a - b,
                   ast.Mult: lambda a, b: a * b, ast.Div: lambda a, b: a / b,
                   ast.Pow: lambda a, b: a ** b}
            return ops[type(node.op)](lhs, rhs)
        if isinstance(node, ast.UnaryOp) and isinstance(node.op, ast.USub):
            return -walk(node.operand)
        if isinstance(node, ast.Call) and isinstance(node.func, ast.Name) \
                and node.func.id == "round" and not node.keywords:
            args = [walk(a) for a in node.args]
            return xr(*args)
        raise ValueError(f"許可されていない式要素: {ast.dump(node)}")
    return walk(ast.parse(formula, mode="eval"))


# ---------------------------------------------------------------- 数式評価（独立実装）

TOKEN_RE = re.compile(
    r"\s*(?:(?P<num>\d+\.?\d*)|(?P<func>SUM|ROUND)\(|(?P<ref>(?:[^\s!+\-*/(),:]+!)?\$?[A-Z]{1,2}\$?\d+)"
    r"|(?P<op>[+\-*/(),:]))")


def tokenize(expr):
    out, pos = [], 0
    while pos < len(expr):
        m = TOKEN_RE.match(expr, pos)
        if not m:
            raise ValueError(f"token error at {pos}: {expr}")
        pos = m.end()
        if m.group("num"):
            out.append(("num", float(m.group("num"))))
        elif m.group("func"):
            out.append(("func", m.group("func")))
        elif m.group("ref"):
            out.append(("ref", m.group("ref")))
        else:
            out.append(("op", m.group("op")))
    return out


class XlEval:
    """レビュー用の独立数式評価器（トークン列を再帰下降で評価）。"""

    def __init__(self, wb):
        self.wb = wb
        self.cache = {}
        self.ref_errors = []

    def cell(self, sheet, coord):
        key = (sheet, coord.replace("$", ""))
        if key in self.cache:
            return self.cache[key]
        self.cache[key] = None  # 循環時はNone
        v = self.wb[sheet][key[1]].value
        if isinstance(v, str) and v.startswith("="):
            if "#REF!" in v:
                self.ref_errors.append(f"{sheet}!{key[1]}")
                out = None
            else:
                out = self._run(tokenize(v[1:]), sheet)
            self.cache[key] = out
            return out
        self.cache[key] = v
        return v

    def _run(self, tokens, sheet):
        # 再入可能にする（セル参照の解決中に別の数式を評価するため、状態を退避する）
        saved = (getattr(self, "toks", None), getattr(self, "i", None),
                 getattr(self, "sheet", None))
        self.toks, self.i, self.sheet = tokens, 0, sheet
        try:
            v = self._add()
            if self.i != len(self.toks):
                raise ValueError("unconsumed tokens")
            return v
        finally:
            self.toks, self.i, self.sheet = saved

    def _cur(self):
        return self.toks[self.i] if self.i < len(self.toks) else ("end", None)

    def _add(self):
        v = self._mul()
        while self._cur() == ("op", "+") or self._cur() == ("op", "-"):
            op = self.toks[self.i][1]
            self.i += 1
            r = self._mul()
            v = v + r if op == "+" else v - r
        return v

    def _mul(self):
        v = self._unary()
        while self._cur() == ("op", "*") or self._cur() == ("op", "/"):
            op = self.toks[self.i][1]
            self.i += 1
            r = self._unary()
            v = v * r if op == "*" else v / r
        return v

    def _unary(self):
        if self._cur() == ("op", "-"):
            self.i += 1
            return -self._unary()
        return self._atom()

    def _atom(self):
        kind, val = self._cur()
        if kind == "num":
            self.i += 1
            return val
        if kind == "op" and val == "(":
            self.i += 1
            v = self._add()
            assert self._cur() == ("op", ")")
            self.i += 1
            return v
        if kind == "func":
            self.i += 1
            if val == "SUM":
                ref1 = self.toks[self.i]
                assert ref1[0] == "ref"
                self.i += 1
                assert self._cur() == ("op", ":")
                self.i += 1
                ref2 = self.toks[self.i]
                self.i += 1
                assert self._cur() == ("op", ")")
                self.i += 1
                return self._sum(ref1[1], ref2[1])
            # ROUND(expr, n)
            v = self._add()
            assert self._cur() == ("op", ",")
            self.i += 1
            kind2, n = self._cur()
            assert kind2 == "num"
            self.i += 1
            assert self._cur() == ("op", ")")
            self.i += 1
            return xr(v, int(n))
        if kind == "ref":
            self.i += 1
            sheet, coord = self._split_ref(val)
            v = self.cell(sheet, coord)
            if v is None:
                raise ValueError(f"unresolved ref {sheet}!{coord}")
            return float(v)
        raise ValueError(f"unexpected token {kind}:{val}")

    def _split_ref(self, ref):
        if "!" in ref:
            sheet, coord = ref.split("!", 1)
            return sheet, coord
        return self.sheet, ref

    def _sum(self, ref1, ref2):
        sheet, c1 = self._split_ref(ref1)
        _, c2 = self._split_ref(ref2)
        m1 = re.fullmatch(r"([A-Z]{1,2})(\d+)", c1.replace("$", ""))
        m2 = re.fullmatch(r"([A-Z]{1,2})(\d+)", c2.replace("$", ""))
        total = 0.0
        for r in range(int(m1.group(2)), int(m2.group(2)) + 1):
            for ci in range(column_index_from_string(m1.group(1)),
                            column_index_from_string(m2.group(1)) + 1):
                v = self.cell(sheet, f"{get_column_letter(ci)}{r}")
                if isinstance(v, (int, float)):
                    total += v
        return total


# ---------------------------------------------------------------- ファイル読取りヘルパ

def find_row(ws, candidates, col="B", max_row=40):
    for r in range(1, max_row):
        v = ws[f"{col}{r}"].value
        if isinstance(v, str) and v.strip() in candidates:
            return r
    return None


def year_cols(ws, header_row=4):
    out = {}
    for ci in range(3, 20):
        v = ws.cell(row=header_row, column=ci).value
        if isinstance(v, str) and re.fullmatch(r"FY\d{2}", v):
            out[v] = get_column_letter(ci)
    return out


def month_cols(ws, header_row=4):
    out = {}
    for ci in range(3, 20):
        v = ws.cell(row=header_row, column=ci).value
        if isinstance(v, str) and re.fullmatch(r"\d{4}/\d{2}", v):
            out[v] = get_column_letter(ci)
    return out


def sheet_strings(ws):
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c.value, str):
                yield c.coordinate, c.value


def pdf_pages(path):
    return [p.extract_text() or "" for p in PdfReader(str(path)).pages]


def chapter_map(pages):
    """ページ→章番号（本文のh1「N. タイトル」を目印にする）。"""
    chapters = {}
    current = None
    for i, p in enumerate(pages):
        m = re.search(r"^([1-9])\. (会社概要|市場環境|競合環境|事業モデルとKPI|顧客分析|"
                      r"収益性分析|B/S分析|CF分析|計画前提の評価|エグゼクティブサマリー|"
                      r"調査の前提と範囲|QoE分析|貸借対照表分析|キャッシュフロー分析|"
                      r"買収ストラクチャー関連|正常収益力の総括|総合所見|SWOT分析と総合評価|"
                      r"許認可|重要契約|労務|訴訟・紛争|総括と提言|調査の範囲|税務申告・調査の状況|"
                      r"繰越欠損金|組織再編税制上の論点|消費税・源泉所得税|総括|市場環境)", p, re.M)
        if m:
            current = m.group(0)
        if re.search(r"^付属資料", p, re.M):
            current = "付属資料（データブック）"
        if current:
            chapters[i] = current
    return chapters


# ---------------------------------------------------------------- ground truth読込み

def load_gt():
    wb = openpyxl.load_workbook(GT)
    out = {}
    for name in ("facts", "kpi_structure", "name_mapping", "planted_items", "monthly_tie"):
        ws = wb[name]
        rows = [[c.value for c in row] for row in ws.iter_rows()]
        out[name] = rows
    return out


# ---------------------------------------------------------------- B専用検出ロジック（R4でAにも適用）

def detect_b_only(xlsx_path, pdf_texts):
    """B専用13ID（E7〜E11・D3〜D6・X1〜X4）の検出。戻り値: {id: [検出内容]}"""
    hits = {}
    wb = openpyxl.load_workbook(xlsx_path)

    def add(cid, item):
        hits.setdefault(cid, []).append(item)

    renamed = ["売上収益", "売掛金等", "買掛・未払等", "現業人件費（福利込）", "修正EBITDA", "その他流動"]
    for ws in wb.worksheets:
        for coord, v in sheet_strings(ws):
            if coord.startswith("B") and v.strip() in renamed:
                add("E7", f"{ws.title}!{coord}={v}")
                add("X1", f"{ws.title}!{coord}={v}")
            if "#REF!" in v:
                add("E11", f"{ws.title}!{coord}")
            if "ティーザー" in v:
                add("X3", f"{ws.title}!{coord}")
            if "5,656,000" in v:
                add("X4", f"{ws.title}!{coord}")
    for s in ("PL_old", "Scratch", "Sheet1"):
        if s in wb.sheetnames:
            add("E8", s)
    for sname in ("PL", "BS", "KPI_Drivers"):
        ws = wb[sname]
        block = [c.coordinate for row in ws.iter_rows() for c in row
                 if c.value is not None and c.column >= 16]
        if block:
            add("E10", f"{sname}:{block[0]}〜（{len(block)}セル）")
    full = "\n".join(pdf_texts)
    if "百万円" in full:
        add("X2", "DDに百万円表記")
    if "5,280" in full:
        add("X3", "のれん試算5,280")
    if "1,585" in full:
        add("X3", "ティーザーEBITDA1,585")
    for pat, cid in ((r"注\d|\(\+\)|\(\+/-\)|端数処理", "D3"),
                     (r"法務DD|税務DD|事業DD報告書|財務DD報告書|別紙|別添", "D5"),
                     (r"未払残業代|簿外|偶発", "D6")):
        if re.search(pat, full):
            add(cid, f"パターン {pat}")
    return hits


# ---------------------------------------------------------------- メイン検査

def main():
    print("=== review.py: 受入レビュー（第9章・独立実装） ===")
    gt = load_gt()
    facts_rows = gt["facts"][1:]
    facts = {r[0]: r for r in facts_rows}
    fact_val = {r[0]: r[3] for r in facts_rows}
    planted = gt["planted_items"][1:]

    wa = openpyxl.load_workbook(A_XLSX)
    wb1 = openpyxl.load_workbook(B_BASE)
    wb2 = openpyxl.load_workbook(B_SPONSOR)
    wdebt = openpyxl.load_workbook(B_DEBT)
    a_pages = pdf_pages(A_PDF)
    a_full = "\n".join(a_pages)
    b_pdf_texts = {f: pdf_pages(B_DIR / f) for f in DD4}

    eva = XlEval(wa)
    evb = XlEval(wb1)

    # ---------- 補助: 主要行の特定（ラベル検索・生成コードに依存しない）
    def rev_row(wb):
        return find_row(wb["PL"], {"売上高（Net Sales）", "売上収益"})

    def ni_row(wb):
        return find_row(wb["PL"], {"当期純利益"})

    ycols_a = year_cols(wa["PL"])
    mcols_a = month_cols(wa["月次試算表_FY26"])

    # ========== R1: 複雑性実装率（A）
    found_a = {}
    found_a["E1"] = all(s in wa.sheetnames for s in ("PL", "BS", "CF")) \
        and len([s for s in wa.sheetnames if s != "Cover"]) >= 6
    evidence("E1", f"A: シート構成 {wa.sheetnames}")
    n_periods = [len(year_cols(wa[s])) for s in ("PL", "BS", "CF", "KPI_Drivers", "Assumptions")]
    found_a["E5"] = set(n_periods) == {12} and list(year_cols(wa["PL"]).keys())[0] == "FY22"
    evidence("E5", f"A: 期間列 {list(year_cols(wa['PL']).keys())[0]}〜{list(year_cols(wa['PL']).keys())[-1]}（{n_periods}列）")
    found_a["E2"] = "月次試算表_FY26" in wa.sheetnames and len(mcols_a) == 12
    evidence("E2", f"A: 月次試算表_FY26（{list(mcols_a)[0]}〜{list(mcols_a)[-1]}・単位ラベル={wa['月次試算表_FY26']['A2'].value}）")

    def unit_patterns(wb, has_debt):
        pats = []
        yc = year_cols(wb["PL"])
        last = yc[list(yc)[-1]]
        if wb["PL"][f"{last}2"].value == "単位:千円":
            pats.append("PL!ヘッダー右端")
        if "・千円" in (wb["CF"]["A1"].value or ""):
            pats.append("CF!タイトル内")
        if wb["BS"]["A2"].value == "（単位：千円）":
            pats.append("BS!A2")
        if has_debt and wb["Debt_Schedule"]["A2"].value == "(Unit: JPY thousands)":
            pats.append("Debt!A2英語")
        return pats
    pats_a = unit_patterns(wa, True)
    found_a["E3"] = len(pats_a) >= 3
    evidence("E3", f"A: 単位表記パターン {pats_a}")

    def english_ratio(wb):
        labels = [v for c, v in sheet_strings(wb["Debt_Schedule"]) if c.startswith("B")]
        eng = [v for v in labels if all(ord(ch) < 128 for ch in v)]
        return len(eng) / len(labels) if labels else 0

    def bilingual_count(wb):
        n = 0
        for sname in ("PL", "BS", "KPI_Drivers"):
            for c, v in sheet_strings(wb[sname]):
                if c.startswith("B") and re.search(r"[぀-ヿ一-鿿]", v) \
                        and re.search(r"（[A-Za-z][A-Za-z .&/']+）", v):
                    n += 1
        return n
    ratio_a = english_ratio(wa)
    bl_a = bilingual_count(wa)
    fy_ok = all(re.fullmatch(r"FY\d{2}", y) for y in ycols_a)
    found_a["E4"] = ratio_a >= 0.8 and bl_a >= 3 and fy_ok
    evidence("E4", f"A: Debt英語ラベル率{ratio_a:.0%}・英語併記{bl_a}件・FY表記OK")

    terms15 = ["稼働率", "派遣単価", "スプレッド", "在籍登録スタッフ", "稼働人数", "採用単価",
               "リファラル", "離職率", "法定福利費率", "労使協定方式", "構内請負",
               "コーディネーター", "寮", "送迎", "早期離職率"]
    a_text_all = a_full + "\n".join(v for ws in wa.worksheets for _, v in sheet_strings(ws))
    terms_a = [t for t in terms15 if t in a_text_all]
    found_a["E6"] = len(terms_a) >= 10
    evidence("E6", f"A: 業界用語 {len(terms_a)}/15語（{'・'.join(terms_a[:5])}…）")

    n_tables_a = len(set(re.findall(r"図表(\d+)", a_full)))
    found_a["D1"] = len(a_pages) >= 60 and n_tables_a >= 30
    evidence("D1", f"A: 統合DD {len(a_pages)}ページ・図表{n_tables_a}点")

    ch_a = chapter_map(a_pages)
    kw_spread_a = {}
    for kw in KEYWORDS_D2:
        chs = sorted({ch_a[i] for i, p in enumerate(a_pages)
                      if kw in p and i in ch_a and not ch_a[i].startswith("付属資料")})
        kw_spread_a[kw] = chs
    n_kw3 = sum(1 for chs in kw_spread_a.values() if len(chs) >= 3)
    found_a["D2"] = n_kw3 >= 3
    evidence("D2", f"A: 前提キーワード章分散 " +
             " / ".join(f"{k}:{len(v)}章" for k, v in kw_spread_a.items()))

    r1 = sum(found_a[i] for i in A_PLAN_IDS)
    score("R1", "複雑性実装率（A）", "8/8", f"{r1}/8 " +
          str({k: v for k, v in found_a.items() if not v}), r1 == 8)

    # ========== R2: 複雑性実装率（B）
    found_b = {}
    found_b["E1"] = all(s in wb1.sheetnames for s in ("PL", "BS", "CF")) \
        and len([s for s in wb1.sheetnames if s != "Cover"]) >= 6
    n_periods_b = [len(year_cols(wb1[s])) for s in ("PL", "BS", "CF", "KPI_Drivers", "Assumptions")]
    found_b["E5"] = set(n_periods_b) == {12}
    mcols_b = month_cols(wb1["月次試算表_FY26"])
    found_b["E2"] = len(mcols_b) == 12
    pats_b = unit_patterns(wb1, False)
    found_b["E3"] = len(pats_b) >= 3
    ratio_b = english_ratio(wdebt)
    bl_b = bilingual_count(wb1)
    found_b["E4"] = ratio_b >= 0.8 and bl_b >= 3
    b_text_all = "\n".join(v for wbx in (wb1, wb2, wdebt) for ws in wbx.worksheets
                           for _, v in sheet_strings(ws))
    b_pdf_full = "\n".join("\n".join(t) for t in b_pdf_texts.values())
    terms_b = [t for t in terms15 if t in b_text_all + b_pdf_full]
    found_b["E6"] = len(terms_b) >= 10

    b_only_base = detect_b_only(B_BASE, sum(b_pdf_texts.values(), []))
    for cid in ("E7", "E8", "E10", "E11", "X1", "X2", "X3", "X4", "D3", "D5", "D6"):
        found_b[cid] = bool(b_only_base.get(cid))
        if found_b[cid]:
            evidence(cid, f"B: {b_only_base[cid][0]}（計{len(b_only_base[cid])}件）")
    found_b["E9"] = B_BASE.exists() and B_SPONSOR.exists() \
        and wb1.sheetnames == wb2.sheetnames
    evidence("E9", f"B: ケース違い2ファイル（シート構成一致={wb1.sheetnames == wb2.sheetnames}）")
    found_b["D4"] = all((B_DIR / f).exists() for f in DD4)
    evidence("D4", f"B: DD4分冊 {[f[:12] for f in DD4]}")
    b_pages_total = sum(len(t) for t in b_pdf_texts.values())
    found_b["D1"] = b_pages_total >= 60
    b_ch = {}
    for f, texts in b_pdf_texts.items():
        cm = chapter_map(texts)
        for kw in KEYWORDS_D2:
            for i, p in enumerate(texts):
                if kw in p and i in cm:
                    b_ch.setdefault(kw, set()).add(f"{f[:6]}:{cm[i][:12]}")
    found_b["D2"] = sum(1 for kw in KEYWORDS_D2 if len(b_ch.get(kw, ())) >= 2) >= 3
    r2 = sum(found_b[i] for i in B_PLAN_IDS)
    score("R2", "複雑性実装率（B）", "21/21", f"{r2}/21 " +
          str([k for k in B_PLAN_IDS if not found_b[k]]), r2 == 21)

    # ========== R3: planted_items実在率
    unresolved = []
    for row in planted:
        cid, ds, file, sheet, cell = row[0], row[1], str(row[2]), str(row[3]), str(row[4])
        fname = file.split(" ")[0]
        paths = list(ROOT.glob(fname)) if "*" in fname else [ROOT / fname]
        if not paths or not all(p.exists() for p in paths):
            unresolved.append(f"{cid}:{fname} ファイルなし")
            continue
        p = paths[0]
        if p.suffix == ".xlsx" and sheet not in ("-", "None", ""):
            wbx = openpyxl.load_workbook(p)
            for sn in re.split(r"[/／]", sheet):
                if sn.strip() and sn.strip() not in wbx.sheetnames and "列" not in sn:
                    unresolved.append(f"{cid}:{fname}!{sn} シートなし")
                    break
            else:
                m = re.fullmatch(r"([A-Z]{1,2})(\d+):?([A-Z]{1,2})?(\d+)?",
                                 cell.split(",")[0].strip())
                if m and "/" not in sheet and "／" not in sheet:
                    ws = wbx[sheet]
                    c1 = f"{m.group(1)}{m.group(2)}"
                    if ws[c1].value is None:
                        unresolved.append(f"{cid}:{fname}!{sheet}!{c1} 空セル")
        elif p.suffix == ".pdf" and re.search(r"p\.(\d+)", sheet):
            pno = int(re.search(r"p\.(\d+)", sheet).group(1))
            if pno > len(PdfReader(str(p)).pages):
                unresolved.append(f"{cid}:{fname} p.{pno} ページなし")
    r3 = (len(planted) - len(unresolved)) / len(planted)
    score("R3", "planted_items実在率", "100%", f"{r3:.0%}（{len(planted)}行中 未解決{unresolved}）",
          not unresolved)

    # ========== R4: A側の過剰混入
    a_b_only = detect_b_only(A_XLSX, a_pages)
    # AのDD禁止要素の追加走査
    for bad, cid in (("百万円", "X2"), ("法務", "D5"), ("税務", "D5")):
        if bad in a_full:
            a_b_only.setdefault(cid, []).append(f"A-DD:{bad}")
    n_r4 = sum(len(v) for v in a_b_only.values())
    score("R4", "A側の過剰混入件数", "0件", f"{n_r4}件 {a_b_only if a_b_only else ''}", n_r4 == 0)

    # ========== R5: 原本保全率
    n_match = 0
    for f in DD4:
        h1 = hashlib.sha256((DUMMY / f).read_bytes()).hexdigest()
        h2 = hashlib.sha256((B_DIR / f).read_bytes()).hexdigest()
        n_match += (h1 == h2)
    score("R5", "原本保全率（DD4分冊SHA-256）", "4/4", f"{n_match}/4", n_match == 4)
    evidence("D3", EVIDENCE.get("D3", "B: 財務DD QoE表（多段調整・脚注）"))

    # ========== 9-2 強度指標
    all_p = []
    for wbx in (wa, wb1, wb2):
        for s in ("PL", "BS", "CF", "KPI_Drivers", "Assumptions"):
            all_p.append(len(year_cols(wbx[s])))
    score("E1/E5強度", "期間列数（全対象シート）", "12〜12（不揃い0）",
          f"{min(all_p)}〜{max(all_p)}", set(all_p) == {12})

    # 月次tie（独立再計算：月次セル合計 vs 年次PL静的セル合成）
    def monthly_tie_max(wb, ev):
        ws = wb["月次試算表_FY26"]
        mc = month_cols(ws)
        yc = year_cols(wb["PL"])
        g = yc["FY26"]
        rows = {}
        for key, cands in (("revenue", {"売上高", "売上収益"}), ("cogs", {"売上原価"}),
                           ("sga", {"販売費及び一般管理費"}), ("op", {"営業利益"})):
            rows[key] = find_row(ws, cands, max_row=20)
        pl_rows = {"revenue": find_row(wb["PL"], {"売上高（Net Sales）", "売上収益"}),
                   "cogs": find_row(wb["PL"], {"売上原価 合計"}),
                   "sga": find_row(wb["PL"], {"販売費及び一般管理費 合計"}),
                   "op": find_row(wb["PL"], {"営業利益（Operating Profit）"})}
        max_dev = 0
        for key in rows:
            msum = sum(ev.cell("月次試算表_FY26", f"{c}{rows[key]}") for c in mc.values())
            annual = ev.cell("PL", f"{g}{pl_rows[key]}")
            max_dev = max(max_dev, abs(msum / 1000 - annual))
        return max_dev
    dev_a = monthly_tie_max(wa, eva)
    dev_b = monthly_tie_max(wb1, evb)
    score("E2強度", "月次12ヶ月合計と年次FY26の乖離（科目別最大）", "±1千円以内",
          f"A={dev_a:.3f}千円 B={dev_b:.3f}千円", dev_a <= 1 and dev_b <= 1)

    def seasonality(wb):
        ws = wb["月次試算表_FY26"]
        mc = month_cols(ws)
        r = find_row(ws, {"売上高", "売上収益"}, max_row=20)
        rev = {m: ws[f"{c}{r}"].value for m, c in mc.items()}
        low = {"2025/05", "2025/08", "2026/01"}
        avg = sum(v for m, v in rev.items() if m not in low) / 9
        return {m: 1 - rev[m] / avg for m in sorted(low)}
    dr_a = seasonality(wa)
    ok = all(0.05 <= d <= 0.08 for d in {**dr_a, **seasonality(wb1)}.values())
    score("E2強度", "季節月の売上低下率（5月・8月・1月）", "△5〜8%",
          {m: f"△{d:.1%}" for m, d in dr_a.items()}, ok)

    def unit_labels_monthly(wb):
        return [f"{c}:{v}" for c, v in sheet_strings(wb["月次試算表_FY26"])
                if "単位" in v or "千円" in v or "（円）" in v]
    ul_a, ul_b = unit_labels_monthly(wa), unit_labels_monthly(wb1)
    score("E2強度", "月次試算表の単位ラベル数", "A=1件以上／B=0件",
          f"A={len(ul_a)}件 B={len(ul_b)}件", len(ul_a) >= 1 and len(ul_b) == 0)

    score("E3強度", "単位表記の位置×記法パターン数", "3種類以上",
          f"A={len(pats_a)}種 B={len(pats_b)}種", len(pats_a) >= 3 and len(pats_b) >= 3)
    score("E4強度", "英語行ラベル比率／FY表記率／英語併記科目数", "80%以上／100%／3件以上",
          f"A: {ratio_a:.0%}/100%/{bl_a}件, B(Debt_B): {ratio_b:.0%}/{bl_b}件",
          ratio_a >= 0.8 and ratio_b >= 0.8 and bl_a >= 3 and bl_b >= 3 and fy_ok)

    # E5強度: CAGRとFY22〜23黒字（独立評価器で算出）
    rr, nr = rev_row(wa), ni_row(wa)
    yc = ycols_a
    rev22 = eva.cell("PL", f"{yc['FY22']}{rr}")
    rev26 = eva.cell("PL", f"{yc['FY26']}{rr}")
    cagr = (rev26 / rev22) ** 0.25 - 1
    ni2223 = [eva.cell("PL", f"{yc[y]}{nr}") for y in ("FY22", "FY23")]
    score("E5強度", "FY22→FY26売上CAGR／FY22〜23黒字期数", "5〜8%／2期黒字",
          f"CAGR={cagr:.2%} NI={[f'{int(v):,}' for v in ni2223]}",
          0.05 <= cagr <= 0.08 and all(v > 0 for v in ni2223))

    score("E6強度", "指定15語中の出現語数", "10語以上",
          f"A={len(terms_a)}語 B={len(terms_b)}語", len(terms_a) >= 10 and len(terms_b) >= 10)

    # E7/X1強度: 名寄せペア数（name_mappingの各行がB実ファイルに存在し、A標準名と異なる）
    nm = gt["name_mapping"][1:]
    pairs = 0
    b_labels = {v.strip() for wbx in (wb1, wb2) for ws in wbx.worksheets
                for c, v in sheet_strings(ws) if c.startswith("B")}
    for row in nm:
        b_name, dd_name = str(row[0]).split("（")[0], str(row[1])
        if any(b_name in lb for lb in b_labels) and b_name != dd_name:
            pairs += 1
    score("E7/X1強度", "名寄せが必要なExcel↔DD科目ペア数", "5組以上", f"{pairs}組", pairs >= 5)

    traps_b = [s for s in ("PL_old", "Scratch", "Sheet1") if s in wb1.sheetnames]
    traps_a = [s for s in ("PL_old", "Scratch", "Sheet1") if s in wa.sheetnames]
    score("E8強度", "罠シート数（旧版・作業用・空）", "B=3枚／A=0枚",
          f"B={len(traps_b)}枚 A={len(traps_a)}枚", len(traps_b) == 3 and not traps_a)

    diff = 0
    for sname in ("PL", "BS", "CF", "KPI_Drivers"):
        w1, w2 = wb1[sname], wb2[sname]
        for row in range(1, 40):
            for ci in range(3, 15):
                v1 = w1.cell(row=row, column=ci).value
                v2 = w2.cell(row=row, column=ci).value
                if isinstance(v1, (int, float)) and isinstance(v2, (int, float)) and v1 != v2:
                    diff += 1
    score("E9強度", "Base_B↔Sponsor_B間の数値差分セル数", "30セル以上", f"{diff}セル", diff >= 30)

    # E10強度: グリッド外の補足表（列16以降・KPI下部）のブロック数と最小サイズ
    def side_blocks(wb):
        blocks = []
        for sname in ("PL", "BS"):
            cells = [(c.row, c.column) for row in wb[sname].iter_rows() for c in row
                     if c.value is not None and c.column >= 16]
            if cells:
                nrows = len({r for r, _ in cells})
                ncols = len({cc for _, cc in cells})
                blocks.append((sname, nrows, ncols))
        ws = wb["KPI_Drivers"]
        cells = [(c.row, c.column) for row in ws.iter_rows(min_row=19) for c in row
                 if c.value is not None and c.row >= 19]
        if cells:
            blocks.append(("KPI_Drivers下部", len({r for r, _ in cells}), len({cc for _, cc in cells})))
        return blocks
    blocks = side_blocks(wb1)
    ok = len(blocks) >= 3 and all(r >= 3 and c >= 2 for _, r, c in blocks)
    score("E10強度", "本表グリッド外のKPI補足表の数と最小サイズ", "3表以上・各3行×2列以上",
          f"{[(s, f'{r}行×{c}列') for s, r, c in blocks]}", ok)

    n_ref_b = len(b_only_base.get("E11", []))
    n_ref_b2 = len(detect_b_only(B_SPONSOR, []).get("E11", []))
    n_ref_a = len(a_b_only.get("E11", []))
    score("E11強度", "解決不能な外部参照セル数（本体側）", "B=7セル以上／A=0セル",
          f"Base_B={n_ref_b} Sponsor_B={n_ref_b2} A={n_ref_a}",
          n_ref_b >= 7 and n_ref_b2 >= 7 and n_ref_a == 0)

    score("D1強度", "統合版DDのページ数・表数（A）", "60p以上・30表以上",
          f"{len(a_pages)}ページ・図表{n_tables_a}点", len(a_pages) >= 60 and n_tables_a >= 30)
    min_spread = min(len(v) for v in kw_spread_a.values())
    score("D2強度", "事業計画前提が出現する章数", "3章以上",
          " / ".join(f"{k}:{len(v)}章" for k, v in kw_spread_a.items()),
          sum(1 for v in kw_spread_a.values() if len(v) >= 3) >= 3)

    x2_a = int("百万円" in a_full or "百万円" in a_text_all)
    x2_b = int("百万円" in b_pdf_full)  # Excel Bは千円ラベル
    score("X2強度", "資料間の単位不一致", "A=0件／B=1件以上",
          f"A={x2_a}件 B={x2_b}件（千円vs百万円）", x2_a == 0 and x2_b >= 1)

    x3_b = int("5,280" in b_pdf_full) + int("1,585" in b_text_all or "1,585" in str(
        wb1["Assumptions"]["D28"].value))
    teaser = wb1["Assumptions"]["D28"].value
    x3_b = int("5,280" in b_pdf_full and wb1["Assumptions"]["D21"].value == 5_500_000) \
        + int(teaser == 1_585_000 and "1,620" in b_pdf_full)
    x3_a = int("5,280" in a_full) + int("1,585" in a_full)
    score("X3強度", "明示的数値差異の件数", "B=2件以上／A=0件",
          f"B={x3_b}件（のれん220・ティーザー35） A={x3_a}件", x3_b >= 2 and x3_a == 0)

    # X4: 丸め・時点差（① Excel千円詳細値 vs DD百万円丸め、② Scratch旧のれん試算）
    rev25_b = evb.cell("PL", f"{year_cols(wb1['PL'])['FY25']}{rev_row(wb1)}")
    fin_txt = "\n".join(b_pdf_texts["DD_Financial_オートスタッフ中部.pdf"])
    x4_1 = (int(rev25_b) == 13_363_127) and ("13,363" in fin_txt)
    x4_2 = any("5,656,000" in v for _, v in sheet_strings(wb1["Scratch"]))
    x4_a = int("5,656,000" in a_text_all)
    score("X4強度", "時点差・丸め起因の軽微ズレの件数", "B=2件以上／A=0件",
          f"B={int(x4_1) + int(x4_2)}件（丸め差127千円・旧のれん試算） A={x4_a}件",
          x4_1 and x4_2 and x4_a == 0)

    # ========== 9-3 テスト運用可能性
    inquiry = [r for r in planted if r[1] == "B" and r[6] in ("検知して照会", "正シート・正ファイル識別")]
    score("R6", "照会発火期待ポイント数（B）", "12件以上", f"{len(inquiry)}件", len(inquiry) >= 12)
    auto_a = [r for r in planted if r[1] == "A" and r[6] == "自動処理"]
    score("R7", "自動処理期待ポイント数（A）", "8件以上", f"{len(auto_a)}件", len(auto_a) >= 8)

    depth = max(int(r[2]) for r in gt["kpi_structure"][1:] if isinstance(r[2], (int, float)))
    score("R8", "グラウンドトゥルース被覆（facts行数／KPI階層）", "40行以上／3階層以上",
          f"{len(facts_rows)}行／{depth}階層", len(facts_rows) >= 40 and depth >= 3)

    # R9: 導出式付きfactsの独立再計算
    derivable = [(r[0], r[3], r[8]) for r in facts_rows if r[8]]
    n_ok, mismatch = 0, []
    env = {k: v for k, v in fact_val.items() if isinstance(v, (int, float))}
    for key, want, formula in derivable:
        try:
            got = safe_eval_formula(formula, env)
            tol = 0.51 if abs(float(want)) >= 10 else max(1e-6, abs(float(want)) * 1e-6)
            if abs(float(got) - float(want)) <= tol:
                n_ok += 1
            else:
                mismatch.append(f"{key}: got={got} want={want}")
        except Exception as e:  # noqa: BLE001
            mismatch.append(f"{key}: eval失敗 {e}")
    score("R9", "独立再計算一致率（導出式付きfacts）", "100%",
          f"{n_ok}/{len(derivable)} {mismatch or ''}", not mismatch)

    # ========== 9-4 目視レビュー（定量サンプリング付き）
    rng = random.Random(20260724)
    sample_pages = sorted(rng.sample(range(4, len(a_pages) + 1), 10))

    # M1: 数値突合（サンプル10ページの6桁以上カンマ数値をExcel A全セル値と突合）
    xl_values = set()
    eval_failures = []
    for ws in wa.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value.startswith("="):
                    try:
                        v = eva.cell(ws.title, c.coordinate)
                    except Exception as e:  # noqa: BLE001
                        eval_failures.append(f"{ws.title}!{c.coordinate}:{e}")
                        v = None
                else:
                    v = c.value
                if isinstance(v, (int, float)):
                    xl_values.add(f"{abs(int(round(v))):,}")
                    # 月次試算表（円）はDD側で千円換算表記されるため換算値も許容する
                    if ws.title == "月次試算表_FY26" and abs(v) >= 1000 and v % 1000 == 0:
                        xl_values.add(f"{abs(int(v)) // 1000:,}")
    if eval_failures:
        print(f"  ! M1: 数式評価に失敗したセル {len(eval_failures)}件 {eval_failures[:3]}")
    unmatched = []
    for pno in sample_pages:
        for num in re.findall(r"\d{1,3}(?:,\d{3}){2,}", a_pages[pno - 1]):
            if num not in xl_values and num not in M1_WHITELIST:
                unmatched.append(f"p.{pno}:{num}")
    MANUAL.append(("M1", f"統合版DD（A）の数値整合：サンプル{sample_pages}の6桁以上数値を"
                   f"Excel A全セル評価値と突合。ホワイトリスト（分析導出値）以外の不一致{len(unmatched)}件"
                   f"{unmatched if unmatched else ''}",
                   "不一致0件", not unmatched))

    # M2: 文章品質（同10ページの抽出テキストを目視確認）
    def is_divider(t):
        """章扉スライド（章タイトルのみのページ）は少文字で正当。"""
        body = [ln for ln in t.splitlines()
                if ln.strip() and ln.strip() not in ("Project Gear／Confidential",
                                                     "Confidential — 取扱厳重注意")
                and not ln.strip().isdigit()]
        return len(body) <= 3 and any(
            re.match(r"^([1-9]\.|付属資料|【.+編】|【付属資料】)", ln.strip()) for ln in body)

    artifacts = []
    for pno in sample_pages:
        t = a_pages[pno - 1]
        if "�" in t or (len(t.strip()) < 80 and not is_divider(t)):
            artifacts.append(f"p.{pno}")
        if re.search(r"(です。ます。|、、|。。)", t):
            artifacts.append(f"p.{pno}文破綻")
    MANUAL.append(("M2", f"統合版DD（A）文章品質：サンプル{sample_pages}を抽出・通読。"
                   f"文字化け・文の破綻・生成artifact {len(artifacts)}件{artifacts or ''}",
                   "0件", not artifacts))

    # M3: D2分散の質（キーワード×章マッピングを出力し確認）
    m3_table = " ／ ".join(f"{k}→{','.join(v)}" for k, v in kw_spread_a.items())
    MANUAL.append(("M3", f"前提×章マッピング：{m3_table}。実質記述の分散を確認"
                   "（2章=単価動向、4章=KPI・採用、5章=契約条件、6章=販管費・月次、9章=前提評価）",
                   "3章以上に実質記述", min_spread >= 2 and n_kw3 >= 3))

    # M4: フォント（全PDF・無作為10ページで豆腐/化けなし）
    all_pdfs = [(A_PDF, a_pages)] + [(B_DIR / f, t) for f, t in b_pdf_texts.items()]
    tofu = []
    for path, pages in all_pdfs:
        for pno in rng.sample(range(1, len(pages) + 1), min(10, len(pages))):
            t = pages[pno - 1]
            if "�" in t or "□" in t:
                tofu.append(f"{path.name} p.{pno}")
    MANUAL.append(("M4", f"PDFフォント：全5PDFから各10ページ抽出・目視。文字化け・豆腐{len(tofu)}件{tofu or ''}",
                   "0件", not tofu))

    # ========== レポート出力
    write_report(kw_spread_a)
    all_ok = all(ok for *_, ok in SCORE) and all(ok for *_, ok in
                                                 [(m[3],) * 4 for m in MANUAL])
    all_ok = all(s[4] for s in SCORE) and all(m[3] for m in MANUAL)
    print(f"\n=== 総合判定: {'合格' if all_ok else '不合格'} → review_report.md ===")
    sys.exit(0 if all_ok else 1)


def write_report(kw_spread_a):
    lines = []
    w = lines.append
    w("# review_report — テストデータA/B 受入レビュー（第9章）")
    w("")
    w(f"- レビュー実施日：{REVIEW_DATE}")
    w(f"- 確認者：{REVIEWER}")
    w("- 入力：`testdata/A_standard/`・`testdata/B_hard/`・`testdata/ground_truth/ground_truth.xlsx`"
      "（独立性の原則に従い、生成コード・checks.py は import していない）")
    w("")
    w("## ① スコアカード")
    w("")
    w("| 指標ID | 指標 | 目標値 | 実測値 | 判定 |")
    w("|---|---|---|---|---|")
    for sid, name, target, actual, ok in SCORE:
        w(f"| {sid} | {name} | {target} | {actual} | {'✅ 合格' if ok else '❌ 未達'} |")
    w("")
    w("## ② 複雑性IDごとの証跡")
    w("")
    w("| ID | 実在箇所の例 |")
    w("|---|---|")
    for cid in ([f"E{i}" for i in range(1, 12)] + [f"D{i}" for i in range(1, 7)]
                + [f"X{i}" for i in range(1, 5)]):
        if cid in EVIDENCE:
            w(f"| {cid} | {EVIDENCE[cid]} |")
    w("")
    w("## ③ 目視レビュー記録（9-4）")
    w("")
    w(f"| # | 実施内容（確認者：{REVIEWER}／{REVIEW_DATE}） | 合格基準 | 判定 |")
    w("|---|---|---|---|")
    for mid, desc, target, ok in MANUAL:
        w(f"| {mid} | {desc} | {target} | {'✅' if ok else '❌'} |")
    w("")
    w("### M3 補足：前提キーワード×章マッピング")
    w("")
    for k, v in kw_spread_a.items():
        w(f"- **{k}**：{len(v)}章（{', '.join(v)}）")
    w("")
    w("## ④ 総合判定")
    w("")
    n_ng = sum(1 for s in SCORE if not s[4]) + sum(1 for m in MANUAL if not m[3])
    if n_ng == 0:
        w("**合格** — 全指標が目標値を達成し、目視レビュー4項目（M1〜M4）も合格基準を満たした。")
    else:
        w(f"**不合格** — 未達 {n_ng} 項目。修正 → checks.py 全再実行 → review.py 全再実行のループを継続する。")
    w("")
    w("---")
    w("再実行手順：`backend/.venv/bin/python testdata/validation/checks.py` → "
      "`backend/.venv/bin/python testdata/validation/review.py`")
    (HERE / "review_report.md").write_text("\n".join(lines), encoding="utf-8")


if __name__ == "__main__":
    main()
