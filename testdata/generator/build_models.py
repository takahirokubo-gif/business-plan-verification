# -*- coding: utf-8 -*-
"""テストデータA/Bの財務モデルExcel群を生成する。

生成物：
- testdata/A_standard/AutostaffChubu_Model_A.xlsx        （Base派生・クリーン・12期）
- testdata/B_hard/AutostaffChubu_Model_Base_B.xlsx       （B改修版）
- testdata/B_hard/AutostaffChubu_Model_Sponsor_B.xlsx    （B改修版・Sponsor）
- testdata/B_hard/AutostaffChubu_Debt_B.xlsx             （分離デットスケジュール）
- testdata/B_hard/DD_*.pdf ×4                            （原本の無変更コピー）
- testdata/generator/build/planted_items_models.json     （仕込み一覧・機械生成）
- testdata/generator/build/b_hashes.json                 （DD4分冊のSHA-256）

方針：A・Bとも spec12（数値正本）から全数値を導出し、実績年=値・計画年=数式で書く。
B本体は Debt_Schedule シートを持たず、PLの支払利息（FY27〜33）が =#REF! の破損状態。
"""
import hashlib
import json
import shutil
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Side, Border

import spec12 as S
import xl12 as L
from spec12 import ACTUAL_YEARS, COL, PLAN_YEARS, YEARS

ROOT = Path(__file__).resolve().parents[1]          # testdata/
DUMMY = ROOT.parent / "dummy_input"
BUILD = Path(__file__).resolve().parent / "build"

TITLE_FONT = Font(name="Yu Gothic", size=14, bold=True, color="1A4F8B")
HEADER_FONT = Font(name="Yu Gothic", size=10, bold=True)
BASE_FONT = Font(name="Yu Gothic", size=10)
NOTE_FONT = Font(name="Yu Gothic", size=9, color="737781")
HEADER_FILL = PatternFill("solid", fgColor="EDEDF3")
ACTUAL_FILL = PatternFill("solid", fgColor="F5F5F5")
SIDE_FILL = PatternFill("solid", fgColor="FFF7E6")
THIN = Side(style="thin", color="C2C6D1")
BORDER = Border(bottom=THIN)

NUM_FMT = '#,##0;"△" #,##0'
PCT_FMT = "0.0%"

PLANTED: list[dict] = []


def plant(cid, dataset, file, sheet, cell, desc, expect):
    PLANTED.append(dict(id=cid, dataset=dataset, file=file, sheet=sheet,
                        cell=cell, desc=desc, expect=expect))


# ---------------------------------------------------------------- 共通部品

def _sheet_header(ws, title, years=True, case_label=None,
                  a2_note=None, right_note=None):
    """タイトル＋（オプションで）12期ヘッダー・実績/計画区分行を書く。"""
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    if a2_note:
        ws["A2"] = a2_note
        ws["A2"].font = NOTE_FONT
    if case_label:
        ws["E2"] = f"ケース: {case_label}"
        ws["E2"].font = HEADER_FONT
    if right_note:
        c = ws[f"{COL[YEARS[-1]]}2"]
        c.value = right_note
        c.font = NOTE_FONT
        c.alignment = Alignment(horizontal="right")
    if years:
        for y in YEARS:
            c = ws[f"{COL[y]}{L.HEADER_ROW}"]
            c.value = y
            c.font = HEADER_FONT
            c.fill = HEADER_FILL
            c.alignment = Alignment(horizontal="center")
            k = ws[f"{COL[y]}{L.KIND_ROW}"]
            k.value = "実績" if y in ACTUAL_YEARS else "計画"
            k.font = NOTE_FONT
            k.alignment = Alignment(horizontal="center")
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 34
    for y in YEARS:
        ws.column_dimensions[COL[y]].width = 12.5


def _put_row(ws, row_def, values_or_formulas, fmt=NUM_FMT, label_override=None):
    row, label = row_def
    lc = ws[f"B{row}"]
    lc.value = label_override if label_override is not None else label
    lc.font = BASE_FONT
    for y in YEARS:
        v = values_or_formulas.get(y)
        if v is None:
            continue
        c = ws[f"{COL[y]}{row}"]
        c.value = v
        c.number_format = fmt
        c.font = BASE_FONT
        if y in ACTUAL_YEARS:
            c.fill = ACTUAL_FILL


def _mixed(formula_tpl, actual_values):
    out = {y: actual_values[y] for y in ACTUAL_YEARS}
    for y in PLAN_YEARS:
        out[y] = formula_tpl.format(c=COL[y])
    return out


def _all_formula(formula_tpl, years=YEARS):
    return {y: formula_tpl.format(c=COL[y]) for y in years}


# ---------------------------------------------------------------- 各シート

def _build_cover(wb, case, variant):
    ws = wb.create_sheet(L.SHEET_COVER)
    ws["B4"] = "Project Gear"
    ws["B4"].font = Font(name="Yu Gothic", size=24, bold=True, color="1A4F8B")
    ws["B6"] = "株式会社オートスタッフ中部　財務モデル"
    ws["B6"].font = Font(name="Yu Gothic", size=16, bold=True)
    ws["B8"] = f"ケース：{S.CASE_LABEL[case]}（{S.CASE_LABEL_JA[case]}）"
    ws["B8"].font = Font(name="Yu Gothic", size=12, bold=True)
    ws["B10"] = "作成：日本橋キャピタルパートナーズ株式会社"
    ws["B11"] = "作成日：2026年7月10日（Ver. 3.4・12期拡張版）"
    ws["B13"] = "【取扱厳重注意】"
    ws["B14"] = ("本資料は貴行における本件検討のためにのみ作成されたものであり、"
                 "作成者の事前の書面による同意なく第三者への開示・複製を禁じます。")
    ws["B15"] = ("本資料に含まれる将来予測は多くの前提・仮定に基づくものであり、"
                 "その実現性を保証するものではありません。")
    for r in range(10, 16):
        ws[f"B{r}"].font = NOTE_FONT if r >= 13 else BASE_FONT
    ws.column_dimensions["B"].width = 90


def _build_assumptions(wb, case, variant):
    ws = wb.create_sheet(L.SHEET_ASSUMPTIONS)
    _sheet_header(ws, "前提条件（Assumptions）", years=True,
                  case_label=S.CASE_LABEL[case], a2_note="（単位：千円）")
    ws[f"B{L.ASM_YEARLY_SECTION_ROW}"] = "オペレーティング前提（年度別）"
    ws[f"B{L.ASM_YEARLY_SECTION_ROW}"].font = HEADER_FONT
    for key, (row, label, tpl, fmt) in L.ASM_YEARLY_ROWS.items():
        _put_row(ws, (row, label), _all_formula(tpl), fmt)
    ws[f"B{L.ASM_STRUCT_SECTION_ROW}"] = "ストラクチャー（千円）"
    ws[f"B{L.ASM_STRUCT_SECTION_ROW}"].font = HEADER_FONT
    values = dict(ev=12_000_000, equity=5_500_000, senior=6_500_000, bank_a=2_500_000,
                  tenor=7, rate=S.INTEREST_RATE, repay=S.ANNUAL_REPAYMENT,
                  close="2026年9月末", goodwill=S.GOODWILL,
                  tax_rate=S.TAX_RATE, ar_days=S.AR_DAYS, cash_adj=-750_000)
    for row, label, key, fmt in L.ASM_KV_ROWS + L.ASM_OTHER_ROWS:
        ws[f"B{row}"] = label
        ws[f"B{row}"].font = BASE_FONT
        c = ws[f"D{row}"]
        c.value = values[key]
        if fmt:
            c.number_format = fmt
        c.font = BASE_FONT
    ws[f"B{L.ASM_OTHER_SECTION_ROW}"] = "その他前提"
    ws[f"B{L.ASM_OTHER_SECTION_ROW}"].font = HEADER_FONT
    if variant == "B":
        ws[f"B{L.ASM_TEASER_ROW}"] = L.ASM_TEASER_LABEL
        ws[f"B{L.ASM_TEASER_ROW}"].font = BASE_FONT
        c = ws[f"D{L.ASM_TEASER_ROW}"]
        c.value = S.DEAL["sponsor_ebitda_mm"] * 1000
        c.number_format = NUM_FMT
        c.font = BASE_FONT
        ws[f"B{L.ASM_OLDNOTE_ROW}"] = L.ASM_OLDNOTE
        ws[f"B{L.ASM_OLDNOTE_ROW}"].font = NOTE_FONT


def _build_kpi(wb, case, variant):
    ws = wb.create_sheet(L.SHEET_KPI)
    _sheet_header(ws, "KPIドライバー（Operating Assumptions）",
                  case_label=S.CASE_LABEL[case], a2_note="（単位：名・時間・円・％）")
    drv = S.drivers(case)
    K = L.KPI_ROWS
    _put_row(ws, K["enrolled"], {y: drv[y]["enrolled"] for y in YEARS})
    _put_row(ws, K["util"], {y: drv[y]["util"] for y in YEARS}, PCT_FMT)
    _put_row(ws, K["active"],
             _all_formula("=ROUND({c}%d*{c}%d,0)" % (K["enrolled"][0], K["util"][0])))
    _put_row(ws, K["hours"], {y: drv[y]["hours"] for y in YEARS})
    _put_row(ws, K["bill"], {y: drv[y]["bill"] for y in YEARS})
    _put_row(ws, K["wage"], {y: drv[y]["wage"] for y in YEARS})
    _put_row(ws, K["welfare"], {y: drv[y]["welfare"] for y in YEARS}, PCT_FMT)
    _put_row(ws, K["hires"], {y: drv[y]["hires"] for y in YEARS})
    _put_row(ws, K["cpa"], {y: drv[y]["cpa"] for y in YEARS})
    _put_row(ws, K["attrition"], {y: drv[y]["attrition"] for y in YEARS}, PCT_FMT)
    ws["B18"] = "※ 稼働人数 = 在籍登録スタッフ数 × 稼働率（四捨五入）"
    ws["B18"].font = NOTE_FONT
    if variant == "B":
        st = L.SIDE_TABLE_KPI
        ws[st["title_cell"]] = st["title"]
        ws[st["title_cell"]].font = HEADER_FONT
        col0, row0 = st["origin"]
        headers = ["採用チャネル", "採用人数（名）", "構成比"]
        for j, h in enumerate(headers):
            c = ws[f"{chr(ord(col0) + j)}{row0}"]
            c.value = h
            c.font = HEADER_FONT
            c.fill = SIDE_FILL
        total = 0
        for i, (name, n, share) in enumerate(L.KPI_CHANNELS):
            r = row0 + 1 + i
            ws[f"{col0}{r}"] = name
            ws[f"{chr(ord(col0) + 1)}{r}"] = n
            ws[f"{chr(ord(col0) + 1)}{r}"].number_format = NUM_FMT
            ws[f"{chr(ord(col0) + 2)}{r}"] = share
            ws[f"{chr(ord(col0) + 2)}{r}"].number_format = "0%"
            for j in range(3):
                ws[f"{chr(ord(col0) + j)}{r}"].font = BASE_FONT
            total += n
        r = row0 + 1 + len(L.KPI_CHANNELS)
        ws[f"{col0}{r}"] = "合計"
        ws[f"{chr(ord(col0) + 1)}{r}"] = total
        ws[f"{chr(ord(col0) + 1)}{r}"].number_format = NUM_FMT
        ws[f"{chr(ord(col0) + 2)}{r}"] = 1.0
        ws[f"{chr(ord(col0) + 2)}{r}"].number_format = "0%"
        for j in range(3):
            ws[f"{chr(ord(col0) + j)}{r}"].font = HEADER_FONT
        assert total == S.drivers(case)["FY26"]["hires"]


def _build_debt_rows(ws, broken=False):
    """デットスケジュールの中身（12期ヘッダー・計画7期のみデータ）。"""
    D = L.DEBT_ROWS
    sched = S.debt_schedule()
    opening = {}
    for i, y in enumerate(PLAN_YEARS):
        opening[y] = S.SENIOR_TOTAL if i == 0 else f"={COL[PLAN_YEARS[i - 1]]}{D['closing'][0]}"
    _put_row(ws, D["opening"], opening)
    repay = {y: S.ANNUAL_REPAYMENT for y in PLAN_YEARS[:-1]}
    repay["FY33"] = f"={COL['FY33']}{D['opening'][0]}"   # 最終回残額一括
    _put_row(ws, D["repayment"], repay)
    _put_row(ws, D["closing"],
             _all_formula("={c}%d-{c}%d" % (D["opening"][0], D["repayment"][0]), PLAN_YEARS))
    _put_row(ws, D["avg"],
             _all_formula("=({c}%d+{c}%d)/2" % (D["opening"][0], D["closing"][0]), PLAN_YEARS))
    _put_row(ws, D["rate"], {y: S.INTEREST_RATE for y in PLAN_YEARS}, PCT_FMT)
    _put_row(ws, D["interest"],
             _all_formula("=ROUND({c}%d*{c}%d,0)" % (D["avg"][0], D["rate"][0]), PLAN_YEARS))
    ws["B14"] = L.DEBT_NOTE
    ws["B14"].font = NOTE_FONT
    return sched


def _build_debt_sheet(wb, case):
    ws = wb.create_sheet(L.SHEET_DEBT)
    _sheet_header(ws, "Debt Schedule (Senior Loan)", case_label=None,
                  a2_note="(Unit: JPY thousands)")
    ws["E2"] = f"Case: {S.CASE_LABEL[case]}"
    ws["E2"].font = HEADER_FONT
    _build_debt_rows(ws)


def _build_pl(wb, case, variant):
    ws = wb.create_sheet(L.SHEET_PL)
    _sheet_header(ws, "損益計算書（Projected P&L）", case_label=S.CASE_LABEL[case],
                  right_note="単位:千円")
    pl = S.compute_pl(case)
    P, K = L.PL_ROWS, L.KPI_ROWS
    ren = {key: new for (_, key, _, new, _) in
           [(s, k, o, n, d) for (s, k, o, n, d) in L.RENAMES if s == L.SHEET_PL]} \
        if variant == "B" else {}

    _put_row(ws, P["staffing_rev"], _mixed(
        "=ROUND(KPI_Drivers!{c}%d*KPI_Drivers!{c}%d*KPI_Drivers!{c}%d*12/1000,0)"
        % (K["active"][0], K["hours"][0], K["bill"][0]),
        {y: pl[y]["staffing_rev"] for y in ACTUAL_YEARS}))
    _put_row(ws, P["other_revenue"], {y: pl[y]["other_revenue"] for y in YEARS})
    _put_row(ws, P["revenue"],
             _all_formula("=SUM({c}%d:{c}%d)" % (P["staffing_rev"][0], P["other_revenue"][0])),
             label_override=ren.get("revenue"))
    _put_row(ws, P["labor"], _mixed(
        "=ROUND(KPI_Drivers!{c}%d*KPI_Drivers!{c}%d*KPI_Drivers!{c}%d*(1+KPI_Drivers!{c}%d)*12/1000,0)"
        % (K["active"][0], K["hours"][0], K["wage"][0], K["welfare"][0]),
        {y: pl[y]["labor"] for y in ACTUAL_YEARS}), label_override=ren.get("labor"))
    _put_row(ws, P["other_cogs"], {y: pl[y]["other_cogs"] for y in YEARS})
    _put_row(ws, P["cogs"], _all_formula("={c}%d+{c}%d" % (P["labor"][0], P["other_cogs"][0])))
    _put_row(ws, P["gross"], _all_formula("={c}%d-{c}%d" % (P["revenue"][0], P["cogs"][0])))
    _put_row(ws, P["gross_margin"],
             _all_formula("={c}%d/{c}%d" % (P["gross"][0], P["revenue"][0])), PCT_FMT)
    _put_row(ws, P["recruiting"], _mixed(
        "=KPI_Drivers!{c}%d*KPI_Drivers!{c}%d" % (K["hires"][0], K["cpa"][0]),
        {y: pl[y]["recruiting"] for y in ACTUAL_YEARS}))
    _put_row(ws, P["hq_cost"], {y: pl[y]["hq_cost"] for y in YEARS})
    _put_row(ws, P["other_sga"], {y: pl[y]["other_sga"] for y in YEARS})
    _put_row(ws, P["sga"],
             _all_formula("=SUM({c}%d:{c}%d)" % (P["recruiting"][0], P["other_sga"][0])))
    _put_row(ws, P["op"], _all_formula("={c}%d-{c}%d" % (P["gross"][0], P["sga"][0])))
    _put_row(ws, P["op_margin"],
             _all_formula("={c}%d/{c}%d" % (P["op"][0], P["revenue"][0])), PCT_FMT)
    _put_row(ws, P["depreciation"], {y: pl[y]["depreciation"] for y in YEARS})
    _put_row(ws, P["ebitda"], _all_formula("={c}%d+{c}%d" % (P["op"][0], P["depreciation"][0])),
             label_override=ren.get("ebitda"))
    _put_row(ws, P["ebitda_margin"],
             _all_formula("={c}%d/{c}%d" % (P["ebitda"][0], P["revenue"][0])), PCT_FMT)
    if variant == "A":
        _put_row(ws, P["interest"], _mixed(
            "=Debt_Schedule!{c}%d" % L.DEBT_ROWS["interest"][0],
            {y: pl[y]["interest"] for y in ACTUAL_YEARS}))
    else:
        # E11：Debt_Schedule分離により計画期の支払利息が解決不能な参照（#REF!）
        vals = {y: pl[y]["interest"] for y in ACTUAL_YEARS}
        for y in PLAN_YEARS:
            vals[y] = L.BROKEN_INTEREST_FORMULA.format(c=COL[y])
        _put_row(ws, P["interest"], vals)
    _put_row(ws, P["non_op"], {y: pl[y]["non_op"] for y in YEARS})
    if variant == "A":
        _put_row(ws, P["ordinary"],
                 _all_formula("={c}%d-{c}%d+{c}%d" % (P["op"][0], P["interest"][0], P["non_op"][0])))
        _put_row(ws, P["tax"], _all_formula("=ROUND({c}%d*0.31,0)" % P["ordinary"][0]))
        _put_row(ws, P["ni"], _all_formula("={c}%d-{c}%d" % (P["ordinary"][0], P["tax"][0])))
    else:
        # 破損セルの下流は値貼り付け状態（Debt_B入手後に検算可能な正値）
        ord_vals = _all_formula(
            "={c}%d-{c}%d+{c}%d" % (P["op"][0], P["interest"][0], P["non_op"][0]), ACTUAL_YEARS)
        tax_vals = _all_formula("=ROUND({c}%d*0.31,0)" % P["ordinary"][0], ACTUAL_YEARS)
        ni_vals = _all_formula("={c}%d-{c}%d" % (P["ordinary"][0], P["tax"][0]), ACTUAL_YEARS)
        for y in PLAN_YEARS:
            ord_vals[y] = pl[y]["ordinary"]
            tax_vals[y] = pl[y]["tax"]
            ni_vals[y] = pl[y]["ni"]
        _put_row(ws, P["ordinary"], ord_vals)
        _put_row(ws, P["tax"], tax_vals)
        _put_row(ws, P["ni"], ni_vals)
    for key in ("revenue", "gross", "op", "ebitda", "ordinary", "ni"):
        row = P[key][0]
        for y in YEARS:
            ws[f"{COL[y]}{row}"].font = HEADER_FONT
        ws[f"B{row}"].font = HEADER_FONT
    if variant == "B":
        # E10：単価改定履歴の補足小表（本表グリッド外・P列以降）
        st = L.SIDE_TABLE_PL
        ws[st["title_cell"]] = st["title"]
        ws[st["title_cell"]].font = HEADER_FONT
        col0, row0 = st["origin"]
        for j, h in enumerate(["年度", "改定率", "改定後単価（円/h）"]):
            c = ws[f"{chr(ord(col0) + j)}{row0}"]
            c.value = h
            c.font = HEADER_FONT
            c.fill = SIDE_FILL
        drv = S.drivers(case)
        prev = drv["FY26"]["bill"]
        for i, y in enumerate(PLAN_YEARS):
            r = row0 + 1 + i
            bill = drv[y]["bill"]
            ws[f"{chr(ord(col0))}{r}"] = y
            ws[f"{chr(ord(col0) + 1)}{r}"] = bill / prev - 1
            ws[f"{chr(ord(col0) + 1)}{r}"].number_format = "0.0%"
            ws[f"{chr(ord(col0) + 2)}{r}"] = bill
            ws[f"{chr(ord(col0) + 2)}{r}"].number_format = NUM_FMT
            for j in range(3):
                ws[f"{chr(ord(col0) + j)}{r}"].font = BASE_FONT
            prev = bill


def _build_bs(wb, case, variant):
    ws = wb.create_sheet(L.SHEET_BS)
    _sheet_header(ws, "貸借対照表（Projected B/S）", case_label=S.CASE_LABEL[case],
                  a2_note="（単位：千円）")
    bs, _ = S.compute_bs_cf(case)
    B = L.BS_ROWS
    ren = {key: new for (s, key, _, new, _) in L.RENAMES if s == L.SHEET_BS} \
        if variant == "B" else {}
    _put_row(ws, B["cash"], {y: bs[y]["cash"] for y in YEARS})
    _put_row(ws, B["ar"], {y: bs[y]["ar"] for y in YEARS}, label_override=ren.get("ar"))
    _put_row(ws, B["oca"], {y: bs[y]["oca"] for y in YEARS}, label_override=ren.get("oca"))
    _put_row(ws, B["current_assets"],
             _all_formula("=SUM({c}%d:{c}%d)" % (B["cash"][0], B["oca"][0])))
    _put_row(ws, B["ppe"], {y: bs[y]["ppe"] for y in YEARS})
    _put_row(ws, B["goodwill"], {y: bs[y]["goodwill"] for y in YEARS})
    _put_row(ws, B["intangible"], {y: bs[y]["intangible"] for y in YEARS})
    _put_row(ws, B["fixed_assets"],
             _all_formula("=SUM({c}%d:{c}%d)" % (B["ppe"][0], B["intangible"][0])))
    _put_row(ws, B["total_assets"],
             _all_formula("={c}%d+{c}%d" % (B["current_assets"][0], B["fixed_assets"][0])))
    _put_row(ws, B["ap"], {y: bs[y]["ap"] for y in YEARS}, label_override=ren.get("ap"))
    _put_row(ws, B["other_liab"], {y: bs[y]["other_liab"] for y in YEARS})
    if variant == "A":
        _put_row(ws, B["debt"], _mixed(
            "=Debt_Schedule!{c}%d" % L.DEBT_ROWS["closing"][0],
            {y: bs[y]["debt"] for y in ACTUAL_YEARS}))
    else:
        _put_row(ws, B["debt"], {y: bs[y]["debt"] for y in YEARS})  # 値貼り付け状態
    _put_row(ws, B["total_liab"], _all_formula("=SUM({c}%d:{c}%d)" % (B["ap"][0], B["debt"][0])))
    _put_row(ws, B["net_assets"], {y: bs[y]["net_assets"] for y in YEARS})
    _put_row(ws, B["total_le"],
             _all_formula("={c}%d+{c}%d" % (B["total_liab"][0], B["net_assets"][0])))
    ws["B26"] = ("※ FY27以降は買収後連結（のれん計上・買収時調整を純資産に含む）。"
                 "のれんは暫定PPA前の想定額。")
    ws["B26"].font = NOTE_FONT
    for key in ("total_assets", "total_le", "net_assets"):
        ws[f"B{B[key][0]}"].font = HEADER_FONT
    if variant == "B":
        st = L.SIDE_TABLE_BS
        ws[st["title_cell"]] = st["title"]
        ws[st["title_cell"]].font = HEADER_FONT
        col0, row0 = st["origin"]
        for i, (k, v) in enumerate(L.DORM_MEMO):
            r = row0 + i
            ws[f"{col0}{r}"] = k
            ws[f"{col0}{r}"].font = BASE_FONT
            ws[f"{col0}{r}"].fill = SIDE_FILL
            c = ws[f"{chr(ord(col0) + 1)}{r}"]
            c.value = v
            c.number_format = NUM_FMT
            c.font = BASE_FONT


def _build_cf(wb, case, variant):
    ws = wb.create_sheet(L.SHEET_CF)
    _sheet_header(ws, "キャッシュフロー計算書（Projected C/F・千円）",
                  case_label=S.CASE_LABEL[case])
    pl = S.compute_pl(case)
    _, cf = S.compute_bs_cf(case)
    C = L.CF_ROWS
    _put_row(ws, C["ni"], _all_formula("=PL!{c}%d" % L.PL_ROWS["ni"][0]))
    _put_row(ws, C["depreciation"], _all_formula("=PL!{c}%d" % L.PL_ROWS["depreciation"][0]))
    wc_values = {y: cf[y]["op_cf"] - pl[y]["ni"] - pl[y]["depreciation"] for y in YEARS}
    _put_row(ws, C["wc"], wc_values)
    _put_row(ws, C["op_cf"], _all_formula("=SUM({c}%d:{c}%d)" % (C["ni"][0], C["wc"][0])))
    _put_row(ws, C["inv_cf"], {y: cf[y]["inv_cf"] for y in YEARS})
    _put_row(ws, C["fcf"], _all_formula("={c}%d+{c}%d" % (C["op_cf"][0], C["inv_cf"][0])))
    sched = S.debt_schedule()
    repay, fin_other = {}, {}
    for y in ACTUAL_YEARS:
        repay[y] = -50_000
        fin_other[y] = cf[y]["fin_cf"] + 50_000
    for y in PLAN_YEARS:
        if variant == "A":
            repay[y] = f"=-Debt_Schedule!{COL[y]}{L.DEBT_ROWS['repayment'][0]}"
        else:
            repay[y] = -sched[y]["repayment"]   # 値貼り付け状態
        fin_other[y] = cf[y]["fin_cf"] + sched[y]["repayment"]
    _put_row(ws, C["repay"], repay)
    _put_row(ws, C["fin_other"], fin_other)
    _put_row(ws, C["fin_cf"], _all_formula("={c}%d+{c}%d" % (C["repay"][0], C["fin_other"][0])))
    _put_row(ws, C["net_change"], _all_formula("={c}%d+{c}%d" % (C["fcf"][0], C["fin_cf"][0])))
    opening = {"FY22": S.OPENING_CASH_FY22, "FY27": S.POST_CLOSE_CASH}
    for i, y in enumerate(YEARS):
        if y in opening:
            continue
        opening[y] = f"={COL[YEARS[i - 1]]}{C['closing_cash'][0]}"
    _put_row(ws, C["opening_cash"], opening)
    _put_row(ws, C["closing_cash"],
             _all_formula("={c}%d+{c}%d" % (C["net_change"][0], C["opening_cash"][0])))
    ws["B24"] = "※ FY27期首現金は買収時調整後（取引費用等750,000千円控除後）。"
    ws["B24"].font = NOTE_FONT


def _build_monthly(wb, variant):
    ws = wb.create_sheet(L.SHEET_MONTHLY)
    ws["A1"] = "月次試算表_FY26（2025年4月～2026年3月）"
    ws["A1"].font = TITLE_FONT
    if variant == "A":
        ws["A2"] = "（単位：円）"
        ws["A2"].font = NOTE_FONT
    mon = S.compute_monthly()
    for m in S.MONTHS:
        c = ws[f"{L.MCOL[m]}{L.HEADER_ROW}"]
        c.value = m
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")
    tc = ws[f"{L.MONTHLY_TOTAL_COL}{L.HEADER_ROW}"]
    tc.value = "年度計"
    tc.font = HEADER_FONT
    tc.fill = HEADER_FILL
    tc.alignment = Alignment(horizontal="center")
    for key, (row, label_a, label_b) in L.MONTHLY_ROWS.items():
        ws[f"B{row}"] = label_a if variant == "A" else label_b
        ws[f"B{row}"].font = BASE_FONT
        for m in S.MONTHS:
            c = ws[f"{L.MCOL[m]}{row}"]
            if key == "gross":
                c.value = f"={L.MCOL[m]}{L.MONTHLY_ROWS['revenue'][0]}-{L.MCOL[m]}{L.MONTHLY_ROWS['cogs'][0]}"
            elif key == "op":
                c.value = f"={L.MCOL[m]}{L.MONTHLY_ROWS['gross'][0]}-{L.MCOL[m]}{L.MONTHLY_ROWS['sga'][0]}"
            else:
                c.value = mon[key][m]
            c.number_format = NUM_FMT
            c.font = BASE_FONT
        if key in ("revenue", "cogs", "gross", "sga", "op"):
            t = ws[f"{L.MONTHLY_TOTAL_COL}{row}"]
            t.value = f"=SUM(C{row}:N{row})"
            t.number_format = NUM_FMT
            t.font = HEADER_FONT
    ws["B17"] = "※ 勤怠・請求データより作成（税理士法人巡回監査済）。"
    ws["B17"].font = NOTE_FONT
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 26
    for m in S.MONTHS:
        ws.column_dimensions[L.MCOL[m]].width = 14
    ws.column_dimensions[L.MONTHLY_TOTAL_COL].width = 15


def _build_pl_old(wb, case):
    """旧版シート（8期・FY24〜31・旧科目名・0.965倍）。原本と同じ構成のノイズ。"""
    ws = wb.create_sheet(L.SHEET_PL_OLD)
    ws["A1"] = "損益計算書【旧版 v2.1】※使用しないこと（FY26速報反映前）"
    ws["A1"].font = Font(name="Yu Gothic", size=14, bold=True, color="BA1A1A")
    ws["A2"] = "（単位：千円）"
    ws["A2"].font = NOTE_FONT
    import spec as spec8  # dummy_input（8期版）
    old_years = spec8.YEARS
    old_col = {y: chr(ord("C") + i) for i, y in enumerate(old_years)}
    for y in old_years:
        c = ws[f"{old_col[y]}{L.HEADER_ROW}"]
        c.value = y
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
    pl = S.compute_pl(case)
    for key in ("staffing_rev", "other_revenue", "revenue", "labor", "cogs",
                "gross", "sga", "op", "depreciation", "ebitda", "ni"):
        row, label = L.PL_ROWS[key]
        ws[f"B{row}"] = label   # 旧表記（改称前ラベル）のまま
        ws[f"B{row}"].font = BASE_FONT
        for y in old_years:
            c = ws[f"{old_col[y]}{row}"]
            c.value = int(pl[y][key] * 0.965)
            c.number_format = NUM_FMT
            c.font = BASE_FONT
    ws.column_dimensions["B"].width = 34


def _build_scratch(wb):
    ws = wb.create_sheet(L.SHEET_SCRATCH)
    ws["A1"] = "（作業用シート・成果物ではありません）"
    ws["A1"].font = NOTE_FONT
    scratch = [
        ("B3", "チェック: 3,344×166×2,100×12 ="),
        ("D3", 13_988_620_800),
        ("B4", "→ PL派遣売上と一致確認済（6/25 IY）"),
        ("B6", "TODO: 労務費の福利率 13.8%→14.0%に更新（6/28 反映済）"),
        ("B8", "単価改定メモ: トヨタ系 4月改定 +1.2%規定 / デンソー系 10月"),
        ("B10", 4_233_600),
        ("C10", "←年間換算係数（166h×2,125×12）ミス。使わない"),
        ("B12", "旧のれん試算 5,656,000 → PPA仮 5,500,000"),
    ]
    for addr, v in scratch:
        ws[addr] = v
        ws[addr].font = BASE_FONT


# ---------------------------------------------------------------- ワークブック組み立て

TAB_COLORS = {
    L.SHEET_COVER: "1A4F8B", L.SHEET_ASSUMPTIONS: "5B79A8",
    L.SHEET_PL: "2E6DA4", L.SHEET_BS: "2E6DA4", L.SHEET_CF: "2E6DA4",
    L.SHEET_KPI: "3D8B72", L.SHEET_DEBT: "3D8B72", L.SHEET_MONTHLY: "8A6FBF",
    L.SHEET_PL_OLD: "9AA0A6", L.SHEET_SCRATCH: "9AA0A6", L.SHEET_EMPTY: "9AA0A6",
}


def _polish(wb):
    """実務的な体裁：タブ色・フリーズペイン・印刷設定（値・数式には影響しない）。"""
    year_sheets = {L.SHEET_PL, L.SHEET_BS, L.SHEET_CF, L.SHEET_KPI,
                   L.SHEET_ASSUMPTIONS, L.SHEET_DEBT}
    for ws in wb.worksheets:
        color = TAB_COLORS.get(ws.title)
        if color:
            ws.sheet_properties.tabColor = color
        if ws.title in year_sheets:
            ws.freeze_panes = "C6"
        elif ws.title == L.SHEET_MONTHLY:
            ws.freeze_panes = "C5"
        ws.page_setup.orientation = "landscape"
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0


def build_model(case: str, variant: str) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    _build_cover(wb, case, variant)
    _build_assumptions(wb, case, variant)
    _build_kpi(wb, case, variant)
    if variant == "A":
        _build_debt_sheet(wb, case)
    _build_pl(wb, case, variant)
    _build_bs(wb, case, variant)
    _build_cf(wb, case, variant)
    _build_monthly(wb, variant)
    if variant == "B":
        _build_pl_old(wb, case)
        _build_scratch(wb)
        wb.create_sheet(L.SHEET_EMPTY)
        order = [L.SHEET_COVER, L.SHEET_ASSUMPTIONS, L.SHEET_PL, L.SHEET_BS, L.SHEET_CF,
                 L.SHEET_KPI, L.SHEET_MONTHLY, L.SHEET_PL_OLD, L.SHEET_SCRATCH, L.SHEET_EMPTY]
    else:
        order = [L.SHEET_COVER, L.SHEET_ASSUMPTIONS, L.SHEET_PL, L.SHEET_BS, L.SHEET_CF,
                 L.SHEET_KPI, L.SHEET_DEBT, L.SHEET_MONTHLY]
    wb._sheets = [wb[name] for name in order]
    _polish(wb)
    return wb


def build_debt_b() -> Workbook:
    """分離されたデットスケジュール（別送ファイルの再現・英語ラベル）。"""
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet(L.SHEET_DEBT)
    _sheet_header(ws, "Debt Schedule (Senior Loan)", a2_note="(Unit: JPY thousands)")
    ws["E2"] = "Project Gear — Senior Facility"
    ws["E2"].font = HEADER_FONT
    _build_debt_rows(ws)
    _polish(wb)
    return wb


# ---------------------------------------------------------------- planted登録

def _register_common(dataset, fname):
    plant("E1", dataset, fname, "-", "-",
          "PL・BS・CFが別シートの複数シート構成（Cover除き6シート以上）", "自動処理")
    plant("E1/E5", dataset, fname, "PL/BS/CF/KPI_Drivers/Assumptions", "C4:N5",
          "FY22〜FY33の12期ヘッダー（実績5期＋計画7期）と実績/計画区分行", "自動処理")
    plant("E3", dataset, fname, "PL", "N2", "単位表記の位置ゆれ①：ヘッダー行右端「単位:千円」", "自動処理")
    plant("E3", dataset, fname, "CF", "A1", "単位表記の位置ゆれ②：タイトル内「…（Projected C/F・千円）」", "自動処理")
    plant("E3", dataset, fname, "BS", "A2", "単位表記の位置ゆれ③：A2セル「（単位：千円）」", "自動処理")
    plant("E4", dataset, fname, "PL/BS/KPI_Drivers", "B9,B21,B13,B23,B8",
          "英語併記科目：営業利益（Operating Profit）・のれん（Goodwill）・純資産（Net Assets）・稼働率（Utilization）等",
          "自動処理")
    plant("E5", dataset, fname, "PL/BS/CF/KPI_Drivers", "C:D列",
          "FY22〜23実績の新規作成（全期黒字・売上CAGR6.4%・貸借一致）", "自動処理")
    plant("E5", dataset, fname, "PL/BS/CF/KPI_Drivers", "M:N列",
          "FY32〜33計画の新規作成（稼働率88%維持・単価改定+1%/年・FY33残額一括弁済）", "自動処理")
    plant("E6", dataset, fname, "KPI_Drivers/月次試算表_FY26", "-",
          "業界用語（稼働率・派遣単価・在籍登録スタッフ・稼働人数・採用単価（CPA）・離職率・法定福利費率等）", "自動処理")


def register_planted_a():
    f = "A_standard/AutostaffChubu_Model_A.xlsx"
    _register_common("A", f)
    plant("E2", "A", f, "月次試算表_FY26", "C7:N15",
          "月次試算表（単位：円・A2ラベル明記）。12ヶ月合計＝年次PL FY26、3月末残高＝年次BS FY26", "自動処理")
    plant("E2", "A", f, "月次試算表_FY26", "D7,G7,L7",
          "季節性：2025/05△6.0%・2025/08△6.5%・2026/01△7.0%（対他9ヶ月平均）", "自動処理")
    plant("E3", "A", f, "Debt_Schedule", "A2",
          "単位表記の位置ゆれ④：英語表記「(Unit: JPY thousands)」", "自動処理")
    plant("E4", "A", f, "Debt_Schedule", "B7:B14",
          "行ラベル英語100%（Beginning Balance / Mandatory Amortization / Ending Balance等）", "自動処理")


def register_planted_b():
    for f in ("B_hard/AutostaffChubu_Model_Base_B.xlsx",
              "B_hard/AutostaffChubu_Model_Sponsor_B.xlsx"):
        _register_common("B", f)
        plant("E2", "B", f, "月次試算表_FY26", "シート全体",
              "月次試算表に単位ラベル0件（円フル桁のみ）→単位の確認照会が必要", "検知して照会")
        for (sheet, key, old, new, dd) in L.RENAMES:
            rows = {L.SHEET_PL: L.PL_ROWS, L.SHEET_BS: L.BS_ROWS}[sheet]
            plant("E7/X1", "B", f, sheet, f"B{rows[key][0]}",
                  f"科目改称「{old}」→「{new}」（DD側呼称は「{dd}」）", "検知して照会")
        plant("E8", "B", f, "PL_old", "A1",
              "旧版シート（「※使用しないこと」注記付き・旧科目名・全値0.965倍）", "正シート・正ファイル識別")
        plant("E8", "B", f, "Scratch", "B10:C10",
              "作業用シート（誤った年間換算係数4,233,600を含む）", "正シート・正ファイル識別")
        plant("E8", "B", f, "Sheet1", "-", "空シート", "正シート・正ファイル識別")
        plant("E10", "B", f, "PL", "P3:R11",
              "本表グリッド外の補足小表①：単価改定履歴（FY27〜33・KPI_Driversの派遣単価と整合）", "検知して照会")
        plant("E10", "B", f, "KPI_Drivers", "B19:D24",
              "本表グリッド外の補足小表②：採用チャネル内訳（合計1,180名＝FY26新規採用人数と整合）", "検知して照会")
        plant("E10", "B", f, "BS", "P3:Q8",
              "本表グリッド外の補足小表③：寮関連メモ（12棟・約900名収容・借上げ賃料210,000千円）", "検知して照会")
        plant("E11", "B", f, "PL", "H28:N28",
              "支払利息（FY27〜33の7セル）が =#REF! の解決不能参照（Debt_Schedule分離に伴うリンク切れ）",
              "検知して照会")
        plant("X3", "B", f, "Assumptions", "D21",
              "のれん想定額5,500,000千円（財務DD p.31の試算5,280百万円と220百万円の差異）", "検知して照会")
        plant("X3", "B", f, "Assumptions", f"D{L.ASM_TEASER_ROW}",
              "ティーザー記載EBITDA 1,585,000千円（財務DD p.34の確定値1,620百万円と時点差）", "検知して照会")
        plant("X4", "B", f, "Scratch", "B12",
              "旧のれん試算5,656,000千円のメモ（現行5,500,000千円との時点差）", "検知して照会")
    plant("E9", "B", "B_hard/AutostaffChubu_Model_Base_B.xlsx ほか", "-", "-",
          "同一構造・同一科目名のケース違い2ファイル（Base_B/Sponsor_B・計画期の数値差分30セル以上）",
          "正シート・正ファイル識別")
    plant("E11", "B", "B_hard/AutostaffChubu_Debt_B.xlsx", "Debt_Schedule", "C4:N12",
          "分離されたデットスケジュール本体（英語ラベル・12期ヘッダー・FY33残額一括）",
          "正シート・正ファイル識別")
    plant("X2", "B", "B_hard/AutostaffChubu_Model_Base_B.xlsx ほか", "-", "-",
          "単位系不一致：Excelは千円・DD4分冊は百万円", "検知して照会")
    plant("X4", "B", "B_hard/AutostaffChubu_Model_Base_B.xlsx", "PL", "F9",
          "丸め差：Excel FY25売上高13,363,127千円 vs 財務DD p.9「13,363百万円」", "検知して照会")
    plant("D3", "B", "B_hard/DD_Financial_オートスタッフ中部.pdf", "p.6/p.33", "-",
          "多段のQoE調整表（符号付き調整行＋脚注参照）", "検知して照会")
    plant("D4", "B", "B_hard/DD_*.pdf", "-", "-",
          "DD報告書が4分冊（事業・財務・法務・税務）に分かれている", "検知して照会")
    plant("D5", "B", "B_hard/DD_Business_オートスタッフ中部.pdf", "p.11/p.19", "-",
          "分冊間クロスリファレンス（事業DD→法務DD参照等）", "検知して照会")
    plant("D6", "B", "B_hard/DD_Legal_オートスタッフ中部.pdf", "p.12", "-",
          "三表に載らない簿外リスク：未払残業代の潜在債務 最大300百万円", "検知して照会")


# ---------------------------------------------------------------- main

def main():
    BUILD.mkdir(exist_ok=True)
    a_dir = ROOT / "A_standard"
    b_dir = ROOT / "B_hard"
    a_dir.mkdir(exist_ok=True)
    b_dir.mkdir(exist_ok=True)

    S.self_check()

    wb = build_model("base", "A")
    wb.save(a_dir / "AutostaffChubu_Model_A.xlsx")
    print("generated: A_standard/AutostaffChubu_Model_A.xlsx")

    for case, fname in (("base", "AutostaffChubu_Model_Base_B.xlsx"),
                        ("sponsor", "AutostaffChubu_Model_Sponsor_B.xlsx")):
        wb = build_model(case, "B")
        wb.save(b_dir / fname)
        print(f"generated: B_hard/{fname}")

    wb = build_debt_b()
    wb.save(b_dir / "AutostaffChubu_Debt_B.xlsx")
    print("generated: B_hard/AutostaffChubu_Debt_B.xlsx")

    # DD4分冊の無変更コピー＋SHA-256記録
    hashes = {}
    for key, fname in S.spec8.DD_FILES.items():
        src = DUMMY / fname
        dst = b_dir / fname
        shutil.copyfile(src, dst)
        h_src = hashlib.sha256(src.read_bytes()).hexdigest()
        h_dst = hashlib.sha256(dst.read_bytes()).hexdigest()
        assert h_src == h_dst
        hashes[fname] = h_src
        print(f"copied: B_hard/{fname} (sha256={h_src[:16]}…)")
    (BUILD / "b_hashes.json").write_text(
        json.dumps(hashes, ensure_ascii=False, indent=2), encoding="utf-8")

    register_planted_a()
    register_planted_b()
    (BUILD / "planted_items_models.json").write_text(
        json.dumps(PLANTED, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"planted items (models): {len(PLANTED)}")


if __name__ == "__main__":
    main()
