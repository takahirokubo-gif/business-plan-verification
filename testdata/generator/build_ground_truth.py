# -*- coding: utf-8 -*-
"""ground_truth.xlsx の生成。

シート構成（claude_code_prompt v2 第7章）：
- facts        : 正解ファクト一覧（40行以上・出所セル/ページ付き・R9用の導出式付き）
- kpi_structure: 正解KPIツリー（4階層）
- name_mapping : 名寄せ対応表（B改称科目 ↔ DD呼称 ↔ A標準科目）
- planted_items: 仕込んだ複雑性の全量リスト（build_models / build_dd_pdf の出力から機械生成）
- monthly_tie  : 月次12ヶ月合計と年次FY26の突合表

導出式（derivation）は facts の key を変数名とする式で、review.py が独立に再計算する（R9）。
round() はExcelのROUND（0.5を0から遠い方へ）として評価する。
"""
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

import spec12 as S
import xl12 as L

ROOT = Path(__file__).resolve().parents[1]
BUILD = Path(__file__).resolve().parent / "build"
OUT = ROOT / "ground_truth" / "ground_truth.xlsx"

HEADER_FONT = Font(name="Yu Gothic", size=10, bold=True)
BASE_FONT = Font(name="Yu Gothic", size=10)
HEADER_FILL = PatternFill("solid", fgColor="EDEDF3")

_pl = S.compute_pl("base")
_pls = S.compute_pl("sponsor")
_bs, _cf = S.compute_bs_cf("base")
_drv = S.drivers("base")
_drvs = S.drivers("sponsor")
_sched = S.debt_schedule()
_mon = S.compute_monthly()
_season = S.monthly_seasonality(_mon["revenue"])

A_XLSX = "A_standard/AutostaffChubu_Model_A.xlsx"
B_XLSX = "B_hard/AutostaffChubu_Model_Base_B.xlsx"
BS_XLSX = "B_hard/AutostaffChubu_Model_Sponsor_B.xlsx"
DEBT_B = "B_hard/AutostaffChubu_Debt_B.xlsx"
A_PDF = "A_standard/DD_Report_統合版_オートスタッフ中部_A.pdf"
FIN_PDF = "B_hard/DD_Financial_オートスタッフ中部.pdf"
BIZ_PDF = "B_hard/DD_Business_オートスタッフ中部.pdf"
LEG_PDF = "B_hard/DD_Legal_オートスタッフ中部.pdf"
TAX_PDF = "B_hard/DD_Tax_オートスタッフ中部.pdf"


def _facts() -> list[dict]:
    F = []

    def add(key, section, name, value, unit, src_a, src_b, applies="両方", deriv=None):
        F.append(dict(key=key, section=section, name=name, value=value, unit=unit,
                      src_a=src_a, src_b=src_b, applies=applies, deriv=deriv or ""))

    # ---- ストラクチャー（Assumptions KV：D列）
    asm = "Assumptions"
    add("ev", "ストラクチャー", "エンタープライズ・バリュー（EV）", 12_000_000, "千円",
        f"{A_XLSX}|{asm}!D13", f"{B_XLSX}|{asm}!D13")
    add("equity", "ストラクチャー", "エクイティ出資", 5_500_000, "千円",
        f"{A_XLSX}|{asm}!D14", f"{B_XLSX}|{asm}!D14")
    add("senior_loan", "ストラクチャー", "シニアローン総額", 6_500_000, "千円",
        f"{A_XLSX}|{asm}!D15", f"{B_XLSX}|{asm}!D15")
    add("bank_a_commitment", "ストラクチャー", "参加金融機関A行 想定取組額", 2_500_000, "千円",
        f"{A_XLSX}|{asm}!D16", f"{B_XLSX}|{asm}!D16")
    add("tenor_years", "ストラクチャー", "ローン期間", 7, "年",
        f"{A_XLSX}|{asm}!D17", f"{B_XLSX}|{asm}!D17")
    add("interest_rate", "ストラクチャー", "適用金利（TIBOR+200bp想定）", 0.022, "－",
        f"{A_XLSX}|{asm}!D18", f"{B_XLSX}|{asm}!D18")
    add("annual_repayment", "ストラクチャー", "約定弁済（年）", 650_000, "千円",
        f"{A_XLSX}|{asm}!D19", f"{B_XLSX}|{asm}!D19")
    add("goodwill_model", "ストラクチャー", "のれん想定額（モデル・暫定PPA前）", S.GOODWILL, "千円",
        f"{A_XLSX}|{asm}!D21", f"{B_XLSX}|{asm}!D21")
    add("ltv", "ストラクチャー", "LTV（シニア÷EV）", 6_500_000 / 12_000_000, "－",
        f"{A_XLSX}|導出", f"{B_XLSX}|導出", deriv="senior_loan/ev")
    add("leverage_fy26", "ストラクチャー", "レバレッジ（シニア÷FY26 EBITDA）",
        6_500_000 / 1_620_000, "倍", f"{A_XLSX}|導出", f"{B_XLSX}|導出",
        deriv="senior_loan/ebitda_fy26")
    add("goodwill_dd", "ストラクチャー", "のれん想定額（財務DD試算）", 5_280, "百万円",
        "－（Aはのれん差異なし）", f"{FIN_PDF}|p.31", applies="B")
    add("teaser_ebitda", "ストラクチャー", "スポンサー提示EBITDA（ティーザー・速報）", 1_585_000, "千円",
        "－（Aはティーザー行なし）", f"{B_XLSX}|{asm}!D{L.ASM_TEASER_ROW}", applies="B")

    # ---- KPI（FY26：G列）
    kpi = "KPI_Drivers"
    d26 = _drv["FY26"]
    add("enrolled_fy26", "KPI（FY26）", "在籍登録スタッフ数", d26["enrolled"], "名",
        f"{A_XLSX}|{kpi}!G7", f"{B_XLSX}|{kpi}!G7")
    add("util_fy26", "KPI（FY26）", "稼働率", d26["util"], "－",
        f"{A_XLSX}|{kpi}!G8", f"{B_XLSX}|{kpi}!G8")
    add("active_fy26", "KPI（FY26）", "稼働人数", d26["active"], "名",
        f"{A_XLSX}|{kpi}!G9", f"{B_XLSX}|{kpi}!G9",
        deriv="round(enrolled_fy26*util_fy26)")
    add("hours_fy26", "KPI（FY26）", "月間平均稼働時間", d26["hours"], "h/名",
        f"{A_XLSX}|{kpi}!G10", f"{B_XLSX}|{kpi}!G10")
    add("bill_fy26", "KPI（FY26）", "派遣単価", d26["bill"], "円/h",
        f"{A_XLSX}|{kpi}!G11", f"{B_XLSX}|{kpi}!G11")
    add("wage_fy26", "KPI（FY26）", "スタッフ平均時給", d26["wage"], "円/h",
        f"{A_XLSX}|{kpi}!G12", f"{B_XLSX}|{kpi}!G12")
    add("welfare_rate", "KPI（FY26）", "法定福利費率", d26["welfare"], "－",
        f"{A_XLSX}|{kpi}!G13", f"{B_XLSX}|{kpi}!G13")
    add("hires_fy26", "KPI（FY26）", "新規採用人数", d26["hires"], "名",
        f"{A_XLSX}|{kpi}!G14", f"{B_XLSX}|{kpi}!G14")
    add("cpa_fy26", "KPI（FY26）", "採用単価（CPA）", d26["cpa"], "千円/名",
        f"{A_XLSX}|{kpi}!G15", f"{B_XLSX}|{kpi}!G15")
    add("attrition_fy26", "KPI（FY26）", "離職率（年間）", d26["attrition"], "－",
        f"{A_XLSX}|{kpi}!G16", f"{B_XLSX}|{kpi}!G16")
    add("spread_fy26", "KPI（FY26）", "スプレッド（福利費考慮後）",
        d26["bill"] - d26["wage"] * 1.14, "円/h", "導出", "導出",
        deriv="bill_fy26 - wage_fy26*(1+welfare_rate)")

    # ---- PL（FY26：G列）
    p26 = _pl["FY26"]
    add("staffing_rev_fy26", "PL（FY26）", "派遣事業売上", p26["staffing_rev"], "千円",
        f"{A_XLSX}|PL!G7", f"{B_XLSX}|PL!G7",
        deriv="round(active_fy26*hours_fy26*bill_fy26*12/1000)")
    add("other_revenue_fy26", "PL（FY26）", "その他営業収入", p26["other_revenue"], "千円",
        f"{A_XLSX}|PL!G8", f"{B_XLSX}|PL!G8")
    add("revenue_fy26", "PL（FY26）", "売上高", p26["revenue"], "千円",
        f"{A_XLSX}|PL!G9", f"{B_XLSX}|PL!G9（B表記は「売上収益」）",
        deriv="staffing_rev_fy26 + other_revenue_fy26")
    add("labor_fy26", "PL（FY26）", "スタッフ労務費（法定福利費込）", p26["labor"], "千円",
        f"{A_XLSX}|PL!G11", f"{B_XLSX}|PL!G11（B表記は「現業人件費（福利込）」）",
        deriv="round(active_fy26*hours_fy26*wage_fy26*(1+welfare_rate)*12/1000)")
    add("other_cogs_fy26", "PL（FY26）", "その他売上原価", p26["other_cogs"], "千円",
        f"{A_XLSX}|PL!G12", f"{B_XLSX}|PL!G12")
    add("cogs_fy26", "PL（FY26）", "売上原価合計", p26["cogs"], "千円",
        f"{A_XLSX}|PL!G13", f"{B_XLSX}|PL!G13", deriv="labor_fy26 + other_cogs_fy26")
    add("gross_fy26", "PL（FY26）", "売上総利益", p26["gross"], "千円",
        f"{A_XLSX}|PL!G14", f"{B_XLSX}|PL!G14", deriv="revenue_fy26 - cogs_fy26")
    add("recruiting_fy26", "PL（FY26）", "採用費", p26["recruiting"], "千円",
        f"{A_XLSX}|PL!G17", f"{B_XLSX}|PL!G17", deriv="hires_fy26*cpa_fy26")
    add("hq_cost_fy26", "PL（FY26）", "本社人件費", p26["hq_cost"], "千円",
        f"{A_XLSX}|PL!G18", f"{B_XLSX}|PL!G18")
    add("other_sga_fy26", "PL（FY26）", "その他販管費", p26["other_sga"], "千円",
        f"{A_XLSX}|PL!G19", f"{B_XLSX}|PL!G19")
    add("sga_fy26", "PL（FY26）", "販管費合計", p26["sga"], "千円",
        f"{A_XLSX}|PL!G20", f"{B_XLSX}|PL!G20",
        deriv="recruiting_fy26 + hq_cost_fy26 + other_sga_fy26")
    add("op_fy26", "PL（FY26）", "営業利益", p26["op"], "千円",
        f"{A_XLSX}|PL!G21", f"{B_XLSX}|PL!G21", deriv="gross_fy26 - sga_fy26")
    add("depreciation_fy26", "PL（FY26）", "減価償却費", p26["depreciation"], "千円",
        f"{A_XLSX}|PL!G24", f"{B_XLSX}|PL!G24")
    add("ebitda_fy26", "PL（FY26）", "EBITDA", p26["ebitda"], "千円",
        f"{A_XLSX}|PL!G25", f"{B_XLSX}|PL!G25（B表記は「修正EBITDA」）",
        deriv="op_fy26 + depreciation_fy26")
    add("ni_fy26", "PL（FY26）", "当期純利益", p26["ni"], "千円",
        f"{A_XLSX}|PL!G32", f"{B_XLSX}|PL!G32")
    add("normalized_ebitda_fy26", "PL（FY26）", "正常収益力EBITDA（DD評価）", 1_620_000, "千円",
        f"{A_PDF}|6.9節（図表・結論）", f"{FIN_PDF}|p.34（単位は百万円で1,620）",
        deriv="op_fy26 + depreciation_fy26")

    # ---- BS（FY26：G列）
    b26 = _bs["FY26"]
    add("cash_fy26", "BS（FY26）", "現金及び預金", b26["cash"], "千円",
        f"{A_XLSX}|BS!G7", f"{B_XLSX}|BS!G7")
    add("ar_fy26", "BS（FY26）", "売上債権", b26["ar"], "千円",
        f"{A_XLSX}|BS!G8", f"{B_XLSX}|BS!G8（B表記は「売掛金等」）")
    add("ap_fy26", "BS（FY26）", "仕入債務・未払費用", b26["ap"], "千円",
        f"{A_XLSX}|BS!G18", f"{B_XLSX}|BS!G18（B表記は「買掛・未払等」）")
    add("debt_fy26", "BS（FY26）", "有利子負債", b26["debt"], "千円",
        f"{A_XLSX}|BS!G20", f"{B_XLSX}|BS!G20")
    add("net_assets_fy26", "BS（FY26）", "純資産", b26["net_assets"], "千円",
        f"{A_XLSX}|BS!G23", f"{B_XLSX}|BS!G23")
    add("net_cash_fy26", "BS（FY26）", "ネットキャッシュ", b26["cash"] - b26["debt"], "千円",
        "導出", "導出", deriv="cash_fy26 - debt_fy26")

    # ---- 新規作成期（FY22〜23実績・FY32〜33計画）
    for y, col in (("FY22", "C"), ("FY23", "D")):
        add(f"revenue_{y.lower()}", f"実績（{y}・新規作成）", "売上高", _pl[y]["revenue"], "千円",
            f"{A_XLSX}|PL!{col}9", f"{B_XLSX}|PL!{col}9")
        add(f"ebitda_{y.lower()}", f"実績（{y}・新規作成）", "EBITDA", _pl[y]["ebitda"], "千円",
            f"{A_XLSX}|PL!{col}25", f"{B_XLSX}|PL!{col}25")
        add(f"ni_{y.lower()}", f"実績（{y}・新規作成）", "当期純利益", _pl[y]["ni"], "千円",
            f"{A_XLSX}|PL!{col}32", f"{B_XLSX}|PL!{col}32")
    add("rev_cagr_fy22_26", "実績（新規作成）", "売上CAGR（FY22→FY26）",
        (_pl["FY26"]["revenue"] / _pl["FY22"]["revenue"]) ** 0.25 - 1, "－",
        "導出", "導出", deriv="(revenue_fy26/revenue_fy22)**0.25 - 1")
    for y, col in (("FY32", "M"), ("FY33", "N")):
        add(f"revenue_{y.lower()}_base", f"計画（{y}・新規作成）", "売上高（Base）",
            _pl[y]["revenue"], "千円", f"{A_XLSX}|PL!{col}9", f"{B_XLSX}|PL!{col}9")
        add(f"ebitda_{y.lower()}_base", f"計画（{y}・新規作成）", "EBITDA（Base）",
            _pl[y]["ebitda"], "千円", f"{A_XLSX}|PL!{col}25", f"{B_XLSX}|PL!{col}25")
        add(f"bill_{y.lower()}_base", f"計画（{y}・新規作成）", "派遣単価（Base）",
            _drv[y]["bill"], "円/h", f"{A_XLSX}|KPI_Drivers!{col}11", f"{B_XLSX}|KPI_Drivers!{col}11")
        add(f"util_{y.lower()}_base", f"計画（{y}・新規作成）", "稼働率（Base）",
            _drv[y]["util"], "－", f"{A_XLSX}|KPI_Drivers!{col}8", f"{B_XLSX}|KPI_Drivers!{col}8")

    # ---- デットスケジュール
    add("interest_fy27", "デット", "支払利息（FY27）", _sched["FY27"]["interest"], "千円",
        f"{A_XLSX}|Debt_Schedule!H12", f"{DEBT_B}|Debt_Schedule!H12",
        deriv="round((senior_loan + (senior_loan - annual_repayment))/2 * interest_rate)")
    add("debt_service_fy27", "デット", "デットサービス（FY27・約定弁済＋利息）",
        _sched["FY27"]["repayment"] + _sched["FY27"]["interest"], "千円",
        "導出", "導出", deriv="annual_repayment + interest_fy27")
    add("debt_closing_fy31", "デット", "シニアローン期末残高（FY31）",
        _sched["FY31"]["closing"], "千円",
        f"{A_XLSX}|Debt_Schedule!L9", f"{DEBT_B}|Debt_Schedule!L9",
        deriv="senior_loan - annual_repayment*5")
    add("bullet_fy33", "デット", "FY33残額一括弁済", _sched["FY33"]["repayment"], "千円",
        f"{A_XLSX}|Debt_Schedule!N8", f"{DEBT_B}|Debt_Schedule!N8",
        deriv="senior_loan - annual_repayment*6")
    add("debt_closing_fy33", "デット", "シニアローン期末残高（FY33）", 0, "千円",
        f"{A_XLSX}|Debt_Schedule!N9", f"{DEBT_B}|Debt_Schedule!N9")

    # ---- スポンサーケース
    add("sponsor_revenue_fy27", "計画（Sponsor）", "売上高（FY27）",
        _pls["FY27"]["revenue"], "千円", "－（AはBaseのみ）", f"{BS_XLSX}|PL!H9", applies="B")
    add("sponsor_ebitda_fy31", "計画（Sponsor）", "EBITDA（FY31）",
        _pls["FY31"]["ebitda"], "千円", "－（AはBaseのみ）", f"{BS_XLSX}|PL!L25", applies="B")
    add("sponsor_ebitda_fy33", "計画（Sponsor）", "EBITDA（FY33）",
        _pls["FY33"]["ebitda"], "千円", "－（AはBaseのみ）", f"{BS_XLSX}|PL!N25", applies="B")

    # ---- 月次試算表（FY26）
    add("monthly_rev_total", "月次（FY26）", "月次売上12ヶ月合計", sum(_mon["revenue"].values()), "円",
        f"{A_XLSX}|月次試算表_FY26!C7:N7", f"{B_XLSX}|月次試算表_FY26!C7:N7",
        deriv="revenue_fy26*1000")
    add("monthly_rev_may", "月次（FY26）", "2025年5月売上", _mon["revenue"]["2025/05"], "円",
        f"{A_XLSX}|月次試算表_FY26!D7", f"{B_XLSX}|月次試算表_FY26!D7")
    add("monthly_cash_mar", "月次（FY26）", "2026年3月末 現金及び預金",
        _mon["cash"]["2026/03"], "円",
        f"{A_XLSX}|月次試算表_FY26!N13", f"{B_XLSX}|月次試算表_FY26!N13",
        deriv="cash_fy26*1000")
    add("monthly_unit", "月次（FY26）", "月次試算表の単位", "円（Aはラベル明記・Bはラベルなし）", "－",
        f"{A_XLSX}|月次試算表_FY26!A2", f"{B_XLSX}|月次試算表_FY26（ラベル0件）")

    # ---- DD定性ファクト
    add("customer_top3_share", "DD定性", "上位3社への売上依存度", 0.62, "－",
        f"{A_PDF}|5.2節（発見事項5-2）", f"{BIZ_PDF}|p.18")
    add("customer_a_share", "DD定性", "最大手A社の売上構成比", 0.28, "－",
        f"{A_PDF}|5.2節", f"{BIZ_PDF}|p.18")
    add("overtime_liability_max", "DD定性（簿外リスク）", "未払残業代の潜在債務（最大）", 300, "百万円",
        "－（A統合版DDには含めない）", f"{LEG_PDF}|p.12", applies="B")
    add("haken_license", "DD定性", "労働者派遣事業許可番号", "派23-301456",
        "－", "－（A統合版DDには含めない）", f"{LEG_PDF}|p.5", applies="B")
    add("nol_status", "DD定性", "繰越欠損金", "存在しない（直近10年間課税所得を計上）", "－",
        "－（A統合版DDには含めない）", f"{TAX_PDF}|p.6", applies="B")
    add("dorm_rent_annual", "DD定性", "借上げ寮 年間賃料", 210_000, "千円",
        f"{A_PDF}|4.7節", f"{B_XLSX}|BS!P7:Q7（寮関連メモ）")
    return F


def _kpi_structure() -> list[list]:
    return [
        ["ノード", "親ノード", "階層", "算式・定義"],
        ["売上高", "－", 1, "派遣事業売上 ＋ その他営業収入"],
        ["派遣事業売上", "売上高", 2, "稼働人数 × 月間平均稼働時間 × 派遣単価 × 12 ÷ 1,000（千円）"],
        ["その他営業収入", "売上高", 2, "紹介手数料等（実績値）"],
        ["稼働人数", "派遣事業売上", 3, "ROUND(在籍登録スタッフ数 × 稼働率, 0)"],
        ["在籍登録スタッフ数", "稼働人数", 4, "前期在籍 ＋ 新規採用人数 − 離職（離職率×在籍）"],
        ["稼働率", "稼働人数", 4, "就業中スタッフ ÷ 在籍登録スタッフ（マッチング効率）"],
        ["派遣単価", "派遣事業売上", 3, "毎年4月の単価改定交渉で決定（実績+1.8〜2.0%/年、計画+1.0%/年前後）"],
        ["月間平均稼働時間", "派遣事業売上", 3, "顧客工場の稼働カレンダー準拠（165〜167h）"],
        ["売上原価", "－", 1, "スタッフ労務費 ＋ その他売上原価"],
        ["スタッフ労務費", "売上原価", 2, "稼働人数 × 稼働時間 × 平均時給 × (1＋法定福利費率14%) × 12 ÷ 1,000"],
        ["売上総利益", "－", 1, "売上高 − 売上原価"],
        ["スプレッド", "売上総利益", 2, "派遣単価 − 平均時給 × (1＋法定福利費率)"],
        ["販管費", "－", 1, "採用費 ＋ 本社人件費 ＋ その他販管費"],
        ["採用費", "販管費", 2, "新規採用人数 × 採用単価（CPA）"],
        ["営業利益", "－", 1, "売上総利益 − 販管費"],
        ["EBITDA", "－", 1, "営業利益 ＋ 減価償却費"],
        ["支払利息", "－", 2, "デットスケジュールの平均残高 × 金利2.2%（ROUND）"],
        ["当期純利益", "－", 1, "（営業利益 − 支払利息 ＋ 営業外損益）× (1 − 実効税率31%)"],
    ]


def _name_mapping() -> list[list]:
    rows = [["B改称科目（Excel B）", "DD側の呼称", "A標準科目（Excel A）", "出現箇所（B）", "備考"]]
    for sheet, key, old, new, dd in L.RENAMES:
        row_no = {L.SHEET_PL: L.PL_ROWS, L.SHEET_BS: L.BS_ROWS}[sheet][key][0]
        rows.append([new, dd, old, f"{sheet}!B{row_no}", "PL_old・Scratchは旧表記のまま"])
    rows.append(["売掛金等（月次）", "売上債権", "売上債権（月次）", "月次試算表_FY26!B14", "月次試算表も同一改称"])
    rows.append(["買掛・未払等（月次）", "仕入債務・未払費用", "仕入債務・未払費用（月次）",
                 "月次試算表_FY26!B15", "同上"])
    rows.append(["売上収益（月次）", "売上高", "売上高（月次）", "月次試算表_FY26!B7", "同上"])
    return rows


def _monthly_tie() -> list[list]:
    rows = [["月", "売上高（円）", "売上原価（円）", "販管費（円）", "営業利益（円）",
             "現金及び預金（円）", "売上債権（円）", "仕入債務等（円）"]]
    for m in S.MONTHS:
        rows.append([m, _mon["revenue"][m], _mon["cogs"][m], _mon["sga"][m], _mon["op"][m],
                     _mon["cash"][m], _mon["ar"][m], _mon["ap"][m]])
    p26 = _pl["FY26"]
    b26 = _bs["FY26"]
    rows.append(["12ヶ月合計（円）", sum(_mon["revenue"].values()), sum(_mon["cogs"].values()),
                 sum(_mon["sga"].values()), sum(_mon["op"].values()), "－（残高）", "－（残高）", "－（残高）"])
    rows.append(["年次FY26（千円）", p26["revenue"], p26["cogs"], p26["sga"], p26["op"],
                 b26["cash"], b26["ar"], b26["ap"]])
    rows.append(["突合差異（千円）",
                 sum(_mon["revenue"].values()) // 1000 - p26["revenue"],
                 sum(_mon["cogs"].values()) // 1000 - p26["cogs"],
                 sum(_mon["sga"].values()) // 1000 - p26["sga"],
                 sum(_mon["op"].values()) // 1000 - p26["op"],
                 _mon["cash"]["2026/03"] // 1000 - b26["cash"],
                 _mon["ar"]["2026/03"] // 1000 - b26["ar"],
                 _mon["ap"]["2026/03"] // 1000 - b26["ap"]])
    rows.append(["季節性（対他9ヶ月平均）",
                 f"2025/05: △{_season['2025/05'] * 100:.1f}%",
                 f"2025/08: △{_season['2025/08'] * 100:.1f}%",
                 f"2026/01: △{_season['2026/01'] * 100:.1f}%", "", "", "", ""])
    return rows


def _planted() -> list[list]:
    items = []
    for f in ("planted_items_models.json", "planted_items_dd.json"):
        items.extend(json.loads((BUILD / f).read_text(encoding="utf-8")))
    rows = [["複雑性ID", "データセット", "ファイル", "シート・ページ", "セル・範囲", "内容", "期待動作"]]
    for it in items:
        rows.append([it["id"], it["dataset"], it["file"], it["sheet"], it["cell"],
                     it["desc"], it["expect"]])
    return rows


def _write_sheet(wb, name, rows, widths):
    ws = wb.create_sheet(name)
    for i, row in enumerate(rows, start=1):
        for j, v in enumerate(row, start=1):
            c = ws.cell(row=i, column=j, value=v)
            c.font = HEADER_FONT if i == 1 else BASE_FONT
            if i == 1:
                c.fill = HEADER_FILL
            c.alignment = Alignment(vertical="top", wrap_text=(j >= 6 or name == "planted_items"))
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"


def main():
    OUT.parent.mkdir(exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)

    facts = _facts()
    rows = [["key", "セクション", "項目名", "正解値", "単位",
             "出所（データA）", "出所（データB）", "適用データセット", "導出式（review.py再計算用）"]]
    for f in facts:
        rows.append([f["key"], f["section"], f["name"], f["value"], f["unit"],
                     f["src_a"], f["src_b"], f["applies"], f["deriv"]])
    _write_sheet(wb, "facts", rows,
                 {"A": 24, "B": 18, "C": 30, "D": 16, "E": 8, "F": 44, "G": 44, "H": 12, "I": 40})

    _write_sheet(wb, "kpi_structure", _kpi_structure(), {"A": 22, "B": 16, "C": 8, "D": 64})
    _write_sheet(wb, "name_mapping", _name_mapping(), {"A": 26, "B": 26, "C": 30, "D": 24, "E": 28})
    _write_sheet(wb, "planted_items", _planted(),
                 {"A": 10, "B": 12, "C": 44, "D": 22, "E": 14, "F": 64, "G": 18})
    _write_sheet(wb, "monthly_tie", _monthly_tie(),
                 {"A": 16, "B": 16, "C": 16, "D": 16, "E": 16, "F": 16, "G": 16, "H": 16})

    wb.save(OUT)
    print(f"generated: {OUT.relative_to(ROOT.parent)}  (facts: {len(facts)}行)")


if __name__ == "__main__":
    main()
