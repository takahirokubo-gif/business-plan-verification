# -*- coding: utf-8 -*-
"""ダミーインプット一式（v2・ルミエールボーテ）の整合チェック。

検証対象:
 1. spec.py の主要ピン（正本数値）
 2. Excelモデル：ピンセルの値・数式・キャッシュ値（実AIが値を拾えること）・単位注記
 3. Excelモデルが現行の解析パイプライン制約（60行×12列のダイジェスト窓）に収まること
 4. DDレポートPDF：キーファクトのページ位置とページ数（pypdfで抽出＝実AIと同経路）
 5. expected_output.json：期待値が spec の計算値と一致すること
 6. GROUND_TRUTH.md が spec から再生成した内容と一致すること
"""
import json
import sys
from pathlib import Path

from openpyxl import load_workbook
from pypdf import PdfReader

import generate_expected
import spec
import xl_layout as L
from spec import ACTUAL_YEARS, COL, PLAN_YEARS

HERE = Path(__file__).parent

# 現行 AnthropicExtractor._excel_digest の読み取り窓
DIGEST_MAX_ROW = 60
DIGEST_MAX_COL = 12

errors = []


def check(cond, msg):
    if cond:
        print(f"  OK  {msg}")
    else:
        errors.append(msg)
        print(f"  NG  {msg}")


def validate_spec():
    print("[1] spec.py 主要ピン")
    spec.summary_check()
    print("  OK  summary_check（9,240/1,120・9,595/1,191・4.3x/53%・KPI・BSバランス）")


def validate_excel():
    print("[2] Excelモデル")
    expected_order = [L.SHEET_COVER, L.SHEET_INP, L.SHEET_PL, L.SHEET_BS,
                      L.SHEET_CF, L.SHEET_DEBT, L.SHEET_MEMO]
    for case in spec.CASES:
        path = HERE / spec.MODEL_FILES[case]
        check(path.exists(), f"{path.name} が存在する")
        wb = load_workbook(path)          # 数式
        wbv = load_workbook(path, data_only=True)  # キャッシュ値
        check(wb.sheetnames == expected_order,
              f"{case}: シート構成が正しい {wb.sheetnames}")
        pl = spec.compute_pl(case)
        bs, cf = spec.compute_bs_cf(case)
        drv = spec.drivers(case)
        ws = wb[L.SHEET_PL]
        wsv = wbv[L.SHEET_PL]
        check("百万円" in str(ws["A2"].value), f"{case}: PLシートに単位注記（百万円）")
        c26 = COL["FY26"]
        c27 = COL["FY27"]
        check(ws[f"{c26}{L.PL_ROWS['ec_rev'][0]}"].value == pl["FY26"]["ec_rev"],
              f"{case}: PL FY26 EC売上（値貼付）= {pl['FY26']['ec_rev']:,}百万円")
        f = ws[f"{c27}{L.PL_ROWS['ec_rev'][0]}"].value
        check(isinstance(f, str) and "inp" in f and "/10^6" in f,
              f"{case}: PL FY27 EC売上は inp 参照の数式（{f}）")
        f = ws[f"{c27}{L.PL_ROWS['cogs'][0]}"].value
        check(isinstance(f, str) and "inp" in f,
              f"{case}: PL FY27 売上原価は原価率参照の数式")
        # キャッシュ値（実AIが拾う値）が spec と一致
        for y in spec.YEARS:
            got = wsv[f"{COL[y]}{L.PL_ROWS['ebitda'][0]}"].value
            check(got == pl[y]["ebitda"],
                  f"{case}: PL {y} EBITDAキャッシュ値 = {pl[y]['ebitda']:,}（{got}）")
        got = wbv[L.SHEET_CF][f"{c27}{L.CF_ROWS['fcf'][0]}"].value
        check(got == cf["FY27"]["fcf"],
              f"{case}: CF FY27 FCFキャッシュ値 = {cf['FY27']['fcf']:,}（{got}）")
        # inp（ストラクチャー・ドライバー）
        ws_i = wb[L.SHEET_INP]
        ws_iv = wbv[L.SHEET_INP]
        check(ws_i[f"D{L.INP_STRUCT_ROWS['ev'][0]}"].value == spec.DEAL["ev_mm"],
              f"{case}: inp EV = {spec.DEAL['ev_mm']:,}百万円")
        check(ws_i[f"D{L.INP_STRUCT_ROWS['senior'][0]}"].value == spec.SENIOR_TOTAL,
              f"{case}: inp シニア = {spec.SENIOR_TOTAL:,}百万円")
        check(ws_i[f"D{L.INP_STRUCT_ROWS['goodwill'][0]}"].value == spec.GOODWILL,
              f"{case}: inp のれん想定 = {spec.GOODWILL:,}百万円")
        check(ws_i[f"D{L.INP_STRUCT_ROWS['sponsor_ebitda'][0]}"].value
              == spec.DEAL["sponsor_ebitda_mm"],
              f"{case}: inp 提示EBITDA = {spec.DEAL['sponsor_ebitda_mm']:,}百万円")
        check(str(ws_i[f"D{L.INP_STRUCT_ROWS['borrower'][0]}"].value) == spec.DEAL["borrower"],
              f"{case}: inp 借入人SPC = {spec.DEAL['borrower']}")
        f = ws_i[f"{c27}{L.INP_DRIVER_ROWS['member'][0]}"].value
        check(isinstance(f, str) and f.startswith("=ROUND(") and "*" in f,
              f"{case}: inp 会員数（計画）は =ROUND(前期×リピート率＋新規) の数式")
        got = ws_iv[f"{c27}{L.INP_DRIVER_ROWS['member'][0]}"].value
        check(got == drv["FY27"]["member"],
              f"{case}: inp FY27会員数キャッシュ値 = {drv['FY27']['member']:,}（{got}）")
        check(ws_i[f"{c26}{L.INP_DRIVER_ROWS['repeat'][0]}"].value == drv["FY26"]["repeat"],
              f"{case}: inp FY26リピート率 = {drv['FY26']['repeat']:.0%}")
        # BSのれん（計画年）
        ws_b = wbv[L.SHEET_BS]
        check(ws_b[f"{c27}{L.BS_ROWS['goodwill'][0]}"].value == spec.GOODWILL,
              f"{case}: BS FY27のれん = {spec.GOODWILL:,}百万円")
        # Cover
        cover = wb[L.SHEET_COVER]
        check(spec.CASE_LABEL[case] in str(cover["B8"].value),
              f"{case}: Coverにケース名（{spec.CASE_LABEL[case]}）")
        check(spec.DEAL["borrower"] in str(cover["D12"].value),
              f"{case}: Coverに借入人SPC")
        # ダイジェスト窓（60行×12列）にデータが収まる
        for name in (L.SHEET_INP, L.SHEET_PL, L.SHEET_BS, L.SHEET_CF, L.SHEET_DEBT):
            s = wb[name]
            over = []
            for row in s.iter_rows():
                for c in row:
                    if c.value is not None and (c.row > DIGEST_MAX_ROW
                                                or c.column > DIGEST_MAX_COL):
                        over.append(c.coordinate)
            check(not over,
                  f"{case}/{name}: データが解析窓（{DIGEST_MAX_ROW}行×{DIGEST_MAX_COL}列）内"
                  + (f"（超過: {over[:5]}）" if over else ""))


def validate_pdfs():
    print("[3] DDレポートPDF")
    expected_pages = {"business": 20, "financial": 34, "legal": 14, "tax": 10}
    readers = {}
    for key, fname in spec.DD_FILES.items():
        path = HERE / fname
        check(path.exists(), f"{fname} が存在する")
        readers[key] = PdfReader(path)
        check(len(readers[key].pages) == expected_pages[key],
              f"{fname}: {expected_pages[key]}ページ構成")
    for fact in spec.DD_KEY_FACTS:
        r = readers[fact["file"]]
        text = (r.pages[fact["page"] - 1].extract_text() or "").replace("\n", "")
        for phrase in fact["check"]:
            check(phrase in text,
                  f"{spec.DD_FILES[fact['file']]} p.{fact['page']} に「{phrase[:28]}…」")
    # 主要キーファクトの結論文言が指定ページ以外に重複して存在しないこと
    for fact_id in ("normalized_ebitda", "goodwill_dd"):
        fact = next(f for f in spec.DD_KEY_FACTS if f["id"] == fact_id)
        r = readers[fact["file"]]
        hits = [i + 1 for i, p in enumerate(r.pages)
                if fact["check"][0] in (p.extract_text() or "").replace("\n", "")]
        check(hits == [fact["page"]],
              f"{fact_id} の結論文言は p.{fact['page']} のみに存在（{hits}）")
    # 千円表とモデル百万円の丸め整合（財務DD p.11 の売上高）
    text = (readers["financial"].pages[10].extract_text() or "").replace(",", "")
    pl = spec.compute_pl("base")
    check(str(pl["FY26"]["revenue"] * 1000) in text,
          "財務DD p.11 の売上高（千円）がモデル（百万円）×1000と一致")


def validate_expected():
    print("[4] expected_output.json")
    exp = json.loads((HERE / "expected_output.json").read_text(encoding="utf-8"))
    pl_b = spec.compute_pl("base")
    pl_s = spec.compute_pl("sponsor")
    _, cf_b = spec.compute_bs_cf("base")
    bs_b, _ = spec.compute_bs_cf("base")
    items = {it["key"]: it for it in exp["items"]}
    check(len(exp["items"]) == 24, f"抽出項目は24件（{len(exp['items'])}件）")
    check(items["act_revenue"]["values"]
          == {y: pl_b[y]["revenue"] for y in ACTUAL_YEARS},
          "act_revenue = 実績売上（7,788/8,500/9,240）")
    check(items["act_ebitda"]["values"]["FY26"] == 1120, "act_ebitda FY26 = 1,120")
    check(items["normalized_ebitda"]["values"]["FY26"] == spec.NORMALIZED_EBITDA,
          f"normalized_ebitda = {spec.NORMALIZED_EBITDA:,}")
    check(items["base_revenue"]["values"]["FY27"] == 9595, "base_revenue FY27 = 9,595")
    check(items["base_ebitda"]["values"]["FY27"] == 1191, "base_ebitda FY27 = 1,191")
    check(items["sponsor_ebitda"]["values"]
          == {y: pl_s[y]["ebitda"] for y in PLAN_YEARS},
          "sponsor_ebitda = スポンサー計画EBITDA")
    check(items["base_fcf"]["values"] == {y: cf_b[y]["fcf"] for y in PLAN_YEARS},
          "base_fcf = ベースFCF")
    check(items["act_cash"]["values"] == {y: bs_b[y]["cash"] for y in ACTUAL_YEARS},
          "act_cash = 実績現預金")
    gw = items["goodwill"]
    check(gw["values"]["FY27"] == spec.GOODWILL
          and gw["mismatch"]["other_value"] == spec.GOODWILL_DD,
          f"goodwill: モデル{spec.GOODWILL:,} vs DD{spec.GOODWILL_DD:,} の不整合データ")
    all_files = set(spec.DD_FILES.values()) | set(spec.MODEL_FILES.values())
    for it in exp["items"]:
        check(it["evidence"]["file"] in all_files,
              f"{it['key']}: 参照ファイル {it['evidence']['file']} が実在する")
    facts = {f["id"]: f for f in spec.DD_KEY_FACTS}
    check(items["risk_oem"]["evidence"]["page"] == facts["oem_dependency"]["page"],
          "risk_oem の参照ページがキーファクト定義と一致")
    check(items["normalized_ebitda"]["evidence"]["page"]
          == facts["normalized_ebitda"]["page"],
          "normalized_ebitda の参照ページがキーファクト定義と一致")
    check(gw["mismatch"]["other_page"] == facts["goodwill_dd"]["page"],
          "goodwill不整合の参照ページがキーファクト定義と一致")
    check(set(exp["identify"].keys()) == all_files, "identify: 6ファイルすべてに対応")
    di = exp["deal_info"]["fields"]
    check(di["ev_mm"] == spec.DEAL["ev_mm"]
          and di["sponsor_ebitda_mm"] == spec.DEAL["sponsor_ebitda_mm"],
          "deal_info: EV・提示EBITDAが正本と一致")
    check(exp["auto_calc"]["initial_leverage"] == 4.3
          and exp["auto_calc"]["ltv_pct"] == 53,
          "auto_calc: レバレッジ4.3x・LTV53%")


def validate_ground_truth():
    print("[5] GROUND_TRUTH.md")
    current = (HERE / "GROUND_TRUTH.md").read_text(encoding="utf-8")
    regenerated = generate_expected.build_ground_truth(generate_expected.build_expected())
    check(current == regenerated, "GROUND_TRUTH.md は spec からの再生成と一致（手編集なし）")


def main():
    validate_spec()
    validate_excel()
    validate_pdfs()
    validate_expected()
    validate_ground_truth()
    print()
    if errors:
        print(f"NG: {len(errors)}件の不整合")
        for e in errors:
            print(f"  - {e}")
        sys.exit(1)
    print("すべての整合チェックに合格")


if __name__ == "__main__":
    main()
