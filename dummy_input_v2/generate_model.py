"""財務モデルExcel（Sponsor / Bank Base 2ケース）の生成（v2・ルミエールボーテ）。

実在のLBOモデルテンプレートの慣行に寄せる:
- 金額は百万円（Coverと各シートに単位を明記）
- シート構成は Cover / inp（前提・ドライバー）/ PL / BS / CF / debt / memo
- 実績年（2024/3期〜2026/3期）は値貼り付け、計画年（2027/3期〜2031/3期）は実数式
  （KPIツリーはこの数式パースで再現できることが要件）
- 生成後に LibreOffice で再計算・保存し、数式セルにキャッシュ値を持たせる。
  実務でアップロードされるExcel（Excel保存済み＝計算結果キャッシュあり）と同じ状態にする
- 項目名の表記揺れ（売上収益 / EBIT / AOV等）を意図的に含む
- ノイズは軽微（memoシートのみ）。重ノイズ（旧版シート・大量の作業セル）は
  既存デモ側でカバー済みのため、本データセットは「理想的なインプット」に徹する
"""
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

import spec
from spec import ACTUAL_YEARS, CASE_LABEL, COL, PLAN_YEARS, YEAR_LABEL, YEARS
import xl_layout as L

OUT_DIR = Path(__file__).parent

TITLE_FONT = Font(name="Yu Gothic", size=14, bold=True, color="1A4F8B")
HEADER_FONT = Font(name="Yu Gothic", size=10, bold=True)
BASE_FONT = Font(name="Yu Gothic", size=10)
INPUT_FONT = Font(name="Yu Gothic", size=10, color="0000C0")  # 直接入力は青字（実務慣行）
NOTE_FONT = Font(name="Yu Gothic", size=9, color="737781")
HEADER_FILL = PatternFill("solid", fgColor="EDEDF3")
ACTUAL_FILL = PatternFill("solid", fgColor="F5F5F5")
THIN = Side(style="thin", color="C2C6D1")
BORDER = Border(bottom=THIN)

NUM_FMT = "#,##0"
PCT_FMT = "0.0%"
DEC_FMT = "#,##0.00"


def _year_header(ws, header_row: int, kind_row: int):
    for y in YEARS:
        c = ws[f"{COL[y]}{header_row}"]
        c.value = YEAR_LABEL[y]
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")
        k = ws[f"{COL[y]}{kind_row}"]
        k.value = "A（実績）" if y in ACTUAL_YEARS else "P（計画）"
        k.font = NOTE_FONT
        k.alignment = Alignment(horizontal="center")


def _sheet_header(ws, title: str, unit_note: str = "（単位：百万円）",
                  case_label: str | None = None):
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws["A2"] = unit_note
    ws["A2"].font = NOTE_FONT
    if case_label:
        ws["H2"] = f"ケース: {case_label}"
        ws["H2"].font = HEADER_FONT
    _year_header(ws, L.HEADER_ROW, L.KIND_ROW)
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 36
    for y in YEARS:
        ws.column_dimensions[COL[y]].width = 12


def _put_row(ws, row_def, values_or_formulas: dict, fmt=NUM_FMT):
    """1行ぶんを書く。values_or_formulas は {year: value|'=formula'}。"""
    row, label = row_def
    lc = ws[f"B{row}"]
    lc.value = label
    lc.font = BASE_FONT
    for y in YEARS:
        c = ws[f"{COL[y]}{row}"]
        v = values_or_formulas.get(y)
        if v is None:
            continue
        c.value = v
        c.number_format = fmt
        c.font = INPUT_FONT if not (isinstance(v, str) and v.startswith("=")) else BASE_FONT
        if y in ACTUAL_YEARS:
            c.fill = ACTUAL_FILL


def _mixed(formula_tpl: str, actual_values: dict) -> dict:
    """実績年は値、計画年は数式（列記号を差し込み）にする。"""
    out = {}
    for y in ACTUAL_YEARS:
        out[y] = actual_values[y]
    for y in PLAN_YEARS:
        out[y] = formula_tpl.format(c=COL[y])
    return out


def _all_formula(formula_tpl: str, years=YEARS) -> dict:
    return {y: formula_tpl.format(c=COL[y]) for y in years}


def build_workbook(case: str) -> Workbook:
    drv = spec.drivers(case)
    pl = spec.compute_pl(case)
    bs, cf = spec.compute_bs_cf(case)
    sched = spec.debt_schedule()

    wb = Workbook()
    wb.remove(wb.active)

    # ---------------- Cover
    ws = wb.create_sheet(L.SHEET_COVER)
    ws["B4"] = spec.PROJECT_NAME
    ws["B4"].font = Font(name="Yu Gothic", size=24, bold=True, color="1A4F8B")
    ws["B6"] = f"{spec.DEAL['target']}　LBOファイナンシャルモデル"
    ws["B6"].font = Font(name="Yu Gothic", size=16, bold=True)
    ws["B8"] = f"ケース：{CASE_LABEL[case]}（{spec.CASE_LABEL_JA[case]}）"
    ws["B8"].font = Font(name="Yu Gothic", size=12, bold=True)
    kv = [
        ("会社名", spec.DEAL["target"]),
        ("事業内容", spec.DEAL["industry"]),
        ("借入人（SPC）", f"{spec.DEAL['borrower']}（本件買収のための特別目的会社）"),
        ("スポンサー", spec.DEAL["sponsor"]),
        ("直近決算期", "2026年3月期（2026/3期）"),
        ("クローズ予定日", "2026年11月30日"),
        ("通貨・単位", "JPY・百万円（million）"),
        ("作成", f"{spec.DEAL['sponsor']}　投資チーム"),
        ("作成日", "2026年6月20日（Ver. 2.3）"),
    ]
    r = 10
    for k, v in kv:
        ws[f"B{r}"] = k
        ws[f"B{r}"].font = HEADER_FONT
        ws[f"D{r}"] = v
        ws[f"D{r}"].font = BASE_FONT
        r += 1
    ws["B21"] = "【取扱厳重注意】"
    ws["B22"] = ("本資料は貴行における本件検討のためにのみ作成されたものであり、"
                 "作成者の事前の書面による同意なく第三者への開示・複製を禁じます。")
    ws["B23"] = ("本資料に含まれる将来予測は多くの前提・仮定に基づくものであり、"
                 "その実現性を保証するものではありません。")
    for rr in (21, 22, 23):
        ws[f"B{rr}"].font = NOTE_FONT
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["D"].width = 72

    # ---------------- inp（前提条件・KPIドライバー）
    ws = wb.create_sheet(L.SHEET_INP)
    _sheet_header(ws, "前提条件・KPIドライバー（inp）",
                  "（単位：百万円。人数・円・％は行ラベルに明記）", CASE_LABEL[case])
    # 年度ヘッダーはドライバー表側に付け直すため、上部のものを消す
    for y in YEARS:
        ws[f"{COL[y]}{L.HEADER_ROW}"] = None
        ws[f"{COL[y]}{L.KIND_ROW}"] = None
        ws[f"{COL[y]}{L.HEADER_ROW}"].fill = PatternFill()
    ws["B6"] = "ストラクチャー（Sources & Uses・のれん）"
    ws["B6"].font = HEADER_FONT
    S = L.INP_STRUCT_ROWS
    su = spec.SOURCES_USES
    struct_values = {
        "ev": (spec.DEAL["ev_mm"], NUM_FMT),
        "ev_multiple": (8.0, '0.0"x"'),
        "equity_value": (su["equity_value"], NUM_FMT),
        "senior": (su["senior"], NUM_FMT),
        "equity": (su["equity"], NUM_FMT),
        "refi": (su["refi"], NUM_FMT),
        "fees": (su["fees"], NUM_FMT),
        "cash_used": (su["cash_used"], NUM_FMT),
        "goodwill": (spec.GOODWILL, NUM_FMT),
        "tenor": (spec.DEAL["tenor_years"], "0"),
        "rate": (spec.INTEREST_RATE, PCT_FMT),
        "repayment": (spec.ANNUAL_REPAYMENT, NUM_FMT),
        "sponsor_ebitda": (spec.DEAL["sponsor_ebitda_mm"], NUM_FMT),
        "close": ("2026年11月30日", None),
        "borrower": (spec.DEAL["borrower"], None),
    }
    for key, (row, label) in S.items():
        ws[f"B{row}"] = label
        ws[f"B{row}"].font = BASE_FONT
        v, fmt = struct_values[key]
        c = ws[f"D{row}"]
        c.value = v
        if fmt:
            c.number_format = fmt
        c.font = INPUT_FONT
    ws["B25"] = ("※ のれん想定額 = 株式対価9,220 − 時価純資産2,655 − 識別無形資産465"
                 "（暫定PPA前・財務DDの時価純資産評価に基づく）")
    ws["B25"].font = NOTE_FONT

    ws[f"B{L.INP_DRIVER_HEADER_ROW - 1}"] = "KPIドライバー（Operating Assumptions）"
    ws[f"B{L.INP_DRIVER_HEADER_ROW - 1}"].font = HEADER_FONT
    _year_header(ws, L.INP_DRIVER_HEADER_ROW, L.INP_DRIVER_KIND_ROW)
    D = L.INP_DRIVER_ROWS
    member_row = D["member"][0]
    member_vals = {}
    for y in ACTUAL_YEARS:
        member_vals[y] = drv[y]["member"]
    for i, y in enumerate(PLAN_YEARS):
        prev_col = COL[ACTUAL_YEARS[-1]] if i == 0 else COL[PLAN_YEARS[i - 1]]
        member_vals[y] = (f"=ROUND({prev_col}{member_row}*{COL[y]}{D['repeat'][0]}"
                          f"+{COL[y]}{D['new'][0]},0)")
    _put_row(ws, D["member"], member_vals)
    _put_row(ws, D["repeat"], {y: drv[y]["repeat"] for y in YEARS}, PCT_FMT)
    _put_row(ws, D["new"], {y: drv[y]["new"] for y in YEARS})
    _put_row(ws, D["freq"], {y: drv[y]["freq"] for y in YEARS}, DEC_FMT)
    _put_row(ws, D["aov"], {y: drv[y]["aov"] for y in YEARS})
    _put_row(ws, D["doors"], {y: drv[y]["doors"] for y in YEARS})
    _put_row(ws, D["perdoor"], {y: drv[y]["perdoor"] for y in YEARS})
    _put_row(ws, D["cogs_rate"], {y: drv[y]["cogs_rate"] for y in YEARS}, PCT_FMT)
    _put_row(ws, D["ad_rate"], {y: drv[y]["ad_rate"] for y in YEARS}, PCT_FMT)
    _put_row(ws, D["log_rate"], {y: drv[y]["log_rate"] for y in YEARS}, PCT_FMT)
    _put_row(ws, D["capex"], {y: spec._PLAN_CAPEX[case] for y in PLAN_YEARS})
    ws["B43"] = "※ アクティブ会員数（計画年）= 前期会員数 × リピート率 ＋ 新規獲得会員数"
    ws["B43"].font = NOTE_FONT

    # ---------------- debt
    ws = wb.create_sheet(L.SHEET_DEBT)
    _sheet_header(ws, "デットスケジュール（シニアタームローンA）",
                  case_label=CASE_LABEL[case])
    Db = L.DEBT_ROWS
    opening = {}
    for i, y in enumerate(PLAN_YEARS):
        opening[y] = f"=inp!D{L.INP_STRUCT_ROWS['senior'][0]}" if i == 0 else \
            f"={COL[PLAN_YEARS[i - 1]]}{Db['closing'][0]}"
    _put_row(ws, Db["opening"], opening)
    _put_row(ws, Db["repayment"],
             _all_formula("=inp!D%d" % L.INP_STRUCT_ROWS["repayment"][0], PLAN_YEARS))
    _put_row(ws, Db["closing"],
             _all_formula("={c}%d-{c}%d" % (Db["opening"][0], Db["repayment"][0]), PLAN_YEARS))
    _put_row(ws, Db["avg"],
             _all_formula("=({c}%d+{c}%d)/2" % (Db["opening"][0], Db["closing"][0]), PLAN_YEARS))
    _put_row(ws, Db["rate"],
             _all_formula("=inp!D%d" % L.INP_STRUCT_ROWS["rate"][0], PLAN_YEARS), PCT_FMT)
    _put_row(ws, Db["interest"],
             _all_formula("=ROUND({c}%d*{c}%d,0)" % (Db["avg"][0], Db["rate"][0]), PLAN_YEARS))
    ws["B14"] = ("※ 約定弁済は年480百万円、最終回（FY34）残額一括。"
                 "実績年欄は対象外（既存借入はBS参照）。")
    ws["B14"].font = NOTE_FONT

    # ---------------- PL
    ws = wb.create_sheet(L.SHEET_PL)
    _sheet_header(ws, "損益計算書（Projected P&L）", case_label=CASE_LABEL[case])
    P = L.PL_ROWS
    D = L.INP_DRIVER_ROWS
    _put_row(ws, P["ec_rev"], _mixed(
        "=ROUND(inp!{c}%d*inp!{c}%d*inp!{c}%d/10^6,0)"
        % (D["member"][0], D["freq"][0], D["aov"][0]),
        {y: pl[y]["ec_rev"] for y in ACTUAL_YEARS}))
    _put_row(ws, P["ws_rev"], _mixed(
        "=ROUND(inp!{c}%d*inp!{c}%d/1000,0)" % (D["doors"][0], D["perdoor"][0]),
        {y: pl[y]["ws_rev"] for y in ACTUAL_YEARS}))
    _put_row(ws, P["other_revenue"], {y: pl[y]["other_revenue"] for y in YEARS})
    _put_row(ws, P["revenue"],
             _all_formula("=SUM({c}%d:{c}%d)" % (P["ec_rev"][0], P["other_revenue"][0])))
    _put_row(ws, P["cogs"], _mixed(
        "=ROUND({c}%d*inp!{c}%d,0)" % (P["revenue"][0], D["cogs_rate"][0]),
        {y: pl[y]["cogs"] for y in ACTUAL_YEARS}))
    _put_row(ws, P["gross"],
             _all_formula("={c}%d-{c}%d" % (P["revenue"][0], P["cogs"][0])))
    _put_row(ws, P["gross_margin"],
             _all_formula("={c}%d/{c}%d" % (P["gross"][0], P["revenue"][0])), PCT_FMT)
    _put_row(ws, P["ad"], _mixed(
        "=ROUND({c}%d*inp!{c}%d,0)" % (P["revenue"][0], D["ad_rate"][0]),
        {y: pl[y]["ad"] for y in ACTUAL_YEARS}))
    _put_row(ws, P["logistics"], _mixed(
        "=ROUND({c}%d*inp!{c}%d,0)" % (P["revenue"][0], D["log_rate"][0]),
        {y: pl[y]["logistics"] for y in ACTUAL_YEARS}))
    _put_row(ws, P["personnel"], {y: pl[y]["personnel"] for y in YEARS})
    _put_row(ws, P["other_sga"], {y: pl[y]["other_sga"] for y in YEARS})
    _put_row(ws, P["sga"],
             _all_formula("=SUM({c}%d:{c}%d)" % (P["ad"][0], P["other_sga"][0])))
    _put_row(ws, P["op"],
             _all_formula("={c}%d-{c}%d" % (P["gross"][0], P["sga"][0])))
    _put_row(ws, P["op_margin"],
             _all_formula("={c}%d/{c}%d" % (P["op"][0], P["revenue"][0])), PCT_FMT)
    _put_row(ws, P["depreciation"], {y: pl[y]["depreciation"] for y in YEARS})
    _put_row(ws, P["ebitda"],
             _all_formula("={c}%d+{c}%d" % (P["op"][0], P["depreciation"][0])))
    _put_row(ws, P["ebitda_margin"],
             _all_formula("={c}%d/{c}%d" % (P["ebitda"][0], P["revenue"][0])), PCT_FMT)
    _put_row(ws, P["interest"], _mixed(
        "=debt!{c}%d" % L.DEBT_ROWS["interest"][0],
        {y: pl[y]["interest"] for y in ACTUAL_YEARS}))
    _put_row(ws, P["non_op"], {y: pl[y]["non_op"] for y in YEARS})
    _put_row(ws, P["ordinary"],
             _all_formula("={c}%d-{c}%d+{c}%d" % (P["op"][0], P["interest"][0], P["non_op"][0])))
    _put_row(ws, P["tax"],
             _all_formula("=ROUND({c}%d*%s,0)" % (P["ordinary"][0], spec.TAX_RATE)))
    _put_row(ws, P["ni"],
             _all_formula("={c}%d-{c}%d" % (P["ordinary"][0], P["tax"][0])))
    for key in ("revenue", "gross", "op", "ebitda", "ordinary", "ni"):
        row = P[key][0]
        for y in YEARS:
            ws[f"{COL[y]}{row}"].font = HEADER_FONT
        ws[f"B{row}"].font = HEADER_FONT

    # ---------------- BS
    ws = wb.create_sheet(L.SHEET_BS)
    _sheet_header(ws, "貸借対照表（Projected B/S）", case_label=CASE_LABEL[case])
    B = L.BS_ROWS
    _put_row(ws, B["cash"], _mixed(
        "=CF!{c}%d" % L.CF_ROWS["closing_cash"][0],
        {y: bs[y]["cash"] for y in ACTUAL_YEARS}))
    _put_row(ws, B["ar"], _mixed(
        "=ROUND(PL!{c}%d*%d/365,0)" % (L.PL_ROWS["revenue"][0], spec.AR_DAYS),
        {y: bs[y]["ar"] for y in ACTUAL_YEARS}))
    _put_row(ws, B["inv"], _mixed(
        "=ROUND(PL!{c}%d*%d/365,0)" % (L.PL_ROWS["cogs"][0], spec.INV_DAYS),
        {y: bs[y]["inv"] for y in ACTUAL_YEARS}))
    _put_row(ws, B["oca"], {y: bs[y]["oca"] for y in YEARS})
    ppe_vals = {y: bs[y]["ppe"] for y in ACTUAL_YEARS}
    for i, y in enumerate(PLAN_YEARS):
        prev_col = COL[ACTUAL_YEARS[-1]] if i == 0 else COL[PLAN_YEARS[i - 1]]
        ppe_vals[y] = (f"={prev_col}{B['ppe'][0]}+inp!{COL[y]}{L.INP_DRIVER_ROWS['capex'][0]}"
                       f"-PL!{COL[y]}{L.PL_ROWS['depreciation'][0]}")
    _put_row(ws, B["ppe"], ppe_vals)
    _put_row(ws, B["goodwill"], {y: bs[y]["goodwill"] for y in YEARS})
    _put_row(ws, B["intangible"], {y: bs[y]["intangible"] for y in YEARS})
    _put_row(ws, B["total_assets"],
             _all_formula("=SUM({c}%d:{c}%d)" % (B["cash"][0], B["intangible"][0])))
    _put_row(ws, B["ap"], {y: bs[y]["ap"] for y in YEARS})
    _put_row(ws, B["other_liab"], {y: bs[y]["other_liab"] for y in YEARS})
    _put_row(ws, B["debt"], _mixed(
        "=debt!{c}%d" % L.DEBT_ROWS["closing"][0],
        {y: bs[y]["debt"] for y in ACTUAL_YEARS}))
    _put_row(ws, B["total_liab"],
             _all_formula("=SUM({c}%d:{c}%d)" % (B["ap"][0], B["debt"][0])))
    _put_row(ws, B["net_assets"], {y: bs[y]["net_assets"] for y in YEARS})
    _put_row(ws, B["total_le"],
             _all_formula("={c}%d+{c}%d" % (B["total_liab"][0], B["net_assets"][0])))
    ws["B24"] = ("※ 2027/3期以降は買収後連結（のれん計上・買収時調整を純資産に含む）。"
                 "のれんは暫定PPA前の想定額（inp参照）。")
    ws["B24"].font = NOTE_FONT
    for key in ("total_assets", "total_le", "net_assets"):
        ws[f"B{B[key][0]}"].font = HEADER_FONT

    # ---------------- CF
    ws = wb.create_sheet(L.SHEET_CF)
    _sheet_header(ws, "キャッシュフロー計算書（Projected C/F）", case_label=CASE_LABEL[case])
    C = L.CF_ROWS
    _put_row(ws, C["ni"], _all_formula("=PL!{c}%d" % L.PL_ROWS["ni"][0]))
    _put_row(ws, C["depreciation"], _all_formula("=PL!{c}%d" % L.PL_ROWS["depreciation"][0]))
    pl_case = pl
    wc_values = {}
    for y in YEARS:
        wc_values[y] = cf[y]["op_cf"] - pl_case[y]["ni"] - pl_case[y]["depreciation"]
    _put_row(ws, C["wc"], wc_values)
    _put_row(ws, C["op_cf"],
             _all_formula("=SUM({c}%d:{c}%d)" % (C["ni"][0], C["wc"][0])))
    inv_vals = {y: cf[y]["inv_cf"] for y in ACTUAL_YEARS}
    for y in PLAN_YEARS:
        inv_vals[y] = f"=-inp!{COL[y]}{L.INP_DRIVER_ROWS['capex'][0]}"
    _put_row(ws, C["inv_cf"], inv_vals)
    _put_row(ws, C["fcf"],
             _all_formula("={c}%d+{c}%d" % (C["op_cf"][0], C["inv_cf"][0])))
    repay = {}
    fin_other = {}
    for y in ACTUAL_YEARS:
        repay[y] = -100
        fin_other[y] = cf[y]["fin_cf"] + 100
    for y in PLAN_YEARS:
        repay[y] = f"=-debt!{COL[y]}{L.DEBT_ROWS['repayment'][0]}"
        fin_other[y] = cf[y]["fin_cf"] + spec.ANNUAL_REPAYMENT
    _put_row(ws, C["repay"], repay)
    _put_row(ws, C["fin_other"], fin_other)
    _put_row(ws, C["fin_cf"],
             _all_formula("={c}%d+{c}%d" % (C["repay"][0], C["fin_other"][0])))
    _put_row(ws, C["net_change"],
             _all_formula("={c}%d+{c}%d" % (C["fcf"][0], C["fin_cf"][0])))
    opening = {"FY24": spec.OPENING_CASH_FY24, "FY27": spec.POST_CLOSE_CASH}
    for i, y in enumerate(YEARS):
        if y in opening:
            continue
        prev = YEARS[i - 1]
        opening[y] = f"={COL[prev]}{C['closing_cash'][0]}"
    _put_row(ws, C["opening_cash"], opening)
    _put_row(ws, C["closing_cash"],
             _all_formula("={c}%d+{c}%d" % (C["net_change"][0], C["opening_cash"][0])))
    ws["B24"] = ("※ 2027/3期期首現金は買収時調整後（手元現金380百万円を買収資金に充当後の"
                 "240百万円）。")
    ws["B24"].font = NOTE_FONT

    # ---------------- memo（軽微なノイズ：成果物ではない作業メモ）
    ws = wb.create_sheet(L.SHEET_MEMO)
    ws["A1"] = "（作業メモ・成果物ではありません）"
    ws["A1"].font = NOTE_FONT
    memo = [
        ("B3", "6/18 銀行提出用に単位を百万円へ統一（旧版は千円）"),
        ("B4", "TODO: PPA確定後にのれん・無形の内訳を更新（暫定 6,100 / 465）"),
        ("B5", "リピート率の定義＝当期購入会員のうち前期も購入した会員の比率（CRM抽出）"),
        ("B6", "卸の店販は帳合経由の出荷ベース。返品控除後"),
    ]
    for addr, v in memo:
        ws[addr] = v
        ws[addr].font = BASE_FONT
    ws.column_dimensions["B"].width = 80

    # シート順
    order = [L.SHEET_COVER, L.SHEET_INP, L.SHEET_PL, L.SHEET_BS, L.SHEET_CF,
             L.SHEET_DEBT, L.SHEET_MEMO]
    wb._sheets = [wb[name] for name in order]
    return wb


def recalc_with_libreoffice(paths: list[Path]):
    """LibreOfficeで再計算して保存し直し、数式セルにキャッシュ値を持たせる。

    openpyxlが書いたxlsxは数式セルに計算結果（キャッシュ値）を持たない。
    実務でアップロードされるファイルはExcelが保存したもの＝キャッシュ値を持つため、
    再計算・保存を挟んで同じ状態を作る（AI抽出は値の拾い上げができる必要がある）。
    """
    soffice = shutil.which("soffice") or shutil.which("libreoffice")
    if not soffice:
        raise RuntimeError(
            "LibreOffice（soffice）が見つかりません。数式のキャッシュ値を生成できないため"
            "中断します。LibreOfficeをインストールするか、Excelで開いて保存してください。")
    with tempfile.TemporaryDirectory() as td:
        subprocess.run(
            [soffice, "--headless", "--convert-to", "xlsx:Calc MS Excel 2007 XML",
             "--outdir", td, *[str(p) for p in paths]],
            check=True, capture_output=True, timeout=300)
        for p in paths:
            out = Path(td) / p.name
            if not out.exists():
                raise RuntimeError(f"LibreOffice変換に失敗: {p.name}")
            shutil.copy2(out, p)


def verify_cached_values(case: str, path: Path):
    """キャッシュ値（data_only=True）が spec の計算値と一致することを確認する。"""
    wb = load_workbook(path, data_only=True)
    pl = spec.compute_pl(case)
    bs, cf = spec.compute_bs_cf(case)
    drv = spec.drivers(case)
    errors = []

    def chk(sheet, row, year, expected, label):
        got = wb[sheet][f"{COL[year]}{row}"].value
        if got is None or abs(float(got) - float(expected)) > 0.5:
            errors.append(f"{sheet}!{COL[year]}{row} ({label} {year}): {got} != {expected}")

    for y in YEARS:
        chk(L.SHEET_PL, L.PL_ROWS["revenue"][0], y, pl[y]["revenue"], "売上収益")
        chk(L.SHEET_PL, L.PL_ROWS["op"][0], y, pl[y]["op"], "営業利益")
        chk(L.SHEET_PL, L.PL_ROWS["ebitda"][0], y, pl[y]["ebitda"], "EBITDA")
        chk(L.SHEET_PL, L.PL_ROWS["ni"][0], y, pl[y]["ni"], "当期純利益")
        chk(L.SHEET_BS, L.BS_ROWS["cash"][0], y, bs[y]["cash"], "現預金")
        chk(L.SHEET_BS, L.BS_ROWS["net_assets"][0], y, bs[y]["net_assets"], "純資産")
        chk(L.SHEET_BS, L.BS_ROWS["debt"][0], y,
            bs[y]["debt"] if y in ACTUAL_YEARS else spec.debt_schedule()[y]["closing"],
            "有利子負債")
        chk(L.SHEET_CF, L.CF_ROWS["fcf"][0], y, cf[y]["fcf"], "FCF")
    for y in PLAN_YEARS:
        chk(L.SHEET_INP, L.INP_DRIVER_ROWS["member"][0], y, drv[y]["member"], "会員数")
    if errors:
        raise RuntimeError(f"{path.name}: キャッシュ値の不一致 {len(errors)}件\n  "
                           + "\n  ".join(errors[:10]))


def main():
    paths = []
    for case in spec.CASES:
        wb = build_workbook(case)
        path = OUT_DIR / spec.MODEL_FILES[case]
        wb.save(path)
        paths.append((case, path))
        print(f"generated: {path.name}")
    recalc_with_libreoffice([p for _, p in paths])
    print("recalculated with LibreOffice (formula cache values written)")
    for case, path in paths:
        verify_cached_values(case, path)
        print(f"verified cached values: {path.name}")


if __name__ == "__main__":
    try:
        main()
    except RuntimeError as e:
        print(f"ERROR: {e}", file=sys.stderr)
        sys.exit(1)
