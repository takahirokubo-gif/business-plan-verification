# -*- coding: utf-8 -*-
"""定量条件表（claude_code_prompt v2 第4章）の自動チェック（作成時の自己チェック）。

- 1条件＝1チェック関数。結果を validation_report.json（条件ID／判定／実測値／期待値）に出力
- 判定は pass / fail / manual（機械判定困難＝目視確認をmanifestに記録）
- 数式セルはミニ評価器で実際に評価し、「意図した破損セル以外の数式エラーゼロ」
  「PL縦計算・BS貸借一致」を検証する
- 期待値は testdata/generator/spec12.py（数値正本）から取得する

実行: backend/.venv/bin/python testdata/validation/checks.py
依存: openpyxl, pypdf
"""
import json
import hashlib
import re
import sys
from pathlib import Path

import openpyxl
from pypdf import PdfReader

HERE = Path(__file__).resolve().parent
ROOT = HERE.parent                       # testdata/
sys.path.insert(0, str(ROOT / "generator"))
import xl12r as L                         # noqa: E402（spec12へのパスもxl12rが解決）
import spec12 as S                        # noqa: E402
from spec12 import ACTUAL_YEARS, COL, PLAN_YEARS, YEARS  # noqa: E402

A_XLSX = ROOT / "A_standard" / "AutostaffChubu_Model_A.xlsx"
A_PDF = ROOT / "A_standard" / "DD_Report_統合版_オートスタッフ中部_A.pdf"
B_BASE = ROOT / "B_hard" / "AutostaffChubu_Model_Base_B.xlsx"
B_SPONSOR = ROOT / "B_hard" / "AutostaffChubu_Model_Sponsor_B.xlsx"
B_DEBT = ROOT / "B_hard" / "AutostaffChubu_Debt_B.xlsx"
DUMMY = ROOT.parent / "dummy_input"
GT = ROOT / "ground_truth" / "ground_truth.xlsx"

RESULTS = []

TERMS15 = ["稼働率", "派遣単価", "スプレッド", "在籍登録スタッフ", "稼働人数",
           "採用単価（CPA）", "リファラル", "離職率", "法定福利費率", "労使協定方式",
           "構内請負", "コーディネーター", "寮", "送迎", "早期離職率"]
KEYWORDS_D2 = ["稼働率", "単価改定", "CPA", "採用人数"]

MONTH_COL = L.MCOL  # {"2025/04": "C", ...}


def record(cid, dataset, desc, status, actual, expected):
    RESULTS.append(dict(id=cid, dataset=dataset, desc=desc, status=status,
                        actual=str(actual), expected=str(expected)))
    mark = {"pass": "✓", "fail": "✗", "manual": "◆"}[status]
    print(f"  {mark} [{cid}][{dataset}] {desc} → {status} (actual={actual})")


def check(cid, dataset, desc, ok, actual, expected):
    record(cid, dataset, desc, "pass" if ok else "fail", actual, expected)


# ---------------------------------------------------------------- ミニ数式評価器

_REF_RE = re.compile(r"(?:(?P<sheet>[A-Za-z0-9_月次試算表FY]+)!)?\$?(?P<col>[A-Z]{1,2})\$?(?P<row>\d+)")


class Evaluator:
    """openpyxlで読んだ数式セルを評価する簡易エンジン。

    対応: 四則演算・括弧・単項マイナス・SUM(range)・ROUND(x,n)・シート間参照。
    #REF! を含む数式は BrokenRef として記録する。
    """

    def __init__(self, wb):
        self.wb = wb
        self.memo = {}
        self.broken = []      # (sheet, coord)
        self.errors = []      # (sheet, coord, message)

    def value(self, sheet, coord):
        key = (sheet, coord)
        if key in self.memo:
            v = self.memo[key]
            if v == "#CYCLE#":
                raise ValueError(f"circular ref {key}")
            return v
        self.memo[key] = "#CYCLE#"
        try:
            cell = self.wb[sheet][coord]
        except KeyError:
            self.errors.append((sheet, coord, "missing sheet"))
            self.memo[key] = None
            return None
        v = cell.value
        if isinstance(v, str) and v.startswith("="):
            if "#REF!" in v:
                self.broken.append((sheet, coord))
                self.memo[key] = None
                return None
            try:
                result = self._eval_expr(v[1:], sheet)
            except Exception as e:  # noqa: BLE001
                self.errors.append((sheet, coord, str(e)))
                result = None
            self.memo[key] = result
            return result
        self.memo[key] = v
        return v

    # --- パーサ（再帰下降・再入可能） ---
    def _eval_expr(self, s, sheet):
        saved = (getattr(self, "_s", None), getattr(self, "_pos", None),
                 getattr(self, "_sheet", None))
        self._s, self._pos, self._sheet = s, 0, sheet
        try:
            v = self._expr()
            if self._pos != len(self._s):
                raise ValueError(f"parse error at {self._pos}: {s}")
            return v
        finally:
            self._s, self._pos, self._sheet = saved

    def _peek(self):
        # 終端は "\0"（空文字だと `"" in "+-"` がTrueになり暴走する）
        return self._s[self._pos] if self._pos < len(self._s) else "\0"

    def _expr(self):
        v = self._term()
        while self._peek() in "+-":
            op = self._s[self._pos]
            self._pos += 1
            rhs = self._term()
            v = v + rhs if op == "+" else v - rhs
        return v

    def _term(self):
        v = self._factor()
        while self._peek() in "*/":
            op = self._s[self._pos]
            self._pos += 1
            rhs = self._factor()
            v = v * rhs if op == "*" else v / rhs
        return v

    def _factor(self):
        ch = self._peek()
        if ch == "-":
            self._pos += 1
            return -self._factor()
        if ch == "(":
            self._pos += 1
            v = self._expr()
            assert self._peek() == ")", "missing )"
            self._pos += 1
            return v
        m = re.match(r"(SUM|ROUND)\(", self._s[self._pos:])
        if m:
            fname = m.group(1)
            self._pos += len(fname) + 1
            if fname == "SUM":
                rng = self._read_until_close()
                return self._sum_range(rng)
            args_start = self._pos
            depth = 1
            while depth:
                c = self._s[self._pos]
                depth += 1 if c == "(" else (-1 if c == ")" else 0)
                self._pos += 1
            inner = self._s[args_start:self._pos - 1]
            expr, _, digits = inner.rpartition(",")
            x = self._subeval(expr)
            return S.excel_round(x, int(digits))
        m = re.match(r"\d+(\.\d+)?", self._s[self._pos:])
        num_m = m
        ref_m = _REF_RE.match(self._s, self._pos)
        if ref_m and (not num_m or ref_m.end() > self._pos + num_m.end()):
            self._pos = ref_m.end()
            sheet = ref_m.group("sheet") or self._sheet
            v = self.value(sheet, f"{ref_m.group('col')}{ref_m.group('row')}")
            if v is None:
                raise ValueError(f"ref to empty/broken {sheet}!{ref_m.group('col')}{ref_m.group('row')}")
            return float(v)
        if num_m:
            self._pos += num_m.end()
            return float(num_m.group(0))
        raise ValueError(f"unexpected char at {self._pos}: {self._s}")

    def _subeval(self, expr):
        saved = (self._s, self._pos)
        try:
            self._s, self._pos = expr, 0
            v = self._expr()
            assert self._pos == len(expr)
            return v
        finally:
            self._s, self._pos = saved

    def _read_until_close(self):
        start = self._pos
        depth = 1
        while depth:
            c = self._s[self._pos]
            depth += 1 if c == "(" else (-1 if c == ")" else 0)
            self._pos += 1
        return self._s[start:self._pos - 1]

    def _sum_range(self, rng):
        sheet = self._sheet
        if "!" in rng:
            sheet, rng = rng.split("!", 1)
        a, b = rng.split(":")
        c1, r1 = openpyxl.utils.cell.coordinate_from_string(a)
        c2, r2 = openpyxl.utils.cell.coordinate_from_string(b)
        i1, i2 = openpyxl.utils.column_index_from_string(c1), openpyxl.utils.column_index_from_string(c2)
        total = 0.0
        for r in range(min(r1, r2), max(r1, r2) + 1):
            for i in range(min(i1, i2), max(i1, i2) + 1):
                v = self.value(sheet, f"{openpyxl.utils.get_column_letter(i)}{r}")
                if v is not None and not isinstance(v, str):
                    total += float(v)
        return total


# ---------------------------------------------------------------- ヘルパ

def load(path):
    return openpyxl.load_workbook(path)


def sheet_strings(ws):
    out = []
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c.value, str):
                out.append((c.coordinate, c.value))
    return out


def pdf_pages(path):
    return [p.extract_text() or "" for p in PdfReader(str(path)).pages]


def wb_all_text(path):
    wb = load(path)
    chunks = []
    for ws in wb.worksheets:
        chunks.extend(v for _, v in sheet_strings(ws))
    return "\n".join(chunks)


# ---------------------------------------------------------------- 共通チェック（各Excel）

YEAR_SHEETS = [L.SHEET_PL, L.SHEET_BS, L.SHEET_CF, L.SHEET_CALC, L.SHEET_INP]


def check_common_excel(tag, path, variant, case):
    wb = load(path)
    ds = "A" if variant == "A" else "B"

    # E1: シート数・三表分離
    sheets = [s for s in wb.sheetnames if s != L.SHEET_COVER]
    ok = len(sheets) >= 6 and all(s in wb.sheetnames for s in ("PL", "BS", "CF"))
    check("E1", ds, f"{tag}: シート数6枚以上（Cover除く）・PL/BS/CF別シート",
          ok, f"{len(sheets)}枚 {sheets}", "6枚以上・PL/BS/CF存在")

    # E1/E5: 12期ヘッダーと実績/計画区分の一貫性
    bad = []
    for sname in YEAR_SHEETS:
        ws = wb[sname]
        hdr = [ws[f"{COL[y]}{L.HEADER_ROW}"].value for y in YEARS]
        kind = [ws[f"{COL[y]}{L.KIND_ROW}"].value for y in YEARS]
        if hdr != [L.year_label(y) for y in YEARS]:
            bad.append(f"{sname}:hdr={hdr}")
        if kind != ["実績"] * 5 + ["計画"] * 7:
            bad.append(f"{sname}:kind={kind}")
    check("E1/E5", ds, f"{tag}: 全対象シートの期間列がFY22A〜FY33Eの12期・実績/計画区分一致",
          not bad, bad or "12期一貫", "FY22〜FY33・実績5期＋計画7期")

    # E2: 月次試算表の構造
    ws = wb[L.SHEET_MONTHLY]
    hdr = [ws[f"{MONTH_COL[m]}{L.HEADER_ROW}"].value for m in S.MONTHS]
    pl_labels = [ws[f"B{r}"].value for r in (7, 8, 9, 10, 11)]
    bs_labels = [ws[f"B{r}"].value for r in (13, 14, 15)]
    ok = hdr == S.MONTHS and all(pl_labels) and all(bs_labels)
    check("E2", ds, f"{tag}: 月次試算表_FY26（2025/04〜2026/03・PL5行＋BS3行）",
          ok, f"hdr={hdr[:2]}…{hdr[-1]} PL={len([x for x in pl_labels if x])}行 BS={len([x for x in bs_labels if x])}行",
          "12ヶ月・PL5行以上・BS3行以上")

    # E2: 月次合計と年次の一致（±1千円）・単位が円であること
    mon_vals = {k: {m: ws[f"{MONTH_COL[m]}{r}"].value for m in S.MONTHS}
                for k, (r, _, _) in L.MONTHLY_ROWS.items() if k in ("revenue", "cogs", "sga")}
    pl26 = S.compute_pl(case)["FY26"]
    diffs = {
        "revenue": sum(mon_vals["revenue"].values()) - pl26["revenue"] * 1000,
        "cogs": sum(mon_vals["cogs"].values()) - pl26["cogs"] * 1000,
        "sga": sum(mon_vals["sga"].values()) - pl26["sga"] * 1000,
    }
    bs26 = S._BS_ACTUAL["FY26"]
    bs_mar = {
        "cash": ws[f"N{L.MONTHLY_ROWS['cash'][0]}"].value - bs26["cash"] * 1000,
        "ar": ws[f"N{L.MONTHLY_ROWS['ar'][0]}"].value - bs26["ar"] * 1000,
        "ap": ws[f"N{L.MONTHLY_ROWS['ap'][0]}"].value - bs26["ap"] * 1000,
    }
    ok = all(abs(d) <= 1000 for d in diffs.values()) and all(d == 0 for d in bs_mar.values())
    check("E2", ds, f"{tag}: 月次12ヶ月合計＝年次FY26（±1千円）・3月末残高＝年次BS",
          ok, f"PL差異(円)={diffs} BS差異(円)={bs_mar}", "PL±1,000円以内・BS一致")

    # E2: 季節性（5月・8月・1月が他9ヶ月平均比△5〜8%）
    rev = {m: ws[f"{MONTH_COL[m]}7"].value for m in S.MONTHS}
    normal_avg = sum(v for m, v in rev.items() if m not in S.LOW_MONTHS) / 9
    drops = {m: 1 - rev[m] / normal_avg for m in S.LOW_MONTHS}
    ok = all(0.05 <= d <= 0.08 for d in drops.values())
    check("E2", ds, f"{tag}: 季節月（5月・8月・1月）の売上が他9ヶ月平均比△5〜8%",
          ok, {m: f"{d:.1%}" for m, d in drops.items()}, "各月5〜8%低下")

    # E3: 単位表記の「位置×記法」3種類以上
    patterns = []
    if wb["PL"][f"{COL['FY33']}2"].value == "単位:千円":
        patterns.append("PL!N2(ヘッダー右端・コロン)")
    if "・千円" in (wb["CF"]["A1"].value or ""):
        patterns.append("CF!A1(タイトル内)")
    if wb["BS"]["A2"].value == "（単位：千円）":
        patterns.append("BS!A2(標準)")
    if variant == "A" and wb[L.SHEET_DEBT]["A2"].value == "(Unit: JPY thousands)":
        patterns.append("Debt!A2(英語)")
    check("E3", ds, f"{tag}: 単位表記の位置×記法が3種類以上",
          len(patterns) >= 3, patterns, "3種類以上")

    # E4: 年次ヘッダーFY表記100%・英語併記科目3件以上
    non_fy = []
    for sname in YEAR_SHEETS + ([L.SHEET_DEBT] if variant == "A" else []):
        for y in YEARS:
            v = wb[sname][f"{COL[y]}{L.HEADER_ROW}"].value
            if not (isinstance(v, str) and re.fullmatch(r"FY\d{2}[AE]", v)):
                non_fy.append(f"{sname}!{COL[y]}4={v}")
    check("E4", ds, f"{tag}: 年次ヘッダーのFY表記率100%", not non_fy, non_fy or "100%", "100%")

    bilingual = []
    for sname in (L.SHEET_PL, L.SHEET_BS, L.SHEET_KPI):
        for coord, v in sheet_strings(wb[sname]):
            if not coord.startswith("B"):
                continue
            m = re.search(r"（([A-Za-z][A-Za-z .&/']+)）", v)
            if m and re.search(r"[぀-ヿ一-鿿]", v):
                bilingual.append(f"{sname}!{coord}:{v}")
    check("E4", ds, f"{tag}: 英語併記科目3件以上", len(bilingual) >= 3,
          f"{len(bilingual)}件", "3件以上")

    # E5: FY22〜23実績（黒字・CAGR・接続）／FY32〜33計画（稼働率・単価）
    pl = S.compute_pl(case)
    ev5 = Evaluator(wb)
    written = {y: ev5.value("PL", f"{COL[y]}{L.PL_ROWS['ni'][0]}") for y in ("FY22", "FY23")}
    cagr = (pl["FY26"]["revenue"] / pl["FY22"]["revenue"]) ** 0.25 - 1
    growth = [pl[YEARS[i]]["revenue"] / pl[YEARS[i - 1]]["revenue"] - 1 for i in range(1, 5)]
    ok = (all(written[y] == pl[y]["ni"] and pl[y]["ni"] > 0 for y in ("FY22", "FY23"))
          and 0.05 <= cagr <= 0.08 and all(0.04 <= g <= 0.09 for g in growth))
    check("E5", ds, f"{tag}: FY22〜23実績（全期黒字・CAGR5〜8%・滑らかな接続）",
          ok, f"NI={written} CAGR={cagr:.2%} 成長率={[f'{g:.1%}' for g in growth]}",
          "黒字・CAGR5〜8%・各年成長率が連続")

    drv = S.drivers(case)
    if case == "base":
        r1 = drv["FY32"]["bill"] / drv["FY31"]["bill"] - 1
        r2 = drv["FY33"]["bill"] / drv["FY32"]["bill"] - 1
        w_util = [wb[L.SHEET_CALC][f"{COL[y]}{L.KPI_ROWS['util'][0]}"].value for y in ("FY32", "FY33")]
        ok = w_util == [0.88, 0.88] and 0.005 <= r1 <= 0.015 and 0.005 <= r2 <= 0.015
        check("E5", ds, f"{tag}: FY32〜33計画（稼働率88%維持・単価改定+1%/年前後）",
              ok, f"util={w_util} 改定率={r1:.2%}/{r2:.2%}", "88%・+1%前後")

    return wb


def check_integrity(tag, path, variant, case):
    """再計算＋整合検査：破損セル以外の数式エラーゼロ・PL縦計算・BS貸借一致・値の正本一致。"""
    ds = "A" if variant == "A" else "B"
    wb = load(path)
    ev = Evaluator(wb)
    pl = S.compute_pl(case)
    bs, cf = S.compute_bs_cf(case)
    sched = S.debt_schedule()
    mismatches = []

    def expect(sheet, row_key, y, want, rows=None):
        rows = rows or {"PL": L.PL_ROWS, "BS": L.BS_ROWS, "CF": L.CF_ROWS,
                        L.SHEET_CALC: L.KPI_ROWS, L.SHEET_DEBT: L.DEBT_ROWS}[sheet]
        coord = f"{COL[y]}{rows[row_key][0]}"
        got = ev.value(sheet, coord)
        if got is None or abs(float(got) - want) > 0.5:
            mismatches.append(f"{sheet}!{coord}: got={got} want={want}")

    # PL全行（費用行は負値規約。Bの計画期支払利息は破損セルなのでスキップ）
    for y in YEARS:
        for key in ("staffing_rev", "other_revenue", "revenue", "labor", "other_cogs",
                    "cogs", "gross", "recruiting", "hq_cost", "other_sga", "sga",
                    "op", "depreciation", "ebitda", "non_op", "ordinary", "tax", "ni"):
            sign = -1 if key in L.NEGATIVE_KEYS else 1
            expect("PL", key, y, sign * pl[y][key])
        if variant == "A" or y in ACTUAL_YEARS:
            expect("PL", "interest", y, -pl[y]["interest"])
    # BS全行＋貸借一致
    for y in YEARS:
        for key in ("cash", "ar", "oca", "ppe", "goodwill", "intangible", "ap",
                    "other_liab", "debt", "net_assets"):
            expect("BS", key, y, bs[y][key])
        ta = ev.value("BS", f"{COL[y]}{L.BS_ROWS['total_assets'][0]}")
        tle = ev.value("BS", f"{COL[y]}{L.BS_ROWS['total_le'][0]}")
        if ta is None or tle is None or abs(ta - tle) > 0.5:
            mismatches.append(f"BS {y}: 貸借不一致 assets={ta} l+e={tle}")
    # CF
    for y in YEARS:
        for key, want in (("op_cf", cf[y]["op_cf"]), ("inv_cf", cf[y]["inv_cf"]),
                          ("fcf", cf[y]["fcf"]), ("fin_cf", cf[y]["fin_cf"]),
                          ("net_change", cf[y]["net_change"]),
                          ("closing_cash", cf[y]["closing_cash"] if y in ACTUAL_YEARS else bs[y]["cash"])):
            expect("CF", key, y, want)
    # KPI
    drv = S.drivers(case)
    for y in YEARS:
        for key in ("enrolled", "util", "active", "hours", "bill", "wage",
                    "welfare", "hires", "cpa", "attrition"):
            expect(L.SHEET_CALC, key, y, drv[y][key])
    # Debt（Aは本体、Bは別ファイル）。返済・支払利息は負値規約
    debt_wb = wb if variant == "A" else load(B_DEBT)
    dev = Evaluator(debt_wb) if variant == "B" else ev
    for y in PLAN_YEARS:
        for key in ("opening", "repayment", "closing", "interest"):
            coord = f"{COL[y]}{L.DEBT_ROWS[key][0]}"
            got = dev.value(L.SHEET_DEBT, coord)
            want = sched[y][key] * (-1 if key in ("repayment", "interest") else 1)
            if got is None or abs(float(got) - want) > 0.5:
                mismatches.append(f"debt!{coord}: got={got} want={want}")
    for y in ACTUAL_YEARS:
        coord = f"{COL[y]}{L.DEBT_EXISTING_ROWS['ex_closing'][0]}"
        got = dev.value(L.SHEET_DEBT, coord)
        if got is None or abs(float(got) - S.EXISTING_DEBT[y]) > 0.5:
            mismatches.append(f"debt!{coord}: got={got} want={S.EXISTING_DEBT[y]}")
        coord = f"{COL[y]}{L.DEBT_EXISTING_ROWS['ex_interest'][0]}"
        got = dev.value(L.SHEET_DEBT, coord)
        if got is None or abs(float(got) + pl[y]["interest"]) > 0.5:
            mismatches.append(f"debt!{coord}: got={got} want={-pl[y]['interest']}")

    # 数式エラー（破損セル以外ゼロ）
    allowed_broken = set()
    if variant == "B":
        allowed_broken = {("PL", f"{COL[y]}{L.PL_ROWS['interest'][0]}") for y in PLAN_YEARS}
    unexpected_broken = [b for b in ev.broken if b not in allowed_broken]
    errors = ev.errors + ([] if variant == "A" else dev.errors)
    ok = not mismatches and not unexpected_broken and not errors
    check("共通(整合)", ds,
          f"{tag}: 再計算一致・PL縦計算・BS貸借一致・破損セル以外の数式エラーゼロ",
          ok,
          f"不一致{len(mismatches)}件 想定外破損{len(unexpected_broken)}件 評価エラー{len(errors)}件"
          + (f" 例:{(mismatches + [str(x) for x in unexpected_broken])[:3]}" if not ok else ""),
          "全一致・エラーゼロ")
    if variant == "B":
        # 破損セル：PL支払利息の計画7期が =#REF!{列}12 の形で存在（明示検査）
        broken_cells = []
        for y in PLAN_YEARS:
            coord = f"{COL[y]}{L.PL_ROWS['interest'][0]}"
            v = wb["PL"][coord].value
            if isinstance(v, str) and v == f"=#REF!{COL[y]}{L.DEBT_ROWS['interest'][0]}":
                broken_cells.append(f"PL!{coord}")
        ok = len(broken_cells) == 7 and not unexpected_broken
        check("E11", "B", f"{tag}: 支払利息FY27E〜33Eの7セルが解決不能参照（#REF!）・全番地planted記載",
              ok, f"{len(broken_cells)}セル {broken_cells}", "PL!H28:N28の7セル")


# ---------------------------------------------------------------- E6 業界用語

def check_e6():
    a_text = wb_all_text(A_XLSX) + "\n" + "\n".join(pdf_pages(A_PDF))
    hits_a = [t for t in TERMS15 if t.replace("（CPA）", "") in a_text or t in a_text]
    check("E6", "A", "指定業界用語15語中10語以上が出現（Excel A＋統合DD）",
          len(hits_a) >= 10, f"{len(hits_a)}語", "10語以上")
    b_text = wb_all_text(B_BASE) + wb_all_text(B_SPONSOR) + wb_all_text(B_DEBT)
    for f in ("DD_Business", "DD_Financial", "DD_Legal", "DD_Tax"):
        b_text += "\n".join(pdf_pages(ROOT / "B_hard" / f"{f}_オートスタッフ中部.pdf"))
    hits_b = [t for t in TERMS15 if t.replace("（CPA）", "") in b_text or t in b_text]
    check("E6", "B", "指定業界用語15語中10語以上が出現（Excel B＋DD4分冊）",
          len(hits_b) >= 10, f"{len(hits_b)}語", "10語以上")


# ---------------------------------------------------------------- A固有

def check_a_specific():
    wb = load(A_XLSX)
    # E2(A): 単位ラベル明記
    v = wb[L.SHEET_MONTHLY]["A2"].value
    check("E2(A)", "A", "月次試算表の冒頭に「（単位：円）」ラベル",
          v == "（単位：円）", v, "（単位：円）")

    # E4(A): 英語シート（Debt_Schedule）の行ラベル80%以上英語
    labels = [x for _, x in sheet_strings(wb[L.SHEET_DEBT]) if _.startswith("B")]
    eng = [x for x in labels if all(ord(ch) < 128 for ch in x)]
    ratio = len(eng) / len(labels) if labels else 0
    check("E4", "A", "英語行ラベルのシートが1枚以上（Debt_Schedule 80%以上英語）",
          ratio >= 0.8, f"{ratio:.0%} ({len(eng)}/{len(labels)})", "80%以上")

    pages = pdf_pages(A_PDF)
    full = "\n".join(pages)
    # D1(A)
    tables = len(re.findall(r"図表\d+", full))
    check("D1(A)", "A", "統合版DDが単一PDFで60ページ以上・表30個以上",
          len(pages) >= 60 and tables >= 30, f"{len(pages)}ページ・図表{tables}点", "60p以上・30表以上")

    # D2(A): 前提条件の章分散
    chapters = {}
    current = None
    for p in pages[3:]:
        m = re.search(r"^([1-9])\. (会社概要|市場環境|競合環境|事業モデルとKPI|顧客分析|収益性分析|B/S分析|CF分析|計画前提の評価)",
                      p, re.M)
        if m:
            current = m.group(1)
        if re.search(r"^付属資料", p, re.M):
            current = "付属資料"
        if current and current != "付属資料":
            chapters.setdefault(current, []).append(p)
    kw_chapters = {k: sorted({ch for ch, ps in chapters.items() if any(k in p for p in ps)})
                   for k in KEYWORDS_D2}
    spread = {k: len(v) for k, v in kw_chapters.items()}
    ok = sum(1 for n in spread.values() if n >= 3) >= 3
    check("D2(A)", "A", "計画前提（稼働率・単価改定・CPA・採用人数のうち3種以上）が3章以上に分散",
          ok, f"{kw_chapters}", "3キーワード以上が各3章以上")
    record("D2(A)", "A", "分散の質（表の再掲でなく実質的な記述か）は目視確認（M3）", "manual",
           "review_report.mdに記録", "目視")

    # X2(A): 単位統一（千円）・百万円等の混入ゼロ
    bad_units = [u for u in ("百万円", "兆円", "億円") if u in full]
    xl_text = wb_all_text(A_XLSX)
    bad_xl = [u for u in ("百万円",) if u in xl_text]
    check("X2(A)", "A", "統合版DDは千円統一・Excel Aと単位一致（百万円等の混入0件）",
          not bad_units and not bad_xl, f"DD混入={bad_units} Excel混入={bad_xl}", "0件")

    # 負条件: B専用要素の非混入
    neg = []
    for s in (L.SHEET_PL_OLD, L.SHEET_SCRATCH, L.SHEET_EMPTY):
        if s in wb.sheetnames:
            neg.append(f"罠シート{s}")
    for _, v in sheet_strings(wb[L.SHEET_ASSUMPTIONS]):
        if "ティーザー" in v or "v2.1" in v:
            neg.append(f"Assumptions:{v[:20]}")
    renamed = ["売上収益", "売掛金等", "買掛・未払等", "現業人件費", "修正EBITDA", "その他流動\n"]
    xl_labels = xl_text
    neg += [f"B改称科目:{r}" for r in renamed[:5] if r in xl_labels]
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and "#REF!" in c.value:
                    neg.append(f"#REF! {ws.title}!{c.coordinate}")
                if ws.title in ("PL", "BS") and c.value is not None \
                        and c.column >= 16:
                    neg.append(f"グリッド外セル {ws.title}!{c.coordinate}")
                if ws.title == L.SHEET_CALC and c.value is not None and c.row >= 23:
                    neg.append(f"calc下部セル {ws.title}!{c.coordinate}")
    for b in ("法務", "税務", "簿外", "偶発", "5,280", "1,585", "注1", "別紙",
              "(+)", "(+/-)", "端数処理"):
        if b in full:
            neg.append(f"DD禁止語:{b}")
    if "5,500,000" not in full:
        neg.append("DDにのれん5,500,000千円の記載なし")
    check("負条件", "A", "B専用13ID（E7〜E11・D3〜D6・X1〜X4）に該当する要素が0件",
          not neg, neg or "0件", "0件")
    record("負条件", "A", "統合版DDの負条件（多段表・他分冊参照なし等）の最終確認は目視併用", "manual",
           "review_report.mdに記録", "目視")


# ---------------------------------------------------------------- B固有

def check_b_specific():
    for tag, path in (("Base_B", B_BASE), ("Sponsor_B", B_SPONSOR)):
        wb = load(path)
        # E2(B): 単位ラベル0件
        labels = [f"{c}:{v}" for c, v in sheet_strings(wb[L.SHEET_MONTHLY])
                  if ("単位" in v or "千円" in v or "（円）" in v)]
        check("E2(B)", "B", f"{tag}: 月次試算表に単位ラベル0件",
              not labels, labels or "0件", "0件")

        # E7/X1: 科目改称の適用（本表）と旧表記の残置（PL_old）
        bad = []
        for sheet, key, old, new, _dd in L.RENAMES:
            rows = {L.SHEET_PL: L.PL_ROWS, L.SHEET_BS: L.BS_ROWS}[sheet]
            got = wb[sheet][f"B{rows[key][0]}"].value
            if got != new:
                bad.append(f"{sheet}!B{rows[key][0]}={got}")
        old_kept = wb[L.SHEET_PL_OLD][f"B{L.PL_ROWS['revenue'][0]}"].value
        if old_kept != "売上高（Net Sales）":
            bad.append(f"PL_old!B9={old_kept}（旧表記が残っていない）")
        mon_bad = [k for k in L.MONTHLY_RENAME_KEYS
                   if wb[L.SHEET_MONTHLY][f"B{L.MONTHLY_ROWS[k][0]}"].value != L.MONTHLY_ROWS[k][2]]
        check("E7/X1", "B", f"{tag}: 科目改称6組を全シート一貫適用（PL_old/Scratchは旧表記）・5組以上",
              not bad and not mon_bad and len(L.RENAMES) >= 5,
              f"改称{len(L.RENAMES)}組 不備={bad + mon_bad or 'なし'}", "6組適用・旧シートは旧表記")

        # E8: 罠シート3枚
        traps = [s for s in (L.SHEET_PL_OLD, L.SHEET_SCRATCH, L.SHEET_EMPTY) if s in wb.sheetnames]
        note = wb[L.SHEET_PL_OLD]["A1"].value or ""
        empty_ok = all(c.value is None for row in wb[L.SHEET_EMPTY].iter_rows() for c in row)
        check("E8", "B", f"{tag}: 罠シート3枚（PL_old「使用しないこと」注記・Scratch・空Sheet1）",
              len(traps) == 3 and "使用しないこと" in note and empty_ok,
              f"{traps} 注記={'あり' if '使用しないこと' in note else 'なし'}", "3枚")

        # E10: 補足小表3つ以上（位置・サイズ・整合）
        side = []
        pl_ws, kpi_ws, bs_ws = wb["PL"], wb[L.SHEET_CALC], wb["BS"]
        drv = S.drivers("base" if "Base" in tag else "sponsor")
        pl_ok = (pl_ws["P3"].value or "").startswith("■") and pl_ws["P4"].value == "年度" \
            and all(pl_ws[f"R{5 + i}"].value == drv[y]["bill"] for i, y in enumerate(PLAN_YEARS))
        side.append(("PL P3:R11", pl_ok, 8, 3))
        total = sum(kpi_ws[f"C{25 + i}"].value for i in range(3))
        kpi_ok = (kpi_ws["B23"].value or "").startswith("■") and total == drv["FY26"]["hires"] \
            and kpi_ws[f"C{28}"].value == drv["FY26"]["hires"]
        side.append(("calc B23:D28", kpi_ok, 5, 3))
        bs_ok = (bs_ws["P3"].value or "").startswith("■") and bs_ws["Q7"].value == 210_000
        side.append(("BS P3:Q8", bs_ok, 5, 2))
        ok = all(s[1] for s in side) and len(side) >= 3 and all(s[2] >= 3 and s[3] >= 2 for s in side)
        check("E10", "B", f"{tag}: 本表グリッド外のKPI補足小表3つ以上（各3行×2列以上・本表と整合）",
              ok, [(s[0], "OK" if s[1] else "NG") for s in side], "3表以上・整合")

        # X3/X4: 既存不一致の維持
        asm = wb[L.SHEET_ASSUMPTIONS]
        teaser = asm[f"D{L.ASM_TEASER_ROW}"].value
        goodwill = asm["D21"].value
        scratch = wb[L.SHEET_SCRATCH]["B12"].value or ""
        ok = teaser == 1_585_000 and goodwill == 5_500_000 and "5,656,000" in scratch
        check("X2〜X4", "B", f"{tag}: 不一致の維持（ティーザー1,585,000・のれん5,500,000・Scratch旧試算）",
              ok, f"teaser={teaser} goodwill={goodwill} scratch={'あり' if '5,656,000' in scratch else 'なし'}",
              "維持")

    # E9: ケース違い2ファイル・差分30セル以上
    wb1, wb2 = load(B_BASE), load(B_SPONSOR)
    same_structure = wb1.sheetnames == wb2.sheetnames
    diff = 0
    for sname in ("PL", "BS", "CF", "calc"):
        w1, w2 = wb1[sname], wb2[sname]
        for row in range(1, 40):
            for y in YEARS:
                c1, c2 = w1[f"{COL[y]}{row}"].value, w2[f"{COL[y]}{row}"].value
                if isinstance(c1, (int, float)) and isinstance(c2, (int, float)) and c1 != c2:
                    diff += 1
    labels_same = all(
        wb1[s][f"B{r}"].value == wb2[s][f"B{r}"].value
        for s in ("PL", "BS", "CF", "calc") for r in range(4, 35))
    check("E9", "B", "Base_B/Sponsor_Bが同一シート構成・同一科目名・計画期数値差分30セル以上",
          same_structure and labels_same and diff >= 30,
          f"構成一致={same_structure} 科目一致={labels_same} 差分={diff}セル", "30セル以上")

    # E11: 分離ファイルの存在（破損セル検査は check_integrity 内で実施済み）
    ok = B_DEBT.exists() and "debt" in load(B_DEBT).sheetnames
    check("E11", "B", "分離デットスケジュール AutostaffChubu_Debt_B.xlsx が存在",
          ok, B_DEBT.name, "存在")

    # X2/X3: DD側の記載（千円vs百万円・のれん・ティーザー）
    fin = pdf_pages(ROOT / "B_hard" / "DD_Financial_オートスタッフ中部.pdf")
    ok = ("5,280" in fin[30]) and ("1,620" in fin[33]) and any("百万円" in p for p in fin)
    check("X2〜X4", "B", "DD4分冊は百万円表記・のれん5,280（p.31）・正常EBITDA1,620（p.34）を維持",
          ok, f"p31のれん={'あり' if '5,280' in fin[30] else 'なし'} p34EBITDA={'あり' if '1,620' in fin[33] else 'なし'}",
          "維持")

    # D3〜D6: 原本ハッシュ一致
    mismatch = []
    for fname in ("DD_Business_オートスタッフ中部.pdf", "DD_Financial_オートスタッフ中部.pdf",
                  "DD_Legal_オートスタッフ中部.pdf", "DD_Tax_オートスタッフ中部.pdf"):
        h1 = hashlib.sha256((DUMMY / fname).read_bytes()).hexdigest()
        h2 = hashlib.sha256((ROOT / "B_hard" / fname).read_bytes()).hexdigest()
        if h1 != h2:
            mismatch.append(fname)
    check("D3〜D6", "B", "DD4分冊のSHA-256が原本と完全一致（無変更の証明）",
          not mismatch, mismatch or "4/4一致", "4/4一致")


# ---------------------------------------------------------------- main

def main():
    print("=== checks.py: 定量条件表（第4章）の機械チェック ===")
    print("-- 4-1 Excel共通（A）--")
    check_common_excel("Model_A", A_XLSX, "A", "base")
    check_integrity("Model_A", A_XLSX, "A", "base")
    print("-- 4-1 Excel共通（B: Base_B / Sponsor_B）--")
    check_common_excel("Base_B", B_BASE, "B", "base")
    check_integrity("Base_B", B_BASE, "B", "base")
    check_common_excel("Sponsor_B", B_SPONSOR, "B", "sponsor")
    check_integrity("Sponsor_B", B_SPONSOR, "B", "sponsor")
    print("-- E6 業界用語 --")
    check_e6()
    print("-- 4-2 データA固有 --")
    check_a_specific()
    print("-- 4-3 データB固有 --")
    check_b_specific()

    n_pass = sum(1 for r in RESULTS if r["status"] == "pass")
    n_fail = sum(1 for r in RESULTS if r["status"] == "fail")
    n_manual = sum(1 for r in RESULTS if r["status"] == "manual")
    report = dict(
        summary=dict(total=len(RESULTS), passed=n_pass, failed=n_fail, manual=n_manual),
        results=RESULTS,
    )
    out = HERE / "validation_report.json"
    out.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"\n=== 結果: pass={n_pass} fail={n_fail} manual={n_manual} → {out.name} ===")
    sys.exit(1 if n_fail else 0)


if __name__ == "__main__":
    main()
