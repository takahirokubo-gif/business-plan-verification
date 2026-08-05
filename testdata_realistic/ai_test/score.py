"""実AIテスト結果の採点（results/result_[AB].json を読む）。

- A：3段フォールバックの判定（計画値・FCF・正常収益力EBITDAが「抽出値」か）と
     抽出値のground_truth突合、および誤検知（Aは照会0件が期待）
- B：planted_items（期待動作=検知して照会）の検知率

使い方: backend/.venv/bin/python testdata_realistic/ai_test/score.py A|B
"""
import json
import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
RES = Path(__file__).resolve().parent / "results"
GT = ROOT / "testdata_realistic/ground_truth/ground_truth.xlsx"

PLAN_KEYS = [f"{c}_{m}" for c in ("base", "sponsor")
             for m in ("revenue", "op", "ebitda", "ni", "cash", "net_assets", "fcf")]
FCF_KEYS = ["act_fcf", "base_fcf", "sponsor_fcf"]
SPECIAL = ["normalized_ebitda"]


def load(ds):
    return json.loads((RES / f"result_{ds}.json").read_text(encoding="utf-8"))


def facts():
    ws = load_workbook(GT)["facts"]
    rows = [[c.value for c in r] for r in ws.iter_rows(min_row=2)]
    return {r[0]: dict(section=r[1], label=r[2], value=r[3], unit=r[4], ds=r[7])
            for r in rows}


def planted(ds):
    ws = load_workbook(GT)["planted_items"]
    rows = [[c.value for c in r] for r in ws.iter_rows(min_row=2)]
    return [dict(id=r[0], ds=r[1], file=r[2], sheet=r[3], cell=r[4],
                 content=r[5], expect=r[6])
            for r in rows if r[1] == ds]


def items_of(data):
    return data["full"]["items"]


def report_a():
    data = load("A")
    items = items_of(data)
    print(f"# データA 採点（{data['model']['model']}）\n")
    print(f"抽出項目 {len(items)}件・確認事項 {len(data['full']['inquiries'])}件・"
          f"KPIノード {len(data['full']['kpi_nodes'])}件・"
          f"シナリオ {len(data['full']['scenarios'])}件\n")

    print("## 1. source_type 一覧（全項目）\n")
    print("| key | label | case | unit(元) | source_type | 年度 | 根拠 |")
    print("|---|---|---|---|---|---|---|")
    for it in items:
        vals = it.get("values") or {}
        yrs = f"{min(vals)}〜{max(vals)}（{len(vals)}期）" if vals else "－"
        ev = it.get("evidence") or {}
        print(f"| {it['key']} | {it['label']} | {it.get('case') or '-'} | "
              f"{it.get('source_unit') or it.get('unit')} | **{it.get('source_type')}** | "
              f"{yrs} | {ev.get('file', '')} {ev.get('location', '')} |")

    print("\n## 2. 判定対象（計画値・FCF・正常収益力EBITDA）\n")
    by_key = {it["key"]: it for it in items}
    print("| 判定軸 | key | source_type | 結果 |")
    print("|---|---|---|---|")
    for label, keys in (("計画値", PLAN_KEYS), ("FCF", FCF_KEYS),
                        ("正常収益力EBITDA", SPECIAL)):
        for k in keys:
            it = by_key.get(k)
            if not it:
                print(f"| {label} | {k} | －（項目なし） | 未抽出 |")
                continue
            st = it.get("source_type")
            mark = "○ 抽出値" if st == "extracted" else f"× {st}"
            print(f"| {label} | {k} | {st} | {mark} |")

    print("\n## 3. 確認事項（Aの期待は0件＝誤検知）\n")
    for q in data["full"]["inquiries"]:
        print(f"- [{q['severity']}/{q['category']}] {q['title']}：{q.get('detail', '')[:120]}")
    if not data["full"]["inquiries"]:
        print("（なし）")


def report_b():
    data = load("B")
    inq = data["full"]["inquiries"]
    items = items_of(data)
    print(f"# データB 採点（{data['model']['model']}）\n")
    print(f"抽出項目 {len(items)}件・確認事項 {len(inq)}件・"
          f"KPIノード {len(data['full']['kpi_nodes'])}件・"
          f"シナリオ {len(data['full']['scenarios'])}件\n")

    print("## 1. 出力された確認事項の全文\n")
    for i, q in enumerate(inq, 1):
        src = q.get("source") or {}
        print(f"### {i}. [{q['severity']}/{q['category']}] {q['title']}")
        print(f"- 内容: {q.get('detail')}")
        print(f"- 根拠: {src.get('file', '')} {src.get('location', '')}")
        print(f"- 質問案: {q.get('suggested_question')}\n")

    print("\n## 2. planted（検知して照会）一覧\n")
    for i, p in enumerate([p for p in planted("B") if "照会" in str(p["expect"])], 1):
        print(f"{i}. [{p['id']}] {p['file']} {p['sheet']} {p['cell']}: {p['content']}")

    print("\n## 3. mismatch を付けた抽出項目\n")
    for it in items:
        if it.get("mismatch"):
            m = it["mismatch"]
            print(f"- {it['key']}（{it['label']}）: 相手 {m.get('other_value')} "
                  f"{m.get('other_unit') or ''} @ {m.get('other_file')} "
                  f"{m.get('other_location')} — {m.get('note')}")


if __name__ == "__main__":
    ds = sys.argv[1].upper()
    (report_a if ds == "A" else report_b)()
