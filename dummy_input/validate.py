# -*- coding: utf-8 -*-
"""ダミーインプット一式の整合チェック。

検証対象:
 1. spec.py の主要ピン（正本数値）
 2. Excelモデル：ピンセルの値・数式・単位注記・ノイズシート
 3. DDレポートPDF：キーファクトのページ位置とページ数
 4. mockフィクスチャ：抽出値がspecの計算値と一致すること
 5. GROUND_TRUTH.md が spec から再生成した内容と一致すること
"""
import json
import sys
from pathlib import Path

from openpyxl import load_workbook
from pypdf import PdfReader

import generate_fixtures
import spec
import xl_layout as L
from spec import ACTUAL_YEARS, PLAN_YEARS, to_mm

HERE = Path(__file__).parent
FIXTURES = HERE.parent / "backend" / "app" / "fixtures"

errors = []


def check(cond, msg):
    if cond:
        print(f"  OK  {msg}")
    else:
        errors.append(msg)
        print(f"  NG  {msg}")


def validate_spec():
    print("[1] spec.py 主要ピン")
    spec.summary_check()
    print("  OK  summary_check（14,160/1,620・14,890/1,750・4.1x/54%・KPI・BSバランス）")


def validate_excel():
    print("[2] Excelモデル")
    expected_order = [L.SHEET_COVER, L.SHEET_ASSUMPTIONS, L.SHEET_PL, L.SHEET_BS,
                      L.SHEET_CF, L.SHEET_KPI, L.SHEET_DEBT,
                      L.SHEET_PL_OLD, L.SHEET_SCRATCH, L.SHEET_EMPTY]
    for case in spec.CASES:
        path = HERE / spec.MODEL_FILES[case]
        check(path.exists(), f"{path.name} が存在する")
        wb = load_workbook(path)
        check(wb.sheetnames == expected_order,
              f"{case}: シート構成（ノイズ含む）が正しい {wb.sheetnames}")
        pl = spec.compute_pl(case)
        ws = wb[L.SHEET_PL]
        check("千円" in str(ws["A2"].value), f"{case}: PLシートに単位注記（千円）")
        c26 = spec.COL["FY26"]
        check(ws[f"{c26}{L.PL_ROWS['staffing_rev'][0]}"].value == pl["FY26"]["staffing_rev"],
              f"{case}: PL FY26派遣事業売上（値貼付）= {pl['FY26']['staffing_rev']:,}千円")
        c27 = spec.COL["FY27"]
        f = ws[f"{c27}{L.PL_ROWS['staffing_rev'][0]}"].value
        check(isinstance(f, str) and "KPI_Drivers" in f and "*12/1000" in f,
              f"{case}: PL FY27派遣事業売上は KPI_Drivers 参照の数式（{f}）")
        f = ws[f"{c27}{L.PL_ROWS['labor'][0]}"].value
        check(isinstance(f, str) and f.count("KPI_Drivers") == 4,
              f"{case}: PL FY27労務費は4変数の数式")
        ws_bs = wb[L.SHEET_BS]
        check(ws_bs[f"{c27}{L.BS_ROWS['goodwill'][0]}"].value == spec.GOODWILL,
              f"{case}: BS FY27のれん = {spec.GOODWILL:,}千円")
        ws_k = wb[L.SHEET_KPI]
        d = spec.drivers(case)["FY26"]
        check(ws_k[f"{c26}{L.KPI_ROWS['util'][0]}"].value == d["util"],
              f"{case}: KPI FY26稼働率 = {d['util']:.0%}")
        check(ws_k[f"{c26}{L.KPI_ROWS['bill'][0]}"].value == d["bill"],
              f"{case}: KPI FY26派遣単価 = {d['bill']:,}円/h")
        check(ws_k[f"{c26}{L.KPI_ROWS['cpa'][0]}"].value == d["cpa"],
              f"{case}: KPI FY26採用CPA = {d['cpa']}千円")
        f = ws_k[f"{c26}{L.KPI_ROWS['active'][0]}"].value
        check(isinstance(f, str) and f.startswith("=ROUND("),
              f"{case}: KPI稼働人数は =ROUND(在籍×稼働率) の数式")
        ws_a = wb[L.SHEET_ASSUMPTIONS]
        check(ws_a["D8"].value == spec.DEAL["ev_mm"] * 1000,
              f"{case}: Assumptions EV = 12,000,000千円")
        check(ws_a["D21"].value == spec.GOODWILL,
              f"{case}: Assumptions のれん想定 = {spec.GOODWILL:,}千円")
        cover = wb[L.SHEET_COVER]
        check(spec.CASE_LABEL[case] in str(cover["B8"].value),
              f"{case}: Coverにケース名（{spec.CASE_LABEL[case]}）")


def validate_pdfs():
    print("[3] DDレポートPDF")
    expected_pages = {"business": 20, "financial": 36, "legal": 14, "tax": 10}
    readers = {}
    for key, fname in spec.DD_FILES.items():
        path = HERE / fname
        check(path.exists(), f"{fname} が存在する")
        readers[key] = PdfReader(path)
        check(len(readers[key].pages) == expected_pages[key],
              f"{fname}: {expected_pages[key]}ページ構成")
    for fact in spec.DD_KEY_FACTS:
        r = readers[fact["file"]]
        text = (r.pages[fact["page"] - 1].extract_text() or "").replace("\n", "")
        for phrase in fact["check"]:
            check(phrase in text,
                  f"{spec.DD_FILES[fact['file']]} p.{fact['page']} に「{phrase[:28]}…」")
    # キーファクトが指定ページ以外に（照合フレーズの全文が）重複して存在しないこと
    for fact in spec.DD_KEY_FACTS:
        if fact["id"] != "normalized_ebitda":
            continue
        r = readers[fact["file"]]
        hits = [i + 1 for i, p in enumerate(r.pages)
                if fact["check"][0] in (p.extract_text() or "").replace("\n", "")]
        check(hits == [fact["page"]],
              f"正常収益力EBITDAの結論文言は p.{fact['page']} のみに存在（{hits}）")


def validate_fixtures():
    print("[4] mockフィクスチャ")
    ext = json.loads((FIXTURES / "extraction_autostaff.json").read_text(encoding="utf-8"))
    pl_b = spec.compute_pl("base")
    pl_s = spec.compute_pl("sponsor")
    bs_b, cf_b = spec.compute_bs_cf("base")
    _, cf_s = spec.compute_bs_cf("sponsor")
    items = {it["key"]: it for it in ext["items"]}
    check(len(ext["items"]) == 24, f"抽出項目は24件（{len(ext['items'])}件）")
    check(items["act_revenue"]["values"] == {y: to_mm(pl_b[y]["revenue"]) for y in ACTUAL_YEARS},
          "act_revenue = 実績売上（12,460/13,363/14,160）")
    check(items["act_ebitda"]["values"]["FY26"] == 1620, "act_ebitda FY26 = 1,620")
    check(items["base_revenue"]["values"]["FY27"] == 14890, "base_revenue FY27 = 14,890")
    check(items["base_ebitda"]["values"]["FY27"] == 1750, "base_ebitda FY27 = 1,750")
    check(items["sponsor_ebitda"]["values"] == {y: to_mm(pl_s[y]["ebitda"]) for y in PLAN_YEARS},
          "sponsor_ebitda = スポンサー計画EBITDA")
    check(items["base_fcf"]["values"] == {y: to_mm(cf_b[y]["fcf"]) for y in PLAN_YEARS},
          "base_fcf = ベースFCF")
    gw = items["goodwill"]
    check(gw["values"]["FY27"] == to_mm(spec.GOODWILL)
          and gw["mismatch"]["other_value"] == to_mm(spec.GOODWILL_DD),
          "goodwill: モデル5,500 vs DD5,280 の不整合データ")
    # 参照ファイルが実在すること
    all_files = set(spec.DD_FILES.values()) | set(spec.MODEL_FILES.values())
    for it in ext["items"]:
        f = it["evidence"]["file"]
        check(f in all_files, f"{it['key']}: 参照ファイル {f} が実在する")
    # DDページ参照がキーファクト定義と一致
    facts = {f["id"]: f for f in spec.DD_KEY_FACTS}
    check(f"p.{facts['customer_concentration']['page']}" in items["risk_concentration"]["evidence"]["location"],
          "risk_concentration の参照ページがキーファクト定義と一致")
    check(f"p.{facts['normalized_ebitda']['page']}" in items["normalized_ebitda"]["evidence"]["location"],
          "normalized_ebitda の参照ページがキーファクト定義と一致")
    check(f"p.{facts['goodwill_dd']['page']}" in gw["mismatch"]["other_location"],
          "goodwill不整合の参照ページがキーファクト定義と一致")
    # KPIツリー
    tree = json.loads((FIXTURES / "kpi_tree_autostaff.json").read_text(encoding="utf-8"))
    ids = {n["id"] for n in tree["nodes"]}
    check("concentration" not in ids, "提案ツリーに『大口依存度』は含まれない（チャットで追加）")
    stars = {n["id"] for n in tree["nodes"] if n["star"]}
    check(stars == {"util", "bill", "cpa"}, f"★は稼働率・派遣単価・採用CPA（{stars}）")
    parents = {n["id"]: n["parent"] for n in tree["nodes"]}
    check(parents["active"] == "staffing_rev" and parents["util"] == "active",
          "ツリー構造：売上→派遣売上→稼働人数→稼働率")
    # シナリオ
    sc = json.loads((FIXTURES / "scenarios_autostaff.json").read_text(encoding="utf-8"))
    keys = [c["key"] for c in sc["cards"]]
    check(keys == ["A", "B", "C"], "AI推奨シナリオはA/B/Cの3類型")
    check([c["adopted"] for c in sc["cards"]] == [True, True, False], "採用状態：A採用/B採用/C不採用")
    check("0.9倍" in sc["cards"][0]["impact"], "シナリオA: DSCR 0.9倍前後（AI推定）")
    check("5.0x" in sc["cards"][1]["impact"], "シナリオB: レバレッジ5.0x超（AI推定）")
    check("枯渇" in sc["cards"][2]["impact"], "シナリオC: 現預金枯渇リスク（AI推定）")
    check(sc["human_card"]["key"] == "D" and "28%" in sc["human_card"]["title"],
          "人の仮説シナリオD: 大口派遣先（構成比28%）")
    # チャット台本
    chat = json.loads((FIXTURES / "chat_scripts.json").read_text(encoding="utf-8"))
    check(len(chat["kpi"]["suggestions"]) >= 2 and len(chat["scenario"]["suggestions"]) >= 3,
          "チャットのプレースホルダー候補（KPI2件以上・シナリオ3件以上）")
    check(chat["kpi"]["scripts"][0]["diff"]["node"]["id"] == "concentration",
          "KPIチャット台本1: 大口依存度ノードの追加差分")
    # identify
    ident = json.loads((FIXTURES / "identify.json").read_text(encoding="utf-8"))
    check(set(ident["files"].keys()) == all_files, "identify: 6ファイルすべてに対応")


def validate_ground_truth():
    print("[5] GROUND_TRUTH.md")
    current = (HERE / "GROUND_TRUTH.md").read_text(encoding="utf-8")
    regenerated = generate_fixtures.build_ground_truth(generate_fixtures.build_extraction())
    check(current == regenerated, "GROUND_TRUTH.md は spec からの再生成と一致（手編集なし）")


def main():
    validate_spec()
    validate_excel()
    validate_pdfs()
    validate_fixtures()
    validate_ground_truth()
    print()
    if errors:
        print(f"NG: {len(errors)}件の不整合")
        for e in errors:
            print(f"  - {e}")
        sys.exit(1)
    print("すべての整合チェックに合格")


if __name__ == "__main__":
    main()
