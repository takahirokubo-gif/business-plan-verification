"""財務モデルExcel（Sponsor / Base 2ケース）の生成。

- 金額は千円。各シートに「（単位：千円）」を明記（単位ラベル欠落による桁誤り対策）
- 実績年（FY24-26）は値貼り付け、計画年（FY27-31）は実際の数式を書き込む
  （KPIツリーはこの数式パースで再現できることが要件）
- ノイズ：Cover / Scratch / PL_old（旧版）/ 空シートを混在させる
- 項目名の表記揺れ（Net Sales / Adj. EBITDA / Operating Profit 等）を意図的に含む
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import spec
from spec import ACTUAL_YEARS, CASE_LABEL, COL, PLAN_YEARS, YEARS
import xl_layout as L

OUT_DIR = Path(__file__).parent

TITLE_FONT = Font(name="Yu Gothic", size=14, bold=True, color="1A4F8B")
HEADER_FONT = Font(name="Yu Gothic", size=10, bold=True)
BASE_FONT = Font(name="Yu Gothic", size=10)
NOTE_FONT = Font(name="Yu Gothic", size=9, color="737781")
HEADER_FILL = PatternFill("solid", fgColor="EDEDF3")
ACTUAL_FILL = PatternFill("solid", fgColor="F5F5F5")
THIN = Side(style="thin", color="C2C6D1")
BORDER = Border(bottom=THIN)

NUM_FMT = "#,##0"
PCT_FMT = "0.0%"


def _sheet_header(ws, title: str, unit_note: str = "（単位：千円）",
                  case_label: str | None = None):
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws["A2"] = unit_note
    ws["A2"].font = NOTE_FONT
    if case_label:
        ws["H2"] = f"ケース: {case_label}"
        ws["H2"].font = HEADER_FONT
    ws[f"{COL['FY24']}{L.HEADER_ROW}"]  # noqa: 触って初期化するだけ
    for y in YEARS:
        c = ws[f"{COL[y]}{L.HEADER_ROW}"]
        c.value = y
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")
        k = ws[f"{COL[y]}{L.KIND_ROW}"]
        k.value = "実績" if y in ACTUAL_YEARS else "計画"
        k.font = NOTE_FONT
        k.alignment = Alignment(horizontal="center")
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 34
    for y in YEARS:
        ws.column_dimensions[COL[y]].width = 13


def _put_row(ws, row_def, values_or_formulas: dict, fmt=NUM_FMT, indent=0):
    """1行ぶんを書く。values_or_formulas は {year: value|'=formula'}。"""
    row, label = row_def
    lc = ws[f"B{row}"]
    lc.value = ("　" * indent) + label
    lc.font = BASE_FONT
    for y in YEARS:
        c = ws[f"{COL[y]}{row}"]
        v = values_or_formulas.get(y)
        if v is None:
            continue
        c.value = v
        c.number_format = fmt
        c.font = BASE_FONT
        if y in ACTUAL_YEARS:
            c.fill = ACTUAL_FILL


def _mixed(formula_tpl: str, actual_values: dict) -> dict:
    """実績年は値、計画年は数式（列記号を差し込み）にする。"""
    out = {}
    for y in ACTUAL_YEARS:
        out[y] = actual_values[y]
    for y in PLAN_YEARS:
        out[y] = formula_tpl.format(c=COL[y])
    return out


def _all_formula(formula_tpl: str, years=YEARS) -> dict:
    return {y: formula_tpl.format(c=COL[y]) for y in years}


def build_workbook(case: str) -> Workbook:
    drv = spec.drivers(case)
    pl = spec.compute_pl(case)
    bs, cf = spec.compute_bs_cf(case)
    sched = spec.debt_schedule()

    wb = Workbook()
    wb.remove(wb.active)

    # ---------------- Cover
    ws = wb.create_sheet(L.SHEET_COVER)
    ws["B4"] = "Project Gear"
    ws["B4"].font = Font(name="Yu Gothic", size=24, bold=True, color="1A4F8B")
    ws["B6"] = "株式会社オートスタッフ中部　財務モデル"
    ws["B6"].font = Font(name="Yu Gothic", size=16, bold=True)
    ws["B8"] = f"ケース：{CASE_LABEL[case]}（{spec.CASE_LABEL_JA[case]}）"
    ws["B8"].font = Font(name="Yu Gothic", size=12, bold=True)
    ws["B10"] = "作成：日本橋キャピタルパートナーズ株式会社"
    ws["B11"] = "作成日：2026年6月30日（Ver. 3.2）"
    ws["B13"] = "【取扱厳重注意】"
    ws["B14"] = ("本資料は貴行における本件検討のためにのみ作成されたものであり、"
                 "作成者の事前の書面による同意なく第三者への開示・複製を禁じます。")
    ws["B15"] = ("本資料に含まれる将来予測は多くの前提・仮定に基づくものであり、"
                 "その実現性を保証するものではありません。")
    for r in range(10, 16):
        ws[f"B{r}"].font = NOTE_FONT if r >= 13 else BASE_FONT
    ws.column_dimensions["B"].width = 90

    # ---------------- Assumptions
    ws = wb.create_sheet(L.SHEET_ASSUMPTIONS)
    _sheet_header(ws, "前提条件（Assumptions）", case_label=CASE_LABEL[case])
    rows = [
        ("ストラクチャー", None, None),
        ("エンタープライズ・バリュー（EV）", 12_000_000, NUM_FMT),
        ("エクイティ出資", 5_500_000, NUM_FMT),
        ("シニアローン総額", 6_500_000, NUM_FMT),
        ("　うち参加金融機関A行 想定取組額", 2_500_000, NUM_FMT),
        ("ローン期間（年）", 7, "0"),
        ("適用金利（TIBOR+200bp想定）", 0.022, PCT_FMT),
        ("約定弁済（年・最終回残額一括）", 650_000, NUM_FMT),
        ("クローズ想定", "2026年9月末", None),
        ("", None, None),
        ("オペレーティング前提", None, None),
        ("実効税率", 0.31, PCT_FMT),
        ("売上債権回転日数", 45, "0"),
        ("買収時現金調整（取引費用等）", -750_000, NUM_FMT),
        ("のれん想定額（暫定PPA前）", spec.GOODWILL, NUM_FMT),
        ("", None, None),
        ("スポンサー提示EBITDA（速報・ティーザー記載）", spec.DEAL["sponsor_ebitda_mm"] * 1000, NUM_FMT),
    ]
    r = 7
    for label, value, fmt in rows:
        ws[f"B{r}"] = label
        ws[f"B{r}"].font = HEADER_FONT if value is None and label else BASE_FONT
        if value is not None:
            c = ws[f"D{r}"]
            c.value = value
            if fmt:
                c.number_format = fmt
            c.font = BASE_FONT
        r += 1
    ws["B24"] = "（参考）旧前提 v2.1：シニア7,000,000／金利2.4%　※現行版では使用しない"
    ws["B24"].font = NOTE_FONT
    ws.column_dimensions["B"].width = 44
    ws.column_dimensions["D"].width = 16

    # ---------------- KPI_Drivers
    ws = wb.create_sheet(L.SHEET_KPI)
    _sheet_header(ws, "KPIドライバー（Operating Assumptions）",
                  "（単位：名・時間・円・％）", CASE_LABEL[case])
    K = L.KPI_ROWS
    _put_row(ws, K["enrolled"], {y: drv[y]["enrolled"] for y in YEARS})
    _put_row(ws, K["util"], {y: drv[y]["util"] for y in YEARS}, PCT_FMT)
    _put_row(ws, K["active"],
             _all_formula("=ROUND({c}%d*{c}%d,0)" % (K["enrolled"][0], K["util"][0])))
    _put_row(ws, K["hours"], {y: drv[y]["hours"] for y in YEARS})
    _put_row(ws, K["bill"], {y: drv[y]["bill"] for y in YEARS})
    _put_row(ws, K["wage"], {y: drv[y]["wage"] for y in YEARS})
    _put_row(ws, K["welfare"], {y: drv[y]["welfare"] for y in YEARS}, PCT_FMT)
    _put_row(ws, K["hires"], {y: drv[y]["hires"] for y in YEARS})
    _put_row(ws, K["cpa"], {y: drv[y]["cpa"] for y in YEARS})
    _put_row(ws, K["attrition"], {y: drv[y]["attrition"] for y in YEARS}, PCT_FMT)
    ws["B18"] = "※ 稼働人数 = 在籍登録スタッフ数 × 稼働率（四捨五入）"
    ws["B18"].font = NOTE_FONT

    # ---------------- Debt_Schedule
    ws = wb.create_sheet(L.SHEET_DEBT)
    _sheet_header(ws, "デットスケジュール（シニアローン）", case_label=CASE_LABEL[case])
    D = L.DEBT_ROWS
    opening = {}
    for i, y in enumerate(PLAN_YEARS):
        opening[y] = (SENIOR := spec.SENIOR_TOTAL) if i == 0 else \
            f"={COL[PLAN_YEARS[i - 1]]}{D['closing'][0]}"
    _put_row(ws, D["opening"], opening)
    _put_row(ws, D["repayment"], {y: spec.ANNUAL_REPAYMENT for y in PLAN_YEARS})
    _put_row(ws, D["closing"],
             _all_formula("={c}%d-{c}%d" % (D["opening"][0], D["repayment"][0]), PLAN_YEARS))
    _put_row(ws, D["avg"],
             _all_formula("=({c}%d+{c}%d)/2" % (D["opening"][0], D["closing"][0]), PLAN_YEARS))
    _put_row(ws, D["rate"], {y: spec.INTEREST_RATE for y in PLAN_YEARS}, PCT_FMT)
    _put_row(ws, D["interest"],
             _all_formula("=ROUND({c}%d*{c}%d,0)" % (D["avg"][0], D["rate"][0]), PLAN_YEARS))
    ws["B14"] = "※ 約定弁済は年650,000千円、最終回（FY33）残額一括。実績年欄は対象外。"
    ws["B14"].font = NOTE_FONT

    # ---------------- PL
    ws = wb.create_sheet(L.SHEET_PL)
    _sheet_header(ws, "損益計算書（Projected P&L）", case_label=CASE_LABEL[case])
    P = L.PL_ROWS
    K = L.KPI_ROWS
    _put_row(ws, P["staffing_rev"], _mixed(
        "=ROUND(KPI_Drivers!{c}%d*KPI_Drivers!{c}%d*KPI_Drivers!{c}%d*12/1000,0)"
        % (K["active"][0], K["hours"][0], K["bill"][0]),
        {y: pl[y]["staffing_rev"] for y in ACTUAL_YEARS}), indent=0)
    _put_row(ws, P["other_revenue"], {y: pl[y]["other_revenue"] for y in YEARS})
    _put_row(ws, P["revenue"],
             _all_formula("=SUM({c}%d:{c}%d)" % (P["staffing_rev"][0], P["other_revenue"][0])))
    _put_row(ws, P["labor"], _mixed(
        "=ROUND(KPI_Drivers!{c}%d*KPI_Drivers!{c}%d*KPI_Drivers!{c}%d*(1+KPI_Drivers!{c}%d)*12/1000,0)"
        % (K["active"][0], K["hours"][0], K["wage"][0], K["welfare"][0]),
        {y: pl[y]["labor"] for y in ACTUAL_YEARS}))
    _put_row(ws, P["other_cogs"], {y: pl[y]["other_cogs"] for y in YEARS})
    _put_row(ws, P["cogs"],
             _all_formula("={c}%d+{c}%d" % (P["labor"][0], P["other_cogs"][0])))
    _put_row(ws, P["gross"],
             _all_formula("={c}%d-{c}%d" % (P["revenue"][0], P["cogs"][0])))
    _put_row(ws, P["gross_margin"],
             _all_formula("={c}%d/{c}%d" % (P["gross"][0], P["revenue"][0])), PCT_FMT)
    _put_row(ws, P["recruiting"], _mixed(
        "=KPI_Drivers!{c}%d*KPI_Drivers!{c}%d" % (K["hires"][0], K["cpa"][0]),
        {y: pl[y]["recruiting"] for y in ACTUAL_YEARS}))
    _put_row(ws, P["hq_cost"], {y: pl[y]["hq_cost"] for y in YEARS})
    _put_row(ws, P["other_sga"], {y: pl[y]["other_sga"] for y in YEARS})
    _put_row(ws, P["sga"],
             _all_formula("=SUM({c}%d:{c}%d)" % (P["recruiting"][0], P["other_sga"][0])))
    _put_row(ws, P["op"],
             _all_formula("={c}%d-{c}%d" % (P["gross"][0], P["sga"][0])))
    _put_row(ws, P["op_margin"],
             _all_formula("={c}%d/{c}%d" % (P["op"][0], P["revenue"][0])), PCT_FMT)
    _put_row(ws, P["depreciation"], {y: pl[y]["depreciation"] for y in YEARS})
    _put_row(ws, P["ebitda"],
             _all_formula("={c}%d+{c}%d" % (P["op"][0], P["depreciation"][0])))
    _put_row(ws, P["ebitda_margin"],
             _all_formula("={c}%d/{c}%d" % (P["ebitda"][0], P["revenue"][0])), PCT_FMT)
    _put_row(ws, P["interest"], _mixed(
        "=Debt_Schedule!{c}%d" % L.DEBT_ROWS["interest"][0],
        {y: pl[y]["interest"] for y in ACTUAL_YEARS}))
    _put_row(ws, P["non_op"], {y: pl[y]["non_op"] for y in YEARS})
    _put_row(ws, P["ordinary"],
             _all_formula("={c}%d-{c}%d+{c}%d" % (P["op"][0], P["interest"][0], P["non_op"][0])))
    _put_row(ws, P["tax"],
             _all_formula("=ROUND({c}%d*0.31,0)" % P["ordinary"][0]))
    _put_row(ws, P["ni"],
             _all_formula("={c}%d-{c}%d" % (P["ordinary"][0], P["tax"][0])))
    for key in ("revenue", "gross", "op", "ebitda", "ordinary", "ni"):
        row = P[key][0]
        for y in YEARS:
            ws[f"{COL[y]}{row}"].font = HEADER_FONT
        ws[f"B{row}"].font = HEADER_FONT

    # ---------------- BS
    ws = wb.create_sheet(L.SHEET_BS)
    _sheet_header(ws, "貸借対照表（Projected B/S）", case_label=CASE_LABEL[case])
    B = L.BS_ROWS
    _put_row(ws, B["cash"], {y: bs[y]["cash"] for y in YEARS})
    _put_row(ws, B["ar"], {y: bs[y]["ar"] for y in YEARS})
    _put_row(ws, B["oca"], {y: bs[y]["oca"] for y in YEARS})
    _put_row(ws, B["current_assets"],
             _all_formula("=SUM({c}%d:{c}%d)" % (B["cash"][0], B["oca"][0])))
    _put_row(ws, B["ppe"], {y: bs[y]["ppe"] for y in YEARS})
    _put_row(ws, B["goodwill"], {y: bs[y]["goodwill"] for y in YEARS})
    _put_row(ws, B["intangible"], {y: bs[y]["intangible"] for y in YEARS})
    _put_row(ws, B["fixed_assets"],
             _all_formula("=SUM({c}%d:{c}%d)" % (B["ppe"][0], B["intangible"][0])))
    _put_row(ws, B["total_assets"],
             _all_formula("={c}%d+{c}%d" % (B["current_assets"][0], B["fixed_assets"][0])))
    _put_row(ws, B["ap"], {y: bs[y]["ap"] for y in YEARS})
    _put_row(ws, B["other_liab"], {y: bs[y]["other_liab"] for y in YEARS})
    _put_row(ws, B["debt"], _mixed(
        "=Debt_Schedule!{c}%d" % L.DEBT_ROWS["closing"][0],
        {y: bs[y]["debt"] for y in ACTUAL_YEARS}))
    _put_row(ws, B["total_liab"],
             _all_formula("=SUM({c}%d:{c}%d)" % (B["ap"][0], B["debt"][0])))
    _put_row(ws, B["net_assets"], {y: bs[y]["net_assets"] for y in YEARS})
    _put_row(ws, B["total_le"],
             _all_formula("={c}%d+{c}%d" % (B["total_liab"][0], B["net_assets"][0])))
    ws["B26"] = ("※ FY27以降は買収後連結（のれん計上・買収時調整を純資産に含む）。"
                 "のれんは暫定PPA前の想定額。")
    ws["B26"].font = NOTE_FONT
    for key in ("total_assets", "total_le", "net_assets"):
        ws[f"B{B[key][0]}"].font = HEADER_FONT

    # ---------------- CF
    ws = wb.create_sheet(L.SHEET_CF)
    _sheet_header(ws, "キャッシュフロー計算書（Projected C/F）", case_label=CASE_LABEL[case])
    C = L.CF_ROWS
    _put_row(ws, C["ni"], _all_formula("=PL!{c}%d" % L.PL_ROWS["ni"][0]))
    _put_row(ws, C["depreciation"], _all_formula("=PL!{c}%d" % L.PL_ROWS["depreciation"][0]))
    wc_values = {}
    for y in YEARS:
        wc_values[y] = cf[y]["op_cf"] - pl[y]["ni"] - pl[y]["depreciation"]
    _put_row(ws, C["wc"], wc_values)
    _put_row(ws, C["op_cf"],
             _all_formula("=SUM({c}%d:{c}%d)" % (C["ni"][0], C["wc"][0])))
    _put_row(ws, C["inv_cf"], {y: cf[y]["inv_cf"] for y in YEARS})
    _put_row(ws, C["fcf"],
             _all_formula("={c}%d+{c}%d" % (C["op_cf"][0], C["inv_cf"][0])))
    repay = {}
    fin_other = {}
    for y in ACTUAL_YEARS:
        repay[y] = -50_000
        fin_other[y] = cf[y]["fin_cf"] + 50_000
    for y in PLAN_YEARS:
        repay[y] = f"=-Debt_Schedule!{COL[y]}{L.DEBT_ROWS['repayment'][0]}"
        fin_other[y] = cf[y]["fin_cf"] + spec.ANNUAL_REPAYMENT
    _put_row(ws, C["repay"], repay)
    _put_row(ws, C["fin_other"], fin_other)
    _put_row(ws, C["fin_cf"],
             _all_formula("={c}%d+{c}%d" % (C["repay"][0], C["fin_other"][0])))
    _put_row(ws, C["net_change"],
             _all_formula("={c}%d+{c}%d" % (C["fcf"][0], C["fin_cf"][0])))
    opening = {"FY24": 1_300_000, "FY27": spec.POST_CLOSE_CASH}
    for i, y in enumerate(YEARS):
        if y in opening:
            continue
        prev = YEARS[i - 1]
        opening[y] = f"={COL[prev]}{C['closing_cash'][0]}"
    _put_row(ws, C["opening_cash"], opening)
    _put_row(ws, C["closing_cash"],
             _all_formula("={c}%d+{c}%d" % (C["net_change"][0], C["opening_cash"][0])))
    ws["B24"] = "※ FY27期首現金は買収時調整後（取引費用等750,000千円控除後）。"
    ws["B24"].font = NOTE_FONT

    # ---------------- ノイズ: PL_old（旧版・値のみ）
    ws = wb.create_sheet(L.SHEET_PL_OLD)
    _sheet_header(ws, "損益計算書【旧版 v2.1】", case_label=CASE_LABEL[case])
    ws["A1"] = "損益計算書【旧版 v2.1】※使用しないこと（FY26速報反映前）"
    ws["A1"].font = Font(name="Yu Gothic", size=14, bold=True, color="BA1A1A")
    for key in ("staffing_rev", "other_revenue", "revenue", "labor", "cogs",
                "gross", "sga", "op", "depreciation", "ebitda", "ni"):
        row, label = L.PL_ROWS[key]
        ws[f"B{row}"] = label
        ws[f"B{row}"].font = BASE_FONT
        for y in YEARS:
            base_v = pl[y][key]
            c = ws[f"{COL[y]}{row}"]
            c.value = int(base_v * 0.965)
            c.number_format = NUM_FMT
            c.font = BASE_FONT

    # ---------------- ノイズ: Scratch
    ws = wb.create_sheet(L.SHEET_SCRATCH)
    ws["A1"] = "（作業用シート・成果物ではありません）"
    ws["A1"].font = NOTE_FONT
    scratch = [
        ("B3", "チェック: 3,344×166×2,100×12 ="),
        ("D3", 13_988_620_800),
        ("B4", "→ PL派遣売上と一致確認済（6/25 IY）"),
        ("B6", "TODO: 労務費の福利率 13.8%→14.0%に更新（6/28 反映済）"),
        ("B8", "単価改定メモ: トヨタ系 4月改定 +1.2%規定 / デンソー系 10月"),
        ("B10", 4_233_600),
        ("C10", "←年間換算係数（166h×2,125×12）ミス。使わない"),
        ("B12", "旧のれん試算 5,656,000 → PPA仮 5,500,000"),
    ]
    for addr, v in scratch:
        ws[addr] = v
        ws[addr].font = BASE_FONT

    # ---------------- ノイズ: 空シート
    wb.create_sheet(L.SHEET_EMPTY)

    # シート順を実物らしい順序に整える（PL/BS/CFが前、ドライバー・ノイズが後）
    order = [L.SHEET_COVER, L.SHEET_ASSUMPTIONS, L.SHEET_PL, L.SHEET_BS, L.SHEET_CF,
             L.SHEET_KPI, L.SHEET_DEBT, L.SHEET_PL_OLD, L.SHEET_SCRATCH, L.SHEET_EMPTY]
    wb._sheets = [wb[name] for name in order]

    return wb


def main():
    for case in spec.CASES:
        wb = build_workbook(case)
        path = OUT_DIR / spec.MODEL_FILES[case]
        wb.save(path)
        print(f"generated: {path.name}")


if __name__ == "__main__":
    main()
