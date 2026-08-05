"""単位正規化（services/units）と数式評価エンジン（services/xl_eval）のテスト。"""
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from openpyxl import Workbook  # noqa: E402

from app.services.units import normalize_values  # noqa: E402
from app.services.xl_eval import XlEvaluator, excel_round  # noqa: E402


def test_normalize_thousand_yen_to_million():
    values, unit, source_unit = normalize_values({"FY26": 14_160_000}, "千円")
    assert values == {"FY26": 14_160.0}
    assert unit == "百万円"
    assert source_unit == "千円"


def test_normalize_million_passthrough():
    values, unit, source_unit = normalize_values({"FY26": 14_160}, "百万円")
    assert values == {"FY26": 14_160}
    assert unit == "百万円"
    assert source_unit is None


def test_normalize_man_yen_to_million():
    values, unit, source_unit = normalize_values({"FY26": 12_345}, "万円")
    assert values == {"FY26": 123.45}
    assert unit == "百万円"
    assert source_unit == "万円"


def test_normalize_yen_keeps_one_yen_precision():
    values, _, source_unit = normalize_values({"FY26": 1_234_567}, "円")
    assert values == {"FY26": 1.234567}
    assert source_unit == "円"


def test_normalize_non_money_unchanged():
    values, unit, source_unit = normalize_values({"FY26": 88}, "%")
    assert values == {"FY26": 88}
    assert unit == "%"
    assert source_unit is None


def test_excel_round_half_away_from_zero():
    assert excel_round(0.5) == 1
    assert excel_round(-0.5) == -1
    assert excel_round(2.345, 2) == 2.35


def _wb():
    wb = Workbook()
    ws = wb.active
    ws.title = "PL"
    ws["C7"] = 100
    ws["C8"] = 23
    ws["C9"] = "=SUM(C7:C8)"
    ws["C10"] = "=-ROUND(C9*0.31,0)"
    ws["C11"] = "=C9+C10"
    ws2 = wb.create_sheet("calc")
    ws2["C5"] = "=PL!C9*2"
    ws["D9"] = "=#REF!D12"
    return wb


def test_evaluator_computes_formula_chain():
    ev = XlEvaluator(_wb())
    assert ev.value("PL", "C9") == 123
    assert ev.value("PL", "C10") == -38          # ROUND(123*0.31)=38 → 負値
    assert ev.value("PL", "C11") == 85
    assert ev.value("calc", "C5") == 246         # シート間参照


def test_evaluator_marks_broken_ref():
    ev = XlEvaluator(_wb())
    assert ev.value("PL", "D9") is None
    assert ev.is_broken("PL", "D9")
    assert ("PL", "D9") in ev.broken


def _sum_wb(a2_value):
    """A1=1・A2=可変・A3=SUM(A1:A2) のブックを作る。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    ws["A1"] = 1
    ws["A2"] = a2_value
    ws["A3"] = "=SUM(A1:A2)"
    return wb


def test_sum_ignores_text_and_empty_cells():
    """文字列定数・空セルはExcelのSUMと同じく無視する。"""
    assert XlEvaluator(_sum_wb("メモ")).value("S", "A3") == 1.0
    assert XlEvaluator(_sum_wb(None)).value("S", "A3") == 1.0
    assert XlEvaluator(_sum_wb(2)).value("S", "A3") == 3.0


def test_sum_with_unevaluable_formula_is_not_partial():
    """未対応関数を含む数式セルがあるとき、部分和を返さず計算不能にする。

    部分和を「計算値」として返すと、実際より小さい値が信頼できる値として
    AIに渡ってしまうため。
    """
    assert XlEvaluator(_sum_wb("=IF(A1>0,10,0)")).value("S", "A3") is None


def test_sum_with_broken_ref_is_not_partial():
    """参照切れ（#REF!）を含む範囲も同様に計算不能にする。"""
    assert XlEvaluator(_sum_wb("=#REF!B1")).value("S", "A3") is None
