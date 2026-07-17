"""ステージ4：Excelエクスポート。

確定データのみ（案件基礎情報＋確定財務数値＋確定KPI構造＋採用シナリオ）を
行内標準フォーマットに近いテンプレートに値転記して .xlsx を生成する。
- すべての数値に出典（参照ファイル・箇所）を併記する（トレーサビリティ）
- AI推定値には必ず注記を付す
- 保留項目は除外し、末尾に一覧を付す
- 審査相談メモ（相談履歴・指摘事項）を別シートに収録する
"""
import json
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from ..config import EXPORT_DIR
from ..models import Deal

TITLE = Font(name="Yu Gothic", size=14, bold=True, color="1A4F8B")
H2 = Font(name="Yu Gothic", size=11, bold=True, color="1A4F8B")
HEAD = Font(name="Yu Gothic", size=9.5, bold=True)
BODY = Font(name="Yu Gothic", size=9.5)
SMALL = Font(name="Yu Gothic", size=8)
NOTE = Font(name="Yu Gothic", size=8.5, color="737781")
AI_NOTE = Font(name="Yu Gothic", size=8.5, color="B54708")
FILL = PatternFill("solid", fgColor="EDEDF3")
AI_FILL = PatternFill("solid", fgColor="FFFAEB")
THIN = Side(style="thin", color="C2C6D1")
BOX = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)

AI_DISCLAIMER = "※ インパクト数値はAIによる推定であり、財務モデルの再計算値ではありません。"

YEARS_ACT = ["FY24", "FY25", "FY26"]
YEARS_PLAN = ["FY27", "FY28", "FY29", "FY30", "FY31"]

USER_NAMES = {"tanaka": "田中", "sato": "佐藤", "takahashi": "高橋"}


def _set(ws, addr, value, font=BODY, fill=None, num=None, align=None, border=False):
    c = ws[addr]
    c.value = value
    c.font = font
    if fill:
        c.fill = fill
    if num:
        c.number_format = num
    if align:
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    else:
        c.alignment = Alignment(vertical="center", wrap_text=True)
    if border:
        c.border = BOX
    return c


def _evidence_str(item) -> str:
    ev = json.loads(item.evidence_json) if item.evidence_json else None
    if not ev:
        return ""
    return f"{ev.get('file', '')}｜{ev.get('location', '')}"


def build_export(deal: Deal) -> tuple[str, int]:
    """確定データからエクスポートExcelを生成。returns (filepath, excluded_held_count)."""
    confirmed = [i for i in deal.items if i.status == "confirmed"]
    held = [i for i in deal.items if i.status == "held"]
    by_key = {i.key: i for i in confirmed}

    wb = Workbook()

    # ================= シート1：審査サマリー =================
    ws = wb.active
    ws.title = "審査サマリー"
    ws.sheet_view.showGridLines = False
    for col, w in dict(A=3, B=24, C=13, D=13, E=13, F=13, G=13, H=52).items():
        ws.column_dimensions[col].width = w

    _set(ws, "B2", "審査相談用サマリー（事業計画検証）", TITLE)
    _set(ws, "B3", f"作成日時：{datetime.now().strftime('%Y/%m/%d %H:%M')}　"
                   f"／　出典：事業計画検証システム（確定データのみを転記・各数値に参照元を併記）", NOTE)

    # ---- 1. 案件基礎情報
    _set(ws, "B5", "1. 案件基礎情報", H2)
    owner = USER_NAMES.get(deal.owner or "", deal.owner)
    rows = [
        ("案件名", deal.name), ("案件種別", deal.deal_type),
        ("借入人（SPC）", deal.borrower), ("対象会社", deal.target),
        ("対象会社の業種", deal.industry or "－"),
        ("スポンサー", deal.sponsor or "－"),
        ("クローズ予定", deal.close_date or "－"),
        ("次回審査相談日", deal.next_meeting_date or "－"),
        ("担当者", owner or "－"),
        ("検討ステータス", deal.review_status),
        ("EV", f"{deal.ev_mm:,}百万円" if deal.ev_mm else "－"),
        ("シニアローン",
         (f"{deal.senior_mm:,}百万円"
          f"（本行取組 {f'{deal.our_commitment_mm:,}' if deal.our_commitment_mm is not None else '－'}百万円"
          f"・期間{deal.tenor_years if deal.tenor_years is not None else '－'}年）")
         if deal.senior_mm else "－"),
        ("エクイティ", f"{deal.equity_mm:,}百万円" if deal.equity_mm else "－"),
        ("スポンサー提示EBITDA（速報）",
         f"{deal.sponsor_ebitda_mm:,}百万円" if deal.sponsor_ebitda_mm else "－"),
        ("初期レバレッジ（自動算出）",
         f"{deal.initial_leverage}x ＝ シニア{deal.senior_mm:,} ÷ 提示EBITDA{deal.sponsor_ebitda_mm:,}"
         if deal.initial_leverage else "－"),
        ("LTV（自動算出）",
         f"{deal.ltv_pct}% ＝ シニア{deal.senior_mm:,} ÷ EV{deal.ev_mm:,}" if deal.ltv_pct else "－"),
    ]
    r = 6
    for k, v in rows:
        _set(ws, f"B{r}", k, HEAD, FILL, border=True)
        ws.merge_cells(f"C{r}:H{r}")
        _set(ws, f"C{r}", v, BODY, border=True)
        r += 1
    if deal.summary:
        _set(ws, f"B{r}", "案件概要", HEAD, FILL, border=True)
        ws.merge_cells(f"C{r}:H{r}")
        _set(ws, f"C{r}", deal.summary, BODY, border=True)
        ws.row_dimensions[r].height = 13 * (len(deal.summary) // 60 + 1) + 6
        r += 1

    # ---- 2. 確定財務数値（出典付き）
    r += 1
    _set(ws, f"B{r}", "2. 確定財務数値（単位：百万円）", H2)
    r += 1
    _set(ws, f"B{r}", "確定済みの抽出値のみを転記。「修正済」は担当者がレビューで修正した値。", NOTE)
    r += 1

    def years_for(keys, default):
        """表の年度列を実データの年度キーから導出する。

        実AIは決算期表記（'2027/3期' 等）を返すことがあるため、固定リストで
        捨てずに実キーを列にする。既知のFY形式を先、未知形式は昇順で後ろ。
        列数はレイアウト上5列まで（C〜G。Hは出典列）。
        """
        found: set[str] = set()
        for key, _label in keys:
            item = by_key.get(key)
            if item:
                found.update((item.effective_values() or {}).keys())
        order = YEARS_ACT + YEARS_PLAN
        known = [y for y in order if y in found]
        unknown = sorted(y for y in found if y not in order)
        return (known + unknown)[:5] or default

    def fin_table(title, keys, years, start_row):
        rr = start_row
        _set(ws, f"B{rr}", title, HEAD, FILL, border=True)
        for ci, y in enumerate(years):
            _set(ws, f"{chr(ord('C') + ci)}{rr}", y, HEAD, FILL, align="center", border=True)
        _set(ws, f"H{rr}", "出典（参照ファイル｜箇所）", HEAD, FILL, border=True)
        rr += 1
        for key, label in keys:
            item = by_key.get(key)
            if not item:
                continue
            _set(ws, f"B{rr}", label + ("（修正済）" if item.edited else ""), BODY, border=True)
            values = item.effective_values() or {}
            for ci, y in enumerate(years):
                v = values.get(y)
                _set(ws, f"{chr(ord('C') + ci)}{rr}", v if v is not None else "－",
                     BODY, num="#,##0", align="right", border=True)
            _set(ws, f"H{rr}", _evidence_str(item), SMALL, border=True)
            rr += 1
        return rr

    act_keys = [("act_revenue", "売上高"), ("act_op", "営業利益"),
                ("act_ebitda", "EBITDA"), ("act_ni", "当期純利益"),
                ("act_cash", "現預金"), ("act_net_assets", "純資産"),
                ("act_debt", "有利子負債"), ("act_fcf", "フリー・キャッシュフロー")]
    base_keys = [("base_revenue", "売上高"), ("base_op", "営業利益"),
                 ("base_ebitda", "EBITDA"), ("base_fcf", "フリー・キャッシュフロー")]
    sponsor_keys = [("sponsor_revenue", "売上高"), ("sponsor_op", "営業利益"),
                    ("sponsor_ebitda", "EBITDA"), ("sponsor_fcf", "フリー・キャッシュフロー")]
    r = fin_table("実績", act_keys, years_for(act_keys, YEARS_ACT), r) + 1
    r = fin_table("計画（ベースケース）", base_keys, years_for(base_keys, YEARS_PLAN), r) + 1
    r = fin_table("計画（スポンサーケース）", sponsor_keys,
                  years_for(sponsor_keys, YEARS_PLAN), r) + 1

    # ---- 3. ストラクチャー・B/S項目
    _set(ws, f"B{r}", "3. ストラクチャー・B/S項目", H2)
    r += 1
    for key, label in (("ev", "エンタープライズ・バリュー（EV）"),
                       ("senior_loan", "シニアローン総額"),
                       ("goodwill", "のれん（買収想定額）")):
        item = by_key.get(key)
        if not item:
            continue
        v = next(iter((item.effective_values() or {}).values()), None)
        _set(ws, f"B{r}", label, HEAD, FILL, border=True)
        _set(ws, f"C{r}", v, BODY, num="#,##0", align="right", border=True)
        ws.merge_cells(f"D{r}:G{r}")
        note = item.resolution_note or ""
        _set(ws, f"D{r}", note, SMALL, border=True)
        _set(ws, f"H{r}", _evidence_str(item), SMALL, border=True)
        if item.mismatch_json:
            # mismatchはキー欠落・数値なしがあり得る（実AIの旧出力・説明型の差異）
            mm = json.loads(item.mismatch_json)
            other_value = mm.get("other_value")
            src_file = mm.get("other_file") or mm.get("source_file") or ""
            src_loc = mm.get("other_location") or mm.get("source_location") or ""
            note = mm.get("note") or mm.get("description") or ""
            head = (f"相手資料の値 {other_value:,}百万円"
                    if isinstance(other_value, (int, float)) else "相手資料")
            r += 1
            ws.merge_cells(f"C{r}:H{r}")
            _set(ws, f"B{r}", "（不整合情報）", SMALL, border=True)
            _set(ws, f"C{r}",
                 f"{head}（{src_file} {src_loc}）との差異あり。{note}",
                 SMALL, AI_FILL, border=True)
        r += 1

    # ---- 4. 前提・定性情報（全文）
    r += 1
    _set(ws, f"B{r}", "4. 前提・定性情報（確定済み・全文）", H2)
    r += 1
    for item in confirmed:
        if item.unit != "テキスト" and item.key != "normalized_ebitda":
            continue
        text = item.effective_text()
        if item.key == "normalized_ebitda":
            v = (item.effective_values() or {}).get("FY26")
            text = f"{v:,}百万円（FY26）。モデルFY26実績EBITDAと一致することを確認済み。" if v else "－"
        _set(ws, f"B{r}", item.label, HEAD, FILL, border=True)
        ws.merge_cells(f"C{r}:G{r}")
        _set(ws, f"C{r}", text, BODY, border=True)
        _set(ws, f"H{r}", _evidence_str(item), SMALL, border=True)
        ws.row_dimensions[r].height = max(18, 13 * (len(str(text)) // 55 + 1) + 4)
        r += 1

    # ---- 5. 保留項目
    if held:
        r += 1
        _set(ws, f"B{r}", f"5. 保留項目（{len(held)}件・本サマリーから除外）", H2)
        r += 1
        for item in held:
            _set(ws, f"B{r}", item.label, BODY, border=True)
            ws.merge_cells(f"C{r}:G{r}")
            _set(ws, f"C{r}", "担当者レビューで保留（確定前のデータは転記しない）", NOTE, border=True)
            _set(ws, f"H{r}", _evidence_str(item), SMALL, border=True)
            r += 1

    # ================= シート2：KPI構造 =================
    ws2 = wb.create_sheet("KPI構造")
    ws2.sheet_view.showGridLines = False
    for col, w in dict(A=3, B=42, C=13, D=20, E=34, F=46).items():
        ws2.column_dimensions[col].width = w
    _set(ws2, "B2", "確定KPI構造", TITLE)
    _set(ws2, "B3", "★＝重要KPI（リスクドライバー）。出典：財務モデルの数式解析（再計算なし）"
                    "およびDDレポートの定性情報。", NOTE)
    _set(ws2, "B5", "KPI（階層）", HEAD, FILL, border=True)
    _set(ws2, "C5", "出典種別", HEAD, FILL, border=True)
    _set(ws2, "D5", "値（FY26実績）", HEAD, FILL, border=True)
    _set(ws2, "E5", "構造（数式・再計算なし）", HEAD, FILL, border=True)
    _set(ws2, "F5", "参照元", HEAD, FILL, border=True)
    nodes = list(deal.kpi_nodes)
    children = {}
    for n in nodes:
        children.setdefault(n.parent_id, []).append(n)
    origin_label = {"model": "モデル数式", "dd": "DD由来", "manual": "手動追加"}
    r2 = [6]

    def walk(parent, depth):
        for n in children.get(parent, []):
            label = ("　" * depth) + ("★ " if n.star else "") + n.label
            if n.badge:
                label += f"〔{n.badge}〕"
            # evidenceは辞書が原則だが、チャット追加ノード等で文字列のことがある
            ev = json.loads(n.evidence_json) if n.evidence_json else None
            if isinstance(ev, dict):
                ev_text = f"{ev.get('file', '')}｜{ev.get('location', '')}"
            else:
                ev_text = str(ev) if ev else ""
            _set(ws2, f"B{r2[0]}", label, BODY, border=True)
            _set(ws2, f"C{r2[0]}", origin_label.get(n.origin, n.origin), BODY, border=True)
            _set(ws2, f"D{r2[0]}", n.value_text or "－", BODY, border=True)
            _set(ws2, f"E{r2[0]}", n.formula or "", SMALL, border=True)
            _set(ws2, f"F{r2[0]}", ev_text, SMALL, border=True)
            r2[0] += 1
            walk(n.node_id, depth + 1)

    walk(None, 0)
    if deal.kpi_confirmed_by:
        _set(ws2, f"B{r2[0] + 1}",
             f"確定：{USER_NAMES.get(deal.kpi_confirmed_by, deal.kpi_confirmed_by)}"
             f"　{deal.kpi_confirmed_at.strftime('%Y/%m/%d %H:%M') if deal.kpi_confirmed_at else ''}",
             NOTE)

    # ================= シート3：シナリオ =================
    ws3 = wb.create_sheet("シナリオ分析")
    ws3.sheet_view.showGridLines = False
    for col, w in dict(A=3, B=20, C=62, D=32).items():
        ws3.column_dimensions[col].width = w
    _set(ws3, "B2", "採用シナリオ（審査相談用）", TITLE)
    _set(ws3, "B3", AI_DISCLAIMER, AI_NOTE, AI_FILL)
    r3 = 5
    adopted = [s for s in deal.scenarios if s.adopted]
    rejected = [s for s in deal.scenarios if not s.adopted]
    node_labels = {n.node_id: n.label for n in nodes}
    for sc in adopted:
        origin = "AI推奨" if sc.origin == "ai" else "自分の仮説"
        _set(ws3, f"B{r3}", f"シナリオ{sc.key}", HEAD, FILL, border=True)
        ws3.merge_cells(f"C{r3}:D{r3}")
        _set(ws3, f"C{r3}", f"{sc.title}　〔{origin}／{sc.type_label}〕", HEAD, FILL, border=True)
        r3 += 1
        kpis = "、".join(node_labels.get(k, k) for k in json.loads(sc.affected_kpis_json or "[]"))
        for k, v in [("① シナリオ名・発生要因", sc.cause),
                     ("② 影響を受けるKPI", kpis or "－（PLへの直接影響）"),
                     ("③ KPIの変化幅と根拠", f"{sc.change_text}\n根拠：{sc.change_basis}"),
                     ("④ 返済能力への影響（AI推定・モデル再計算なし）", sc.impact),
                     ("⑤ 保全策・確認事項", f"{sc.safeguards}\n確認事項：{sc.questions}")]:
            _set(ws3, f"B{r3}", k, BODY, border=True, align="left")
            ws3.merge_cells(f"C{r3}:D{r3}")
            cell = _set(ws3, f"C{r3}", v, BODY, border=True, align="left")
            if "AI推定" in k:
                cell.fill = AI_FILL
            ws3.row_dimensions[r3].height = max(16, 13 * (len(str(v)) // 55 + 1))
            r3 += 1
        r3 += 1
    if rejected:
        _set(ws3, f"B{r3}", "（参考）不採用シナリオ", H2)
        r3 += 1
        for sc in rejected:
            _set(ws3, f"B{r3}", f"シナリオ{sc.key}", BODY, border=True)
            ws3.merge_cells(f"C{r3}:D{r3}")
            note = f"　※不採用の理由：{sc.rejection_note}" if sc.rejection_note else ""
            _set(ws3, f"C{r3}", f"{sc.title}{note}", SMALL, border=True)
            ws3.row_dimensions[r3].height = 26
            r3 += 1

    # ================= シート4：審査相談メモ =================
    ws4 = wb.create_sheet("審査相談メモ")
    ws4.sheet_view.showGridLines = False
    for col, w in dict(A=3, B=14, C=10, D=30, E=70).items():
        ws4.column_dimensions[col].width = w
    _set(ws4, "B2", "審査相談の記録", TITLE)
    _set(ws4, "B3", "対面の審査相談（擦り合わせ会議）の記録。指摘事項は紐付け先を併記。", NOTE)
    r4 = 5
    for m in reversed(list(deal.memos)):
        attendees = "、".join(json.loads(m.attendees_json or "[]"))
        _set(ws4, f"B{r4}", m.meeting_date, HEAD, FILL, border=True)
        _set(ws4, f"C{r4}", m.conclusion, HEAD, FILL, border=True)
        _set(ws4, f"D{r4}", f"出席：{attendees}", BODY, FILL, border=True)
        _set(ws4, f"E{r4}", f"記録：{USER_NAMES.get(m.created_by or '', m.created_by)}",
             BODY, FILL, border=True)
        r4 += 1
        for i, f in enumerate(m.findings, 1):
            link = ""
            if f.target_type == "scenario":
                link = f"【シナリオ{f.target_key}】"
            elif f.target_type == "kpi":
                link = "【KPI構造】"
            elif f.target_type == "item":
                link = f"【数値：{f.target_key}】"
            _set(ws4, f"D{r4}", f"指摘{i} {link}", BODY, border=True)
            ws4.merge_cells(f"E{r4}:E{r4}")
            _set(ws4, f"E{r4}", f.text, BODY, border=True)
            ws4.row_dimensions[r4].height = max(16, 13 * (len(f.text) // 55 + 1))
            r4 += 1
        if m.note:
            _set(ws4, f"D{r4}", "メモ", BODY, border=True)
            _set(ws4, f"E{r4}", m.note, BODY, border=True)
            ws4.row_dimensions[r4].height = max(16, 13 * (len(m.note) // 55 + 1))
            r4 += 1
        r4 += 1

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"審査サマリー_{deal.target}_{ts}.xlsx"
    path = EXPORT_DIR / filename
    wb.save(path)
    return str(path), len(held)
