"""検証用ダンプ：AI解析の生結果をそのままExcelに書き出す。

通常のエクスポート（export_xlsx）は「必須項目の確定」が前提のため、
確認事項が多く確定が進まない検証データでは出力できない。
テスト・採点作業のために、確定状態によらず解析結果の全量を出す。

- 抽出項目（値・単位・取得方法・根拠・不整合）
- 確認事項（照会）
- KPIツリー
- シナリオ
- 解析ログ（出力の打ち切り・エラーの有無）
"""
import json
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..config import EXPORT_DIR
from ..models import Deal

HEAD = Font(name="Yu Gothic", size=9.5, bold=True)
BODY = Font(name="Yu Gothic", size=9.5)
FILL = PatternFill("solid", fgColor="EDEDF3")
WARN_FILL = PatternFill("solid", fgColor="FDECEA")

SOURCE_TYPE_LABEL = {
    "extracted": "抽出値（資料の記載）",
    "calculated": "計算値（資料内の数式から算出）",
    "estimated": "AI推定",
    "missing": "未取得",
}


def _sheet(wb, title: str, headers: list[str], widths: list[int]):
    ws = wb.create_sheet(title)
    ws.sheet_view.showGridLines = False
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = HEAD
        c.fill = FILL
        c.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    return ws


def _put(ws, row: int, values: list, fill=None):
    for i, v in enumerate(values, start=1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = BODY
        c.alignment = Alignment(vertical="top", wrap_text=True)
        if fill:
            c.fill = fill


def build_raw_export(deal: Deal) -> str:
    """確定状態によらず解析結果の全量をExcelに書き出す。returns filepath."""
    wb = Workbook()
    wb.remove(wb.active)

    # ---- 抽出項目
    ws = _sheet(wb, "抽出項目",
                ["key", "セクション", "項目名", "ケース", "単位", "取得方法", "元単位",
                 "年度→値", "テキスト", "必須", "状態", "根拠ファイル", "根拠箇所",
                 "根拠引用", "抽出の論理", "不整合(mismatch)"],
                [22, 16, 26, 8, 8, 22, 8, 40, 50, 6, 10, 30, 22, 40, 50, 50])
    for r, it in enumerate(deal.items, start=2):
        ev = json.loads(it.evidence_json) if it.evidence_json else {}
        mm = json.loads(it.mismatch_json) if it.mismatch_json else None
        values = it.effective_values() or {}
        _put(ws, r, [
            it.key, it.section, it.label, it.case_name or "", it.unit,
            SOURCE_TYPE_LABEL.get(it.source_type, it.source_type), it.source_unit or "",
            json.dumps(values, ensure_ascii=False) if values else "",
            it.effective_text() or "", "必須" if it.required else "任意", it.status,
            ev.get("file", ""), ev.get("location", ""), ev.get("quote", ""), ev.get("logic", ""),
            json.dumps(mm, ensure_ascii=False) if mm else "",
        ], fill=WARN_FILL if (mm or it.source_type == "estimated") else None)

    # ---- 確認事項（照会）
    ws = _sheet(wb, "確認事項",
                ["状態", "重要度", "分類", "確認事項", "内容", "根拠ファイル", "根拠箇所",
                 "根拠引用", "確認質問案", "対応メモ"],
                [10, 8, 14, 32, 60, 30, 22, 40, 50, 30])
    sev = {"high": "高", "medium": "中", "low": "低"}
    for r, q in enumerate(deal.inquiries, start=2):
        src = json.loads(q.source_json) if q.source_json else {}
        _put(ws, r, [
            "確認済" if q.status == "resolved" else "未確認",
            sev.get(q.severity, q.severity), q.category, q.title, q.detail or "",
            src.get("file", ""), src.get("location", ""), src.get("quote", ""),
            q.suggested_question or "", q.resolution_note or "",
        ])

    # ---- KPIツリー
    ws = _sheet(wb, "KPIツリー",
                ["ノードID", "親ID", "ラベル", "由来", "重要", "数式", "値", "バッジ",
                 "根拠ファイル", "根拠箇所", "抽出の論理"],
                [12, 12, 30, 8, 6, 50, 20, 12, 30, 22, 50])
    for r, n in enumerate(deal.kpi_nodes, start=2):
        ev = json.loads(n.evidence_json) if n.evidence_json else {}
        _put(ws, r, [
            n.node_id, n.parent_id or "", n.label, n.origin, "★" if n.star else "",
            n.formula or "", n.value_text or "", n.badge or "",
            ev.get("file", ""), ev.get("location", ""), ev.get("logic", ""),
        ])

    # ---- シナリオ
    ws = _sheet(wb, "シナリオ",
                ["key", "由来", "類型", "タイトル", "発生要因", "影響KPI", "変化幅",
                 "変化の根拠", "インパクト", "保全策", "確認事項", "採用"],
                [8, 8, 14, 30, 40, 20, 30, 40, 60, 40, 40, 8])
    for r, s in enumerate(deal.scenarios, start=2):
        _put(ws, r, [
            s.key, s.origin, s.type_label or "", s.title, s.cause or "",
            "、".join(json.loads(s.affected_kpis_json or "[]")),
            s.change_text or "", s.change_basis or "", s.impact or "",
            s.safeguards or "", s.questions or "", "採用" if s.adopted else "",
        ])

    # ---- 解析ログ（出力の打ち切り・エラーの切り分け用）
    ws = _sheet(wb, "解析ログ",
                ["実行時刻", "ステップ", "状態", "エンジン", "モデル", "入力トークン",
                 "出力トークン", "出力上限", "停止理由", "所要(秒)", "結果", "エラー"],
                [18, 12, 20, 12, 24, 12, 12, 10, 14, 10, 30, 60])
    for r, run in enumerate(deal.analysis_runs, start=2):
        near_limit = (run.output_tokens is not None and run.max_tokens is not None
                      and run.output_tokens >= run.max_tokens * 0.95)
        _put(ws, r, [
            run.at.strftime("%Y/%m/%d %H:%M:%S") if run.at else "",
            run.step, run.status, run.mode or "", run.model or "",
            run.input_tokens, run.output_tokens, run.max_tokens,
            run.stop_reason or "",
            round(run.duration_ms / 1000, 1) if run.duration_ms else None,
            run.result_summary or "", run.error or "",
        ], fill=WARN_FILL if (run.status != "ok" or near_limit) else None)

    # ---- サマリー（先頭に置く）
    ws = wb.create_sheet("サマリー", 0)
    ws.sheet_view.showGridLines = False
    for col, w in dict(A=3, B=30, C=60).items():
        ws.column_dimensions[col].width = w
    ws["B2"] = f"検証用ダンプ（確定前の生結果）　{deal.name}"
    ws["B2"].font = Font(name="Yu Gothic", size=13, bold=True)
    ws["B3"] = f"出力日時：{datetime.now().strftime('%Y/%m/%d %H:%M')}"
    ws["B3"].font = Font(name="Yu Gothic", size=9, color="737781")
    rows = [
        ("抽出項目", f"{len(deal.items)}件"),
        ("　うちAI推定", f"{len([i for i in deal.items if i.source_type == 'estimated'])}件"),
        ("　うち計算値", f"{len([i for i in deal.items if i.source_type == 'calculated'])}件"),
        ("　うち不整合あり", f"{len([i for i in deal.items if i.mismatch_json])}件"),
        ("確認事項（照会）", f"{len(deal.inquiries)}件"
                       f"（未確認 {len([q for q in deal.inquiries if q.status == 'open'])}件）"),
        ("KPIノード", f"{len(deal.kpi_nodes)}件"),
        ("シナリオ", f"{len(deal.scenarios)}件"),
        ("解析ログ", f"{len(deal.analysis_runs)}件"
                 f"（異常 {len([r for r in deal.analysis_runs if r.status != 'ok'])}件）"),
    ]
    r = 5
    for k, v in rows:
        ws[f"B{r}"] = k
        ws[f"B{r}"].font = HEAD if not k.startswith("　") else BODY
        ws[f"C{r}"] = v
        ws[f"C{r}"].font = BODY
        r += 1
    r += 1
    ws[f"B{r}"] = "※ 確定状態によらず解析結果の全量を出力しています（検証・採点用）。"
    ws[f"B{r}"].font = Font(name="Yu Gothic", size=9, color="737781")

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"検証ダンプ_{deal.target or deal.name}_{ts}.xlsx"
    path = EXPORT_DIR / filename
    wb.save(path)
    return str(path)
