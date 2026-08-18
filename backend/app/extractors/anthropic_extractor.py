"""AnthropicExtractor：Anthropic API による実抽出の実装。

設計原則（仕様 ③「数値取得の3段フォールバック」）:
1. 資料の記載値（Excelキャッシュ値・DDの表）を抽出する（source_type=extracted）
2. 記載が無くても資料内の数式から一意に定まる値は、システム側の数式評価エンジン
   （services/xl_eval）が「計算値」として資料ダイジェストに埋め込む。
   AI自身は再計算しない（source_type=calculated）
3. どちらも無い場合のみ推定し、推定であることを必ず明示する（source_type=estimated）

その他の原則:
- すべての出力に根拠（参照ファイル・箇所・論理）を必須とするスキーマを強制する
- 構造化出力は tool_use（input_schema強制）で受け取る
- 単位の換算はしない（元資料の単位のまま返し、正規化はシステム側の責務）
- 人に確認すべき事象（単位不明・名寄せ・参照切れ・資料間矛盾等）は inquiries として報告する
"""
import hashlib
import json
from pathlib import Path

from ..config import ANTHROPIC_API_KEY
from .base import Extractor
from .factory import get_model

EVIDENCE_SCHEMA = {
    "type": "object",
    "description": "根拠3点セット。すべての抽出値に必須。",
    "properties": {
        "file": {"type": "string", "description": "参照ファイル名"},
        "location": {"type": "string",
                     "description": "箇所（Excelはシート・セル番地、PDFはページ番号）"},
        "quote": {"type": "string", "description": "該当箇所の原文抜粋"},
        "logic": {"type": "string",
                  "description": "抽出の論理（なぜこの箇所をこの項目にマッピングしたか）"},
    },
    "required": ["file", "location", "quote", "logic"],
}

INQUIRY_SCHEMA = {
    "type": "object",
    "description": "人への確認事項（照会）。機械が勝手に解決してはいけない事象を報告する。",
    "properties": {
        "category": {"type": "string",
                     "enum": ["単位不明", "名寄せ", "参照切れ", "資料間矛盾",
                              "正シート識別", "正ファイル識別", "作成時点差", "その他"]},
        "severity": {"type": "string", "enum": ["high", "medium", "low"],
                     "description": "high=数値の採否に直結／medium=解釈に影響／low=軽微"},
        "title": {"type": "string", "description": "確認事項の見出し（30字以内）"},
        "detail": {"type": "string",
                   "description": "事象の内容・両論併記（どちらの値/解釈があり得るか）・審査上の論点"},
        "source": EVIDENCE_SCHEMA,
        "suggested_question": {"type": "string",
                               "description": "スポンサー・対象会社への確認質問文の案"},
    },
    "required": ["category", "severity", "title", "detail", "source", "suggested_question"],
}

ITEMS_SCHEMA = {
    "type": "object",
    "properties": {
        "items": {
            "type": "array",
            "items": {
                "type": "object",
                "properties": {
                    "key": {"type": "string"},
                    "section": {"type": "string"},
                    "label": {"type": "string"},
                    "unit": {"type": "string",
                             "description": "元資料の単位表記のまま（例: 千円・百万円・円・%・名・円/h）。"
                                            "換算しない（単位の正規化はシステム側で行う）。"
                                            "単価・比率は必ず分母まで書く（例: 千円/名・円/h）。"
                                            "分母を省くと金額合計と誤認され百万円へ誤換算される"},
                    "case": {"type": ["string", "null"], "enum": ["base", "sponsor", None]},
                    "values": {"type": ["object", "null"],
                               "description": "年度→数値（元資料の単位のまま・換算しない）。定性項目はnull"},
                    "text_value": {"type": ["string", "null"]},
                    "required": {"type": "boolean"},
                    "source_type": {
                        "type": "string",
                        "enum": ["extracted", "calculated", "estimated", "missing"],
                        "description": "extracted=資料の記載値（キャッシュ値含む）を転記／"
                                       "calculated=ダイジェスト中の『計算値』（システムが数式から算出）を使用／"
                                       "estimated=推定（根拠と算法をlogicに必ず書く）／"
                                       "missing=資料に無く推定もできない（values=null）"},
                    "evidence": EVIDENCE_SCHEMA,
                    "mismatch": {
                        "type": ["object", "null"],
                        "description": "資料間で値が食い違う場合の相手側情報。無ければnull",
                        "properties": {
                            "other_value": {
                                "type": ["number", "null"],
                                "description": "相手資料側の数値（相手資料の単位のまま）。"
                                               "単一の数値で表せない差異はnull"},
                            "other_unit": {"type": ["string", "null"],
                                           "description": "相手資料側の単位（例: 百万円）"},
                            "other_file": {"type": "string"},
                            "other_location": {"type": "string"},
                            "other_quote": {"type": "string"},
                            "note": {"type": "string",
                                     "description": "差異の内容と審査上の論点（どちらを採用すべきか等）"},
                        },
                        "required": ["other_value", "other_file",
                                     "other_location", "other_quote", "note"],
                    },
                },
                "required": ["key", "section", "label", "unit", "values",
                             "text_value", "required", "source_type", "evidence"],
            },
        },
        "inquiries": {
            "type": "array",
            "description": "人への確認事項（照会）の一覧。該当が無ければ空配列",
            "items": INQUIRY_SCHEMA,
        },
    },
    "required": ["items", "inquiries"],
}

KPI_TREE_SCHEMA = {
    "type": "object",
    "properties": {
        "nodes": {
            "type": "array",
            "items": {
                "type": "object",
                "properties": {
                    "id": {"type": "string"},
                    "parent": {"type": ["string", "null"]},
                    "label": {"type": "string"},
                    "origin": {"type": "string", "enum": ["model", "dd", "manual"]},
                    "star": {"type": "boolean"},
                    "formula": {"type": ["string", "null"],
                                "description": "数式の構造（人が読める形）。再計算はしない"},
                    "value_text": {"type": ["string", "null"],
                                   "description": "最新の値（例: '88%（FY26実績）'）。"
                                                  "数式はここに入れず formula に書く"},
                    "badge": {"type": ["string", "null"]},
                    "evidence": EVIDENCE_SCHEMA,
                },
                "required": ["id", "parent", "label", "origin", "star", "evidence"],
            },
        }
    },
    "required": ["nodes"],
}

SCENARIOS_SCHEMA = {
    "type": "object",
    "properties": {
        "cards": {
            "type": "array",
            "items": {
                "type": "object",
                "properties": {
                    "key": {"type": "string"},
                    "origin": {"type": "string", "enum": ["ai", "human"]},
                    "type_label": {"type": "string",
                                   "enum": ["トップライン", "コスト", "イベント", "自分の仮説"]},
                    "title": {"type": "string"},
                    "cause": {"type": "string", "description": "シナリオ名・発生要因"},
                    "affected_kpis": {"type": "array", "items": {"type": "string"},
                                      "description": "確定KPIツリーのノードIDのみを入れる"
                                                     "（ラベルや括弧書きを付けない。例: 'B31'）"},
                    "change_text": {"type": "string", "description": "KPIの変化幅"},
                    "change_basis": {"type": "string", "description": "変化幅の根拠"},
                    "impact": {"type": "string",
                               "description": "返済能力への影響。DSCR・レバレッジ・現預金等への"
                                              "影響を具体的数値を含む定性推定で記述する。"
                                              "判定ラベル（問題なし等）は出さない"},
                    "safeguards": {"type": "string", "description": "保全策・構造"},
                    "questions": {"type": "string", "description": "スポンサー等への確認事項"},
                    "impact_calc": {
                        "type": "array",
                        "description": "impactに書いた最終指標（DSCR・Net Debt/EBITDA・粗利率等）を"
                                       "左辺とし、ストレス対象KPIを含む式で分解した参考試算。"
                                       "ストレス起因で変化する部分は {{...}} で囲む（UIで強調表示される）",
                        "items": {
                            "type": "object",
                            "properties": {
                                "metric": {"type": "string", "description": "左辺の指標名（例: DSCR（FY28））"},
                                "formula": {"type": "string", "description": "指標の分解式（例: EBITDA ÷（約定弁済 ＋ 支払利息））"},
                                "before": {"type": "string",
                                           "description": "ストレス前の値を入れた計算式（例: 1,836百万円 ÷（650百万円 ＋ 122百万円）＝ 約1.9x）"},
                                "after": {"type": "string",
                                          "description": "ストレス後の値を入れた計算式。変化した値を {{}} で囲む"},
                                "note": {"type": "string",
                                         "description": "どのKPI変化がどう波及して左辺が変わるかの1文の補足。対象KPIを {{}} で囲む"},
                            },
                            "required": ["metric", "formula", "before", "after", "note"],
                        },
                    },
                },
                "required": ["key", "origin", "type_label", "title", "cause",
                             "affected_kpis", "change_text", "change_basis",
                             "impact", "safeguards", "questions", "impact_calc"],
            },
        }
    },
    "required": ["cards"],
}

IDENTIFY_SCHEMA = {
    "type": "object",
    "properties": {
        "company": {"type": "string", "description": "対象会社名"},
        "doc_type": {"type": "string",
                     "enum": ["model_sponsor", "model_base", "dd_business",
                              "dd_financial", "dd_legal", "dd_tax", "dd_integrated",
                              "unknown"]},
        "label": {"type": "string"},
        "detail": {"type": "string", "description": "判定根拠"},
    },
    "required": ["company", "doc_type", "label", "detail"],
}

DEAL_INFO_SCHEMA = {
    "type": "object",
    "properties": {
        "fields": {
            "type": "object",
            "properties": {
                "name": {"type": "string", "description": "案件名（対象会社名＋スキーム）"},
                "deal_type": {"type": "string", "enum": ["LBO", "MBO", "事業承継", "リファイナンス"]},
                "borrower": {"type": "string", "description": "借入人（SPC）"},
                "target": {"type": "string", "description": "対象会社"},
                "industry": {"type": ["string", "null"]},
                "sponsor": {"type": ["string", "null"]},
                "close_date": {"type": ["string", "null"], "description": "YYYY-MM-DD"},
                "ev_mm": {"type": ["number", "null"],
                          "description": "EV。必ず百万円単位に換算した数値を入れる"
                                         "（資料の単位が何であれ百万円に換算する。"
                                         "元の値・単位と換算の過程をsourcesに記載）"},
                "senior_mm": {"type": ["number", "null"], "description": "百万円単位に換算した数値"},
                "equity_mm": {"type": ["number", "null"], "description": "百万円単位に換算した数値"},
                "tenor_years": {"type": ["number", "null"]},
                "sponsor_ebitda_mm": {"type": ["number", "null"],
                                      "description": "スポンサー提示EBITDA（速報値）。百万円単位に換算した数値"},
                "summary": {"type": ["string", "null"], "description": "案件概要の要約（3文程度）"},
            },
            "required": ["name", "deal_type", "borrower", "target"],
        },
        "sources": {"type": "object",
                    "description": "各フィールドの出典（ファイル・シート/セルまたはページ。"
                                   "単位換算した場合は元の値・単位も書く）"},
        "note": {"type": "string"},
    },
    "required": ["fields", "sources"],
}

CHAT_SCHEMA = {
    "type": "object",
    "properties": {
        "reply": {"type": "string"},
        "diff": {
            "type": ["object", "null"],
            "description": "適用可能な変更差分。変更が無ければnull",
            "properties": {
                "type": {"type": "string",
                         "enum": ["add_node", "star_change", "add_card", "update_card"]},
                "node": {"type": ["object", "null"],
                         "description": "add_node時に必須: "
                                        "{id, parent, label, origin:'manual', star, "
                                        "badge, value_text, evidence}"},
                "remove": {"type": ["array", "null"], "items": {"type": "string"},
                           "description": "star_change時: ★を外すノードID"},
                "add": {"type": ["array", "null"], "items": {"type": "string"},
                        "description": "star_change時: ★を付けるノードID"},
                "card": {"type": ["object", "null"],
                         "description": "add_card時に必須: 標準5部構成のカード"
                                        "（key, origin:'human', type_label, title, cause, "
                                        "affected_kpis, change_text, change_basis, impact, "
                                        "safeguards, questions）"},
                "card_key": {"type": ["string", "null"],
                             "description": "update_card時に必須: 対象カードのkey"},
                "fields": {"type": ["object", "null"],
                           "description": "update_card時に必須: 変更するフィールドのみ"},
            },
            "required": ["type"],
        },
    },
    "required": ["reply", "diff"],
}

SYSTEM_PROMPT = """あなたは銀行の融資審査部門を支援する抽出・解釈エンジンです。
LBO等のストラクチャードファイナンス案件の資料（財務モデルExcel・DDレポートPDF）から、
審査相談の叩き台となる情報を抽出します。

数値取得の優先順位（3段フォールバック・絶対に守る）:
1. 資料の記載値を抽出する（Excelのキャッシュ値＝ダイジェスト中の「値」、DDの表・本文）
   → source_type=extracted
2. 記載が無い場合、ダイジェスト中の「計算値」（システムが資料内の数式から算出した値）を使う
   → source_type=calculated。あなた自身は数式の再計算をしない
3. どちらも無い場合のみ推定する → source_type=estimated。推定の算法と根拠をlogicに必ず書く。
   推定すらできない場合は values=null・source_type=missing とし、理由をlogicに書く

その他の絶対ルール:
1. 単位の換算をしない。元資料の単位のまま値を返し、unitに単位を正確に書く
   （単位の正規化はシステム側の責務。ただしスキーマが明示的に「百万円に換算」と
   指定しているフィールドはこの限りではなく、換算過程を出典に記す）
2. すべての出力に根拠3点セット（参照ファイル・箇所・抽出の論理）を付ける
3. 資料間で値が食い違う場合は、どちらかに寄せず mismatch として両方を報告する
4. シナリオのインパクトは定性推定であり、モデルの再計算値ではないことを前提に記述する。
   「問題なし／問題あり」等の判定ラベルは出さない（推定値と含意の記述まで）
5. 資料だけでは決められない事象（単位不明・名寄せ・参照切れ・資料間矛盾・
   正シート/正ファイルの識別・作成時点差）は、勝手に解決せず inquiries で人に照会する。
   逆に、根拠が資料内で一意に決まる事柄（単位ラベルが明記されている等）は照会しない
6. 必須/任意の別によらず、すべての項目について資料全体を同じ深さで探索する
   （任意項目だからといって探索を浅くしない）
7. 人が読む文章（text_value・detail・note・logic・summary・impact等）は、
   審査資料にふさわしい自然な日本語で書く。次のものを文章に混ぜてはいけない:
   - null / values=null / source_type / missing 等の内部用語やスキーマのキー名
   - 生のセル参照だけを並べた記述（例「B16」「D28」のみ）。
     セル番地・ページは evidence（location）に書き、本文では
     「前提条件シートの参加金融機関A行の欄」のように人が分かる言い方をする
   値が資料に無い場合は「資料に記載がないため取得できていない」のように
   日本語で述べる（内部表現をそのまま書かない）"""


def _file_sha(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


DD_SLOT_NAMES = ("dd_business", "dd_financial", "dd_legal", "dd_tax", "dd_integrated")

# 実物の財務モデルは300行を超えることがあるため、上限は保険としてのみ置く。
# 値の無いセルは出力しないので、上限を上げてもトークンは行数に比例しては増えない。
MAX_DIGEST_ROWS = 2000
MAX_DIGEST_COLS = 200


def _excel_digest(path: Path) -> str:
    """Excelのシート構造・ラベル・数式・値をLLM入力用のテキストにする。

    数式セルは「数式＋値」の両方を出す。値の優先順位:
    1. キャッシュ値（Excel保存時の計算結果）→「値」
    2. キャッシュ値が無い場合、システム側の数式評価エンジンで算出 →「計算値」
    3. #REF! を含む数式 →「#REF!（参照切れ）」と明示（照会の材料）
    4. 評価不能 →「計算不能」

    上限に当たって切り落とした場合は、その旨をシート末尾に明記する
    （黙って切ると「AIが見落とした」と「そもそも渡していない」が区別できなくなるため）。
    """
    from openpyxl import load_workbook

    from ..services.xl_eval import XlEvaluator
    wb = load_workbook(path)                    # 数式
    wbv = load_workbook(path, data_only=True)   # キャッシュ値（Excel保存時の計算結果）
    ev = XlEvaluator(wb)
    out = []
    for ws in wb.worksheets:
        wsv = wbv[ws.title]
        out.append(f"=== シート: {ws.title} ===")
        rows, cols = min(ws.max_row, MAX_DIGEST_ROWS), min(ws.max_column, MAX_DIGEST_COLS)
        for row in ws.iter_rows(min_row=1, max_row=rows, max_col=cols):
            cells = []
            for c in row:
                if c.value is None:
                    continue
                if isinstance(c.value, str) and c.value.startswith("="):
                    if "#REF!" in c.value:
                        cells.append(f"{c.coordinate}: {c.value!r} → #REF!"
                                     "（参照切れ。リンク先シート/ファイルの欠落）")
                        continue
                    cached = wsv[c.coordinate].value
                    if cached is not None:
                        cells.append(f"{c.coordinate}: {c.value!r} → 値 {cached!r}")
                        continue
                    computed = ev.value(ws.title, c.coordinate)
                    if computed is not None:
                        cells.append(f"{c.coordinate}: {c.value!r} → 計算値 {computed!r}")
                    else:
                        cells.append(f"{c.coordinate}: {c.value!r} → 計算不能")
                    continue
                cells.append(f"{c.coordinate}: {c.value!r}")
            if cells:
                out.append(" | ".join(cells))
        if ws.max_row > rows or ws.max_column > cols:
            out.append(f"（注意: シート「{ws.title}」は全{ws.max_row}行×{ws.max_column}列のうち"
                       f"先頭{rows}行×{cols}列のみを表示。範囲外は未読であり、"
                       "そこに記載がある可能性を排除できない）")
    return "\n".join(out)


def _pdf_digest(path: Path, max_pages: int | None = None) -> str:
    """PDF全文のテキスト化（既定は全ページ。上限指定時はその旨を明記する）。"""
    from pypdf import PdfReader
    reader = PdfReader(path)
    pages = reader.pages if max_pages is None else reader.pages[:max_pages]
    out = []
    for i, page in enumerate(pages):
        out.append(f"=== p.{i + 1} ===")
        out.append(page.extract_text() or "")
    if max_pages is not None and len(reader.pages) > max_pages:
        out.append(f"（注意: 全{len(reader.pages)}ページ中、先頭{max_pages}ページのみを表示）")
    return "\n".join(out)


def _excel_docs(documents: list[dict]) -> list[dict]:
    return [d for d in documents
            if d.get("stored_path") and str(d["stored_path"]).endswith(".xlsx")]


def _pick_model_doc(documents: list[dict]) -> dict | None:
    """KPIツリー生成に使う財務モデルを1冊選ぶ。

    スロットが付いていればそれを優先し、付いていない（種別なしで取り込んだ）
    Excelもファイル名・判定ラベルから候補にする。ケース違いのファイルが複数ある案件で、
    ベースケースらしいものを選び、デットスケジュール等の付属ファイルを避ける。
    """
    for slot in ("model_base", "model_sponsor"):
        doc = next((d for d in documents if d["slot"] == slot), None)
        if doc:
            return doc
    candidates = [d for d in _excel_docs(documents) if d["slot"] not in DD_SLOT_NAMES]
    if not candidates:
        return None

    def score(doc: dict) -> int:
        text = f"{doc.get('filename', '')} {doc.get('identified_label') or ''}".lower()
        s = 0
        if any(w in text for w in ("base", "ベース")):
            s += 2
        if any(w in text for w in ("sponsor", "スポンサー")):
            s -= 1
        # デットスケジュール・返済予定表の類は主たる財務モデルではない
        if any(w in text for w in ("debt", "デット", "返済", "借入")):
            s -= 3
        return s

    return max(candidates, key=score)


def _docs_digest(documents: list[dict], excel_limit: int | None = None,
                 pdf_pages: int | None = None) -> str:
    """複数資料のダイジェストを結合する。同一内容のファイルは1回だけ展開する
    （同じファイルが複数スロットにアップロードされるケースの冗長・コスト増を防ぐ）。"""
    parts: list[str] = []
    seen: dict[str, str] = {}   # sha -> 最初に展開したファイル表記
    for doc in documents:
        p = Path(doc["stored_path"])
        header = f"--- ファイル: {doc['filename']}（スロット: {doc['slot']}）---"
        sha = _file_sha(p)
        if sha in seen:
            parts.append(header)
            parts.append(f"（内容は {seen[sha]} と完全に同一のため省略）")
            continue
        seen[sha] = f"『{doc['filename']}（スロット: {doc['slot']}）』"
        parts.append(header)
        if p.suffix == ".xlsx":
            digest = _excel_digest(p)
            if excel_limit and len(digest) > excel_limit:
                parts.append(digest[:excel_limit])
                parts.append(f"（注意: このファイルのダイジェストは全{len(digest):,}文字のうち"
                             f"先頭{excel_limit:,}文字のみを表示。以降は未読）")
            else:
                parts.append(digest)
        else:
            parts.append(_pdf_digest(p, max_pages=pdf_pages))
    return "\n".join(parts)


MAX_OUTPUT_TOKENS = 32000


class TruncatedOutputError(RuntimeError):
    """出力が最大長で打ち切られた（＝結果の取りこぼしが確定している）。

    「AIが検知しなかった」のと区別するため専用の例外にする。
    """


class AnthropicExtractor(Extractor):
    def __init__(self):
        import anthropic
        if not ANTHROPIC_API_KEY:
            raise RuntimeError("EXTRACTOR_MODE=anthropic には ANTHROPIC_API_KEY が必要です")
        self.client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
        # 直近のAI呼び出しの診断情報（呼び出し側が解析ログに記録する）
        self.last_call: dict | None = None

    def _call(self, prompt: str, schema: dict, tool_name: str) -> dict:
        # 根拠3点セット付きの抽出項目一覧は長くなる。8192では途中で切れて
        # tool入力のJSONが不完全になる（実APIで確認済み）ため大きめに取る。
        # 大きなmax_tokensはSDKがストリーミングを要求するためstreamで受ける
        import time
        started = time.monotonic()
        model = get_model()  # 設定UIから実行時に切替可能
        self.last_call = dict(mode="anthropic", model=model, prompt_chars=len(prompt),
                              max_tokens=MAX_OUTPUT_TOKENS)
        try:
            with self.client.messages.stream(
                model=model,
                max_tokens=MAX_OUTPUT_TOKENS,
                system=SYSTEM_PROMPT,
                tools=[dict(name=tool_name, description="構造化された抽出結果を返す",
                            input_schema=schema)],
                tool_choice=dict(type="tool", name=tool_name),
                messages=[dict(role="user", content=prompt)],
            ) as stream:
                resp = stream.get_final_message()
        except Exception as e:
            self.last_call.update(status="error", error=f"{type(e).__name__}: {e}",
                                  duration_ms=int((time.monotonic() - started) * 1000))
            raise
        usage = getattr(resp, "usage", None)
        self.last_call.update(
            stop_reason=resp.stop_reason,
            input_tokens=getattr(usage, "input_tokens", None),
            output_tokens=getattr(usage, "output_tokens", None),
            duration_ms=int((time.monotonic() - started) * 1000),
        )
        if resp.stop_reason == "max_tokens":
            self.last_call.update(status="truncated")
            raise TruncatedOutputError(
                f"AI応答が最大出力長（{MAX_OUTPUT_TOKENS:,}トークン）に達して打ち切られました。"
                "結果の一部が失われています（AIが検知しなかったのではありません）。"
                "資料の分量を減らすか、対象を絞って再実行してください")
        for block in resp.content:
            if block.type == "tool_use":
                self.last_call.update(status="ok")
                return block.input
        self.last_call.update(status="error", error="tool_use ブロックが返らなかった")
        raise RuntimeError("構造化出力が得られませんでした")

    def identify_document(self, filename: str, path: Path | None) -> dict:
        digest = ""
        if path and path.suffix == ".pdf":
            digest = _pdf_digest(path, max_pages=3)
        elif path:
            digest = _excel_digest(path)[:8000]
        prompt = (f"次のファイルの社名と資料種別を判定してください。\n"
                  f"ファイル名: {filename}\n先頭部分の内容:\n{digest[:8000]}\n"
                  "事業・財務等の複数領域が1冊にまとまった統合版DDは dd_integrated と判定すること。"
                  "いずれの種別にも当てはまらない資料は unknown とし、"
                  "資料の実際の性格と判定できない理由を detail に書くこと"
                  "（無理にいずれかの種別へ寄せない）。")
        return self._call(prompt, IDENTIFY_SCHEMA, "identify_document")

    def extract_deal_info(self, documents: list[dict]) -> dict:
        parts = [_docs_digest(documents, excel_limit=20000)]
        parts.append(
            "\n上記資料から案件の基本情報（案件名・案件種別・借入人SPC・対象会社・業種・"
            "スポンサー・クローズ予定日・EV・シニアローン総額・エクイティ・期間・"
            "スポンサー提示EBITDA・案件概要）を読み取ってください。"
            "各フィールドの出典（ファイル＋シート/セルまたはページ）を sources に必ず記載すること。"
            "借入人SPC名はDDレポートの買収ストラクチャー・買収後B/S関連の記述も含めて全文を探すこと。"
            "金額フィールド（ev_mm・senior_mm・equity_mm・sponsor_ebitda_mm）は"
            "必ず百万円単位に換算した数値を入れること（資料の単位が何であれ百万円に換算する。"
            "元の値・単位と換算の過程をsourcesに記載）。"
            "本行取組額・担当者・審査相談予定日は行内情報のため対象外。"
            "資料に無い項目は null にすること（推測しない）。")
        return self._call("\n".join(parts), DEAL_INFO_SCHEMA, "extract_deal_info")

    def extract_items(self, deal: dict, documents: list[dict]) -> dict:
        parts = [f"案件情報: {json.dumps(deal, ensure_ascii=False)}", "",
                 _docs_digest(documents)]
        parts.append(
            "\n上記資料から以下を抽出してください：\n"
            "1. 定性情報：事業要約・主要リスク（DDレポート群から。ページ番号を根拠に）\n"
            "2. 過去実績：資料にある全年度のPL（売上高・営業利益・EBITDA・当期純利益）、"
            "BS（現預金・純資産・有利子負債）、CF（FCF）。"
            "実績が5期あれば5期すべて抽出する（年度の取捨選択をしない）\n"
            "3. 予測数値：資料に存在するケースごとのFY計画値"
            "（売上・営業利益・EBITDA・当期純利益・現預金・純資産・FCF）。"
            "計画も資料にある全年度を抽出する（7期あれば7期すべて）。"
            "計画値は財務モデル（キャッシュ値・計算値）とDDレポートの計画表の両方を確認し、"
            "両方にあれば一致を検証（不一致はmismatchに記載）、片方のみならその出典を明記。"
            "ケースの数・名称は資料に従うこと（2つとは限らない）。"
            "判定根拠（シート・セル・プルダウン等の切替方法）を明記。"
            "独立したケースが1つしか存在しない場合は、存在するケースのみ抽出し、"
            "無いケースは values=null・source_type=missing で報告して理由をlogicに書く。"
            "ケースが3つ以上ある場合は、審査の基礎とすべき保守的なケースと"
            "スポンサー提示等の強気ケースを標準キーに割り当て、"
            "残りは case3_revenue 等のキーで追加し、割当の論理をlogicに書く。"
            "どのケースを審査の基礎とすべきかが資料から一意に決まらない場合は"
            "inquiries（category=正ファイル識別）で照会する\n"
            "4. ストラクチャー：のれん・EV・シニアローン。資料間で値が食い違う場合はmismatchに両方を記載\n"
            "5. 財務ハイライト（key=fin_highlights・section=財務ハイライト・text_value）："
            "抽出した実績・計画数値に基づき、利益率や資金フローの特徴的な変動とその要因"
            "（例：どの年度にマージンが変動したか、原価・販管費のどの費目が要因か）を、"
            "審査報告書にふさわしい文体で記述。"
            "「・観点：本文」形式の箇条書き3〜5項目（改行区切り）とし、数値は抽出値のみを引用すること\n"
            "6. ケース前提差異（key=case_assumptions・section=財務ハイライト・text_value）："
            "抽出した各ケースの前提（KPI・成長率・マージン）の違いと乖離の主因を、"
            "審査報告書にふさわしい文体で記述。"
            "「・観点：本文」形式の箇条書き3〜4項目（改行区切り）とし、ケース間の計画値の差を具体的数値で示すこと"
            "（ケースが1つしか無い場合はその旨を述べる）\n"
            "抽出の規律：\n"
            "・FCF等の導出指標も、まず資料の記載値（CF計算書・DDのCF表）を探して抽出する"
            "（source_type=extracted）。記載が無い場合のみ計算値・推定を用い、定義と算法をlogicに明記する\n"
            "・正常収益力EBITDA（normalized_ebitda）はDDの結論部・QoE・正常収益力関連の章まで必ず探すこと\n"
            "・同一の概念が資料ごとに異なる名称・略称・英語表記で書かれている場合は名寄せし、"
            "マッピングの論理をlogicに書く。"
            "確証の無い名寄せは inquiries（category=名寄せ）でも報告する\n"
            "・シート名・注記・記載内容から現行版ではないと判断されるシート"
            "（旧版・作業用・下書き・空シート等）は抽出対象外とし、"
            "検知したら inquiries（category=正シート識別）で報告する\n"
            "・#REF!等の参照切れを検知したら inquiries（category=参照切れ）で報告し、"
            "リンク先ファイルの提供を求める質問を書く\n"
            "・ダイジェストに『計算不能』と示された数式（システムが解釈できない関数・"
            "外部参照等）の値は、推定で埋めず source_type=missing とし、"
            "計画の主要数値に関わる場合は inquiries で照会する\n"
            "・単位ラベルの無いシート・表の数値は勝手に単位を仮定せず、"
            "inquiries（category=単位不明）で照会する（周辺資料から単位が一意に確定できる場合は"
            "その論理をlogicに書いた上で抽出してよい）\n"
            "・金額は元資料の単位のまま返し、unitに単位を正確に書く（換算しない）\n"
            "keyの命名規則：以下に該当する項目は必ずこの標準キーを使うこと（帳票出力が参照する）。"
            "実績: act_revenue／act_op（営業利益）／act_ebitda／act_ni（当期純利益）／"
            "act_cash（現預金）／act_net_assets（純資産）／act_debt（有利子負債）／act_fcf。"
            "審査の基礎とする保守的なケースの計画: base_revenue／base_op／base_ebitda／base_ni／"
            "base_cash／base_net_assets／base_fcf。"
            "スポンサー提示等の強気ケースの計画: sponsor_revenue／sponsor_op／sponsor_ebitda／"
            "sponsor_ni／sponsor_cash／sponsor_net_assets／sponsor_fcf。"
            "ストラクチャー: ev／senior_loan／goodwill。その他: normalized_ebitda（正常収益力EBITDA）。"
            "該当しない項目は内容がわかる英小文字スネークケースで命名。\n"
            "valuesの年度キーは『FY+西暦下2桁』（例: FY24, FY27）で統一すること。"
            "資料側の年度表記がどのような形式（決算期表記・和暦・実績/予想の別を示す記号等）でも"
            "この形式に正規化し、元の表記との対応と実績/計画の区分をlogicに記す。"
            "会計年度の始期が資料から一意に定まらない場合は、割当の前提をlogicに明記する。")
        return self._call("\n".join(parts), ITEMS_SCHEMA, "extract_items")

    def propose_kpi_tree(self, deal: dict, documents: list[dict]) -> dict:
        # 財務モデルはBase優先・無ければSponsor（単一ケース案件に対応）。
        # スロット未指定（種別なしで取り込んだ）Excelも候補に含める
        # ——ここを slot だけで探すと、種別なし投入時にモデルが1冊も渡らない
        model_doc = _pick_model_doc(documents)
        other_models = [d for d in _excel_docs(documents) if d is not model_doc]
        # 事業系DDは分冊（dd_business）と統合版（dd_integrated）の両方に対応
        dd_docs = [d for d in documents if d["slot"] in ("dd_business", "dd_integrated")]
        parts = []
        if model_doc:
            parts.append(f"--- 財務モデル: {model_doc['filename']} ---")
            parts.append(_excel_digest(Path(model_doc["stored_path"])))
            if other_models:
                names = "・".join(d["filename"] for d in other_models)
                parts.append(f"（注記: 本ステップでは上記1冊を財務モデルとして用いた。"
                             f"他に {names} が提出されている。"
                             f"上記が主たる財務モデルでないと判断される場合は、その旨を"
                             f"ツリーの説明に明記すること）")
        seen: set[str] = set()
        for doc in dd_docs:
            sha = _file_sha(Path(doc["stored_path"]))
            if sha in seen:
                continue
            seen.add(sha)
            parts.append(f"--- 事業系DD: {doc['filename']} ---")
            parts.append(_pdf_digest(Path(doc["stored_path"])))
        parts.append(
            "\n財務モデルの計画年の数式を構文解析し（再計算はしない）、変数の連動関係を"
            "ツリー構造（nodes）にしてください。\n"
            "ツリーの構成（必ずこの形にする）：\n"
            "1. 最上位（parent=null）は返済能力に直結する集約指標を1つ置く"
            "（通常はEBITDA。モデルにEBITDA行があればそれを根にする）\n"
            "2. その下に営業利益 → 売上高・売上原価・販管費（各合計）を置き、"
            "さらにモデルに存在する内訳科目の行すべてへ分解する\n"
            "3. 各内訳科目の下に、算式に現れる非財務の変数（数量・比率・単価・時間・率など、"
            "対象会社の事業を動かすドライバー）まで分解する\n"
            "4. 減価償却費のようにEBITDAへ直接加算される項目も根の直下に置く\n"
            "重要な制約：\n"
            "・すべてのノードの parent は、必ず同じ nodes 配列内に存在する id を指すこと"
            "（存在しないidを親に指定すると画面に表示されなくなる）。根は parent=null とする\n"
            "・モデルの三表（PL）に存在する損益科目を取りこぼさないこと。"
            "ツリーは上から下まで連結された1本の木にする\n"
            "事業DDの定性情報と照合し、リスクドライバーとして重要なKPIに star=true を付けてください。"
            "各ノードの evidence には参照シート・行と、数式のどの部分から親子関係を"
            "判定したかを書いてください。")
        return self._call("\n".join(parts), KPI_TREE_SCHEMA, "propose_kpi_tree")

    def propose_scenarios(self, deal: dict, documents: list[dict]) -> list[dict]:
        parts = [f"案件情報: {json.dumps(deal, ensure_ascii=False)}",
                 _docs_digest(documents, excel_limit=20000)]
        parts.append(
            "\n3類型（トップライン悪化／コスト上昇／イベント）のストレスシナリオを"
            "標準5部構成で生成してください（key: A/B/C、origin: ai）。"
            "インパクトはDSCR・レバレッジ・現預金への影響を具体的数値を含めて定性推定しますが、"
            "モデルの再計算はせず、判定ラベルは出さないこと。"
            "変化幅には外部データや資料内の根拠を紐付けること。"
            "impact_calcには、impactに書いた最終指標（DSCR等）を左辺に置き、"
            "抽出済みの数値・ストレス対象KPIを使って before/after の値入り計算式で分解すること"
            "（beforeとafterで式の構造を揃え、変化する値のみ {{}} で囲む）。")
        result = self._call("\n".join(parts), SCENARIOS_SCHEMA, "propose_scenarios")
        return result["cards"]

    def chat(self, context: str, message: str, state: dict) -> dict:
        prompt = (
            f"コンテキスト: {context}（kpi=KPIツリーの修正 / scenario=シナリオカードの修正）\n"
            f"現在の状態: {json.dumps(state, ensure_ascii=False)}\n"
            f"ユーザーの指示: {message}\n\n"
            "指示に対応する変更差分（diff）を返してください。diffの形式:\n"
            "- ノード追加: {type:'add_node', node:{id,parent,label,origin,star,badge,evidence,...}}\n"
            "- ★変更: {type:'star_change', remove:[nodeId], add:[nodeId]}\n"
            "- カード追加: {type:'add_card', card:{key,origin,title,...標準5部構成}}\n"
            "- カード修正: {type:'update_card', card_key:'A', fields:{変更するフィールドのみ}}\n"
            "変更を直接適用してはいけません。人が差分プレビューを確認して適用します。"
            "対応できない指示の場合は diff=null で、replyに理由を書いてください。")
        return self._call(prompt, CHAT_SCHEMA, "chat_diff")

    def chat_suggestions(self, context: str) -> list[str]:
        if context == "kpi":
            return ["（例）重要なKPIを追加・変更したい内容を入力してください"]
        return ["（例）シナリオの深掘り・保全策の追加などを入力してください"]
