"""案件（S1/S2）と資料アップロード・AI解析のAPI。"""
import json
import shutil
from datetime import datetime
from pathlib import Path

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from fastapi.responses import FileResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session

from ..config import UPLOAD_DIR
from ..database import get_db
from ..extractors import get_extractor
from ..extractors.base import UnknownSampleFileError
from ..models import (DOCUMENT_SLOTS, MULTI_SLOTS, Deal, Document, ExtractedItem,
                      HistoryEvent, Inquiry, KpiNode, Scenario, User)
from ..services.text_clean import clean_obj, clean_text
from ..services.units import normalize_values

router = APIRouter(prefix="/api", tags=["deals"])


def add_history(db: Session, deal: Deal, user: str | None, action: str, detail: str = ""):
    db.add(HistoryEvent(deal_id=deal.id, at=datetime.now(), user_key=user,
                        action=action, detail=detail))
    deal.updated_at = datetime.now()


@router.get("/users")
def list_users(db: Session = Depends(get_db)):
    return [u.to_dict() for u in db.query(User).all()]


@router.get("/deals")
def list_deals(review_status: str | None = None, work_status: str | None = None,
               db: Session = Depends(get_db)):
    deals = db.query(Deal).order_by(Deal.updated_at.desc()).all()
    out = [d.to_list_dict() for d in deals]
    if review_status:
        out = [d for d in out if d["review_status"] == review_status]
    if work_status:
        out = [d for d in out if d["work_status"] == work_status]
    counts = {}
    for d in deals:
        counts[d.review_status] = counts.get(d.review_status, 0) + 1
    return dict(deals=out, counts=counts, total=len(deals))


class DealCreate(BaseModel):
    name: str
    deal_type: str
    borrower: str
    target: str
    industry: str | None = None
    sponsor: str | None = None
    close_date: str | None = None
    next_meeting_date: str | None = None
    ev_mm: int | None = None
    senior_mm: int | None = None
    our_commitment_mm: int | None = None
    equity_mm: int | None = None
    tenor_years: int | None = None
    sponsor_ebitda_mm: int | None = None
    summary: str | None = None
    user: str | None = None


@router.post("/deals")
def create_deal(body: DealCreate, db: Session = Depends(get_db)):
    deal = Deal(**body.model_dump(exclude={"user"}), owner=body.user,
                review_status="検討中")
    db.add(deal)
    db.flush()
    add_history(db, deal, body.user, "案件登録", "基本情報を登録")
    db.commit()
    return deal.to_dict()


class DraftBody(BaseModel):
    user: str | None = None


@router.post("/deals/draft")
def create_draft(body: DraftBody, db: Session = Depends(get_db)):
    """資料アップロード起点の登録フロー用の下書き案件。"""
    deal = Deal(name="（下書き）新規案件", deal_type="LBO", borrower="", target="",
                owner=body.user, review_status="検討中")
    db.add(deal)
    db.flush()
    add_history(db, deal, body.user, "案件登録（下書き）", "資料アップロードを開始")
    db.commit()
    return deal.to_dict()


class DealPatch(BaseModel):
    name: str | None = None
    deal_type: str | None = None
    borrower: str | None = None
    target: str | None = None
    industry: str | None = None
    sponsor: str | None = None
    close_date: str | None = None
    next_meeting_date: str | None = None
    ev_mm: int | None = None
    senior_mm: int | None = None
    our_commitment_mm: int | None = None
    equity_mm: int | None = None
    tenor_years: int | None = None
    sponsor_ebitda_mm: int | None = None
    summary: str | None = None
    owner: str | None = None
    user: str | None = None


@router.patch("/deals/{deal_id}")
def update_deal(deal_id: int, body: DealPatch, db: Session = Depends(get_db)):
    deal = _get_deal(db, deal_id)
    data = body.model_dump(exclude_unset=True, exclude={"user"})
    for k, v in data.items():
        setattr(deal, k, v)
    add_history(db, deal, body.user, "案件情報を更新",
                "登録内容を確定" if "name" in data else "")
    db.commit()
    return deal.to_dict()


@router.delete("/deals/{deal_id}")
def delete_deal(deal_id: int, db: Session = Depends(get_db)):
    deal = _get_deal(db, deal_id)
    db.delete(deal)
    db.commit()
    return dict(deleted=deal_id)


@router.post("/deals/{deal_id}/extract-info")
def extract_deal_info(deal_id: int, user: str | None = None,
                      db: Session = Depends(get_db)):
    """アップロード済み資料から案件基本情報をAIで読み取る（フォームへの自動入力用）。"""
    deal = _get_deal(db, deal_id)
    if not deal.documents:
        raise HTTPException(400, "資料がアップロードされていません")
    docs = [dict(filename=d.filename, slot=d.slot, stored_path=d.stored_path)
            for d in deal.documents]
    result = get_extractor().extract_deal_info(docs)
    # 画面に出る説明文（note・出典）から内部表現を除去する
    if isinstance(result, dict):
        if isinstance(result.get("note"), str):
            result["note"] = clean_text(result["note"])
        if isinstance(result.get("sources"), dict):
            result["sources"] = {k: clean_text(v) if isinstance(v, str) else v
                                 for k, v in result["sources"].items()}
    add_history(db, deal, user, "案件情報の自動読み取り",
                f"{len(result.get('fields', {}))}フィールドを資料から抽出")
    db.commit()
    return result


def _get_deal(db: Session, deal_id: int) -> Deal:
    deal = db.get(Deal, deal_id)
    if not deal:
        raise HTTPException(404, "案件が見つかりません")
    return deal


@router.get("/deals/{deal_id}/full")
def deal_full(deal_id: int, db: Session = Depends(get_db)):
    deal = _get_deal(db, deal_id)
    extractor = get_extractor()
    findings = []
    for m in deal.memos:
        findings.extend(f.to_dict() for f in m.findings)
    return dict(
        deal=deal.to_dict(),
        documents=[d.to_dict() for d in deal.documents],
        items=[i.to_dict() for i in deal.items],
        kpi_nodes=[n.to_dict() for n in deal.kpi_nodes],
        scenarios=[s.to_dict() for s in deal.scenarios],
        memos=[m.to_dict() for m in reversed(deal.memos)],
        history=[h.to_dict() for h in reversed(deal.history)],
        exports=[e.to_dict() for e in deal.exports],
        findings=findings,
        inquiries=[q.to_dict() for q in deal.inquiries],
        chat_suggestions=dict(kpi=extractor.chat_suggestions("kpi"),
                              scenario=extractor.chat_suggestions("scenario")),
    )


@router.get("/deals/{deal_id}/documents/{doc_id}/file")
def document_file(deal_id: int, doc_id: int, db: Session = Depends(get_db)):
    """アップロード済み資料の実ファイルを返す（根拠パネルの参照元リンク用）。

    PDFはブラウザ内で開く（inline・#page=N アンカー対応）。Excelはダウンロード。
    """
    deal = _get_deal(db, deal_id)
    doc = next((d for d in deal.documents if d.id == doc_id), None)
    if not doc:
        raise HTTPException(404, "資料が見つかりません")
    # シードデータ等では stored_path が無い（実ファイル未添付）ことがある
    if not doc.stored_path:
        raise HTTPException(404, "この資料は実ファイルが保存されていません（デモ用メタデータのみ）")
    path = Path(doc.stored_path)
    if not path.exists():
        raise HTTPException(404, "ファイルが移動または削除されています")
    is_pdf = path.suffix.lower() == ".pdf"
    return FileResponse(
        path,
        media_type="application/pdf" if is_pdf
        else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=doc.filename,
        content_disposition_type="inline" if is_pdf else "attachment",
    )


@router.get("/deals/{deal_id}/documents/{doc_id}/peek")
def document_peek(deal_id: int, doc_id: int, location: str,
                  db: Session = Depends(get_db)):
    """根拠のExcelセル周辺を抜粋して返す（パネル内で中身を確認できるようにする）。

    Excelはブラウザで開けないため、location（例「PL!G9」「Assumptions!D21」）を解釈し、
    該当セルを中心に前後の行・列を値と数式つきで返す。
    """
    import re

    from openpyxl import load_workbook
    from openpyxl.utils import column_index_from_string, get_column_letter

    from ..services.xl_eval import XlEvaluator

    deal = _get_deal(db, deal_id)
    doc = next((d for d in deal.documents if d.id == doc_id), None)
    if not doc or not doc.stored_path:
        raise HTTPException(404, "資料が見つかりません")
    path = Path(doc.stored_path)
    if not path.exists() or path.suffix.lower() != ".xlsx":
        raise HTTPException(400, "この資料はExcelの抜粋に対応していません")

    # 「PL!G9」「'月次 試算表'!C7」「Assumptions!D21」「PLシート C9:E9」形式に対応する
    loc = (location or "").replace("'", "").replace('"', "")
    m = re.search(r"([^\s!、（）]+)\s*!\s*\$?([A-Z]{1,2})\$?(\d+)", loc) \
        or re.search(r"([^\s!、（）]+?)\s*シート\s*\$?([A-Z]{1,2})\$?(\d+)", loc)
    if not m:
        raise HTTPException(400, "参照箇所（シート・セル）を特定できませんでした")
    sheet_name, col_s, row_s = m.group(1), m.group(2), int(m.group(3))

    wb = load_workbook(path)
    if sheet_name not in wb.sheetnames:
        # 表記ゆれ（「PLシート」等）を救済する。ただし短いシート名（PL）が
        # 別シート（PL_old）の指定に誤ってマッチしないよう、長い候補から順に見る
        by_len = sorted(wb.sheetnames, key=len, reverse=True)
        cand = (next((s for s in by_len if sheet_name.startswith(s)), None)
                or next((s for s in by_len if s in sheet_name), None))
        if not cand:
            raise HTTPException(404, f"シートが見つかりません: {sheet_name}")
        sheet_name = cand
    ws = wb[sheet_name]
    wbv = load_workbook(path, data_only=True)
    wsv = wbv[sheet_name]
    ev = XlEvaluator(wb)

    # Excelの上限（1,048,576行 / 16,384列＝XFD）を超える指定でクラッシュしないよう丸める
    MAX_ROW, MAX_COL = 1_048_576, 16_384
    col = min(column_index_from_string(col_s), MAX_COL)
    row_s = min(max(row_s, 1), MAX_ROW)
    # ヘッダー行（年度行）＋対象セル周辺を抜粋する
    row_from, row_to = max(1, row_s - 3), min(row_s + 3, MAX_ROW)
    col_from, col_to = max(1, col - 3), min(col + 3, MAX_COL)
    header_rows = [r for r in (4, 5) if r < row_from]

    def cell_out(r: int, c: int) -> dict:
        cell = ws.cell(row=r, column=c)
        raw = cell.value
        display = raw
        formula = None
        if isinstance(raw, str) and raw.startswith("="):
            formula = raw
            cached = wsv.cell(row=r, column=c).value
            display = cached if cached is not None else ev.value(sheet_name, cell.coordinate)
        return dict(ref=cell.coordinate, value=display, formula=formula,
                    target=(r == row_s and c == col))

    cols = list(range(col_from, col_to + 1))
    rows = []
    for r in header_rows + list(range(row_from, row_to + 1)):
        # B列（行ラベル）は常に添えて、どの科目の行かが分かるようにする
        label = ws.cell(row=r, column=2).value
        cells = [cell_out(r, c) for c in cols]
        if any(c["value"] is not None for c in cells) or label:
            rows.append(dict(row=r, label=label, cells=cells))
    return dict(filename=doc.filename, sheet=sheet_name,
                target=f"{col_s}{row_s}",
                columns=[get_column_letter(c) for c in cols], rows=rows)


@router.post("/deals/{deal_id}/documents")
def upload_document(deal_id: int, slot: str = Form("unclassified"), user: str = Form(None),
                    file: UploadFile = File(...), db: Session = Depends(get_db)):
    """資料を1件アップロードする。

    slot は複数指定（カンマ区切り）・未指定のいずれも可（仕様②4）。
    - 複数指定：同じファイルを各スロットに登録する（統合版DDが事業・財務を兼ねる等）
    - 未指定：unclassified として保持し、解析対象には含める
    """
    deal = _get_deal(db, deal_id)
    slots = [s.strip() for s in (slot or "").split(",") if s.strip()] or ["unclassified"]
    unknown = [s for s in slots if s not in DOCUMENT_SLOTS]
    if unknown:
        raise HTTPException(400, f"不明な資料種別です: {'、'.join(unknown)}")

    dest_dir = UPLOAD_DIR / str(deal_id)
    dest_dir.mkdir(parents=True, exist_ok=True)
    dest = dest_dir / file.filename
    with dest.open("wb") as f:
        shutil.copyfileobj(file.file, f)

    try:
        info = get_extractor().identify_document(file.filename, dest)
    except UnknownSampleFileError as e:
        dest.unlink(missing_ok=True)
        raise HTTPException(400, e.message)

    docs = []
    for s in slots:
        # 単一スロットは上書き（種別未設定は複数ファイルを並べて保持する）
        if s not in MULTI_SLOTS:
            for d in list(deal.documents):
                if d.slot == s:
                    db.delete(d)
        elif any(d.slot == s and d.filename == file.filename for d in deal.documents):
            continue  # 同じファイルの重複登録は無視
        doc = Document(deal_id=deal.id, slot=s, filename=file.filename,
                       stored_path=str(dest), status="uploaded",
                       identified_company=info.get("company"),
                       identified_label=info.get("label"),
                       identified_detail=info.get("detail"))
        db.add(doc)
        docs.append(doc)
    db.flush()
    labels = "、".join(DOCUMENT_SLOTS.get(s, s) for s in slots)
    add_history(db, deal, user, "資料アップロード", f"{file.filename}（{labels}）")
    db.commit()
    result = docs[0].to_dict() if docs else dict(filename=file.filename, slot=slots[0])
    result["slots"] = slots
    # 案件照合（社名の一致チェック）。対象会社が未設定（下書き）の場合は判定しない
    company = info.get("company") or ""
    if company and deal.target:
        result["company_match"] = company in deal.target or deal.target in company
    else:
        result["company_match"] = None
    return result


@router.post("/deals/{deal_id}/analyze")
def analyze(deal_id: int, user: str | None = None, db: Session = Depends(get_db)):
    """AI解析の実行：抽出項目・KPIツリー提案・AI推奨シナリオを一括生成する。"""
    deal = _get_deal(db, deal_id)
    if not deal.documents:
        raise HTTPException(400, "資料がアップロードされていません")
    extractor = get_extractor()
    deal_info = dict(name=deal.name, deal_type=deal.deal_type, borrower=deal.borrower,
                     target=deal.target, sponsor=deal.sponsor,
                     ev_mm=deal.ev_mm, senior_mm=deal.senior_mm)
    docs = [dict(filename=d.filename, slot=d.slot, stored_path=d.stored_path)
            for d in deal.documents]

    extraction = extractor.extract_items(deal_info, docs)
    # 新形式は {items, inquiries}、モック等の旧形式は items のリストのみ
    if isinstance(extraction, dict):
        items = extraction.get("items", [])
        inquiries = extraction.get("inquiries", [])
    else:
        items, inquiries = extraction, []
    tree = extractor.propose_kpi_tree(deal_info, docs)
    cards = extractor.propose_scenarios(deal_info, docs)
    # 人が読む文章から内部表現（values=null 等）を除去する
    items = clean_obj(items)
    inquiries = clean_obj(inquiries)
    tree = clean_obj(tree)
    cards = clean_obj(cards)

    # 再解析時は既存の提案をリセット（確定済み案件の再解析はデモでは想定しない）
    for coll in (deal.items, deal.kpi_nodes, deal.scenarios, deal.inquiries):
        for row in list(coll):
            db.delete(row)
    db.flush()

    for idx, it in enumerate(items):
        # 単位の正規化はシステム側の責務（仕様 ③4）。金額は百万円へ換算し元単位を残す
        values, unit, source_unit = normalize_values(it.get("values"), it.get("unit", "百万円"))
        db.add(ExtractedItem(
            deal_id=deal.id, key=it["key"], section=it["section"], label=it["label"],
            unit=unit or "百万円",
            source_type=it.get("source_type", "extracted"), source_unit=source_unit,
            case_name=it.get("case"),
            values_json=json.dumps(values, ensure_ascii=False) if values else None,
            text_value=it.get("text_value"), required=it.get("required", True),
            evidence_json=json.dumps(it.get("evidence"), ensure_ascii=False),
            mismatch_json=json.dumps(it.get("mismatch"), ensure_ascii=False) if it.get("mismatch") else None,
            status="proposed", order_index=idx))
    severity_order = {"high": 0, "medium": 1, "low": 2}
    for idx, q in enumerate(sorted(inquiries,
                                   key=lambda x: severity_order.get(x.get("severity"), 1))):
        db.add(Inquiry(
            deal_id=deal.id, category=q.get("category", "その他"),
            severity=q.get("severity", "medium"), title=q.get("title", "確認事項"),
            detail=q.get("detail"),
            source_json=json.dumps(q.get("source"), ensure_ascii=False) if q.get("source") else None,
            suggested_question=q.get("suggested_question"),
            status="open", order_index=idx))
    for idx, n in enumerate(tree["nodes"]):
        db.add(KpiNode(
            deal_id=deal.id, node_id=n["id"], parent_id=n.get("parent"),
            label=n["label"], origin=n.get("origin", "model"), star=n.get("star", False),
            formula=n.get("formula"), value_text=n.get("value_text"), badge=n.get("badge"),
            evidence_json=json.dumps(n.get("evidence"), ensure_ascii=False),
            order_index=idx))
    for idx, c in enumerate(cards):
        db.add(Scenario(
            deal_id=deal.id, key=c["key"], origin=c.get("origin", "ai"),
            type_label=c.get("type_label"), title=c["title"], cause=c.get("cause"),
            affected_kpis_json=json.dumps(c.get("affected_kpis", []), ensure_ascii=False),
            change_text=c.get("change_text") or c.get("change"),
            change_basis=c.get("change_basis"),
            impact=c.get("impact"),
            impact_calc_json=json.dumps(c.get("impact_calc"), ensure_ascii=False)
            if c.get("impact_calc") else None,
            safeguards=c.get("safeguards"),
            questions=c.get("questions"), adopted=False, order_index=idx))
    deal.kpi_status = "proposed"
    for d in deal.documents:
        d.status = "analyzed"
    add_history(db, deal, user, "AI解析完了",
                f"{len(items)}項目を抽出・確認事項{len(inquiries)}件・"
                f"KPIツリー{len(tree['nodes'])}ノード・推奨シナリオ{len(cards)}件を提案")
    db.commit()
    return dict(items=len(items), inquiries=len(inquiries),
                kpi_nodes=len(tree["nodes"]), scenarios=len(cards))


class InquiryPatch(BaseModel):
    status: str | None = None            # open / resolved
    resolution_note: str | None = None
    user: str | None = None


@router.patch("/deals/{deal_id}/inquiries/{inquiry_id}")
def update_inquiry(deal_id: int, inquiry_id: int, body: InquiryPatch,
                   db: Session = Depends(get_db)):
    """確認事項（照会）の状態更新（確認済みチェック・メモ）。"""
    deal = _get_deal(db, deal_id)
    inquiry = next((q for q in deal.inquiries if q.id == inquiry_id), None)
    if not inquiry:
        raise HTTPException(404, "確認事項が見つかりません")
    if body.status is not None:
        if body.status not in ("open", "resolved"):
            raise HTTPException(400, f"不明なステータスです: {body.status}")
        inquiry.status = body.status
        if body.status == "resolved":
            inquiry.resolved_by = body.user
            inquiry.resolved_at = datetime.now()
        else:
            inquiry.resolved_by = None
            inquiry.resolved_at = None
    if body.resolution_note is not None:
        inquiry.resolution_note = body.resolution_note
    add_history(db, deal, body.user,
                "確認事項を更新",
                f"「{inquiry.title}」を{'確認済み' if inquiry.status == 'resolved' else '未確認に戻す'}")
    db.commit()
    return inquiry.to_dict()
